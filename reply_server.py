from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Body, Query, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    RedirectResponse,
    JSONResponse,
    Response,
    StreamingResponse,
)
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from pydantic import BaseModel, Field, field_validator
from typing import Awaitable, Callable, List, Tuple, Optional, Dict, Any, Literal
from pathlib import Path
from urllib.parse import quote, unquote, urljoin, urlsplit
import hashlib
import ipaddress
import secrets
import socket
import tempfile
import time
import json
import os
import re
import uvicorn
import pandas as pd
import io
import csv
import asyncio
import importlib.util
import sqlite3
import threading
from collections import OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime, timedelta
import cookie_manager
from db_manager import (
    CARD_STOCK_IMPORT_MAX_ITEMS,
    CARD_STOCK_ITEM_MAX_BYTES,
    FULFILLMENT_API_PROTOCOL,
    AccountIdentityMismatchError,
    db_manager,
    mask_secret_preview,
)
from file_log_collector import setup_file_logging, get_file_log_collector
from ai_reply_engine import ai_reply_engine
from ai_provider_service import (
    PROVIDER_PRESETS,
    discover_provider_models,
    normalize_provider_models,
    provider_test_tokens,
    test_provider_reply,
)
from settings_service import (
    SETTINGS_SECTION_KEYS,
    apply_secret_action,
    normalize_system_settings,
    resolve_user_basic_settings,
)
from account_session_refresh import (
    active_refresh_registry,
    is_runtime_event_active,
    is_valid_account_login_username,
    login_method_label,
    normalize_login_method,
    official_login_error_message,
    password_refresh_requires_manual_reauth,
    reauth_action_for,
    reauth_message_for,
    remove_verification_image,
)
from utils.qr_login import qr_login_manager
from utils.browser_interaction import (
    InteractionQueueFull,
    InteractionRateLimited,
    InteractionUnavailable,
    InteractionValidationError,
    StaleFrameRevision,
)
from utils.xianyu_utils import trans_cookies
from utils.image_utils import image_manager
from utils.outbound_http import (
    OutboundRequestError,
    parse_public_http_url,
    request_public_http_sync,
)
from utils.verification_images import (
    latest_private_verification_image,
    list_private_verification_images,
    resolve_private_verification_image,
)
from order_sync_service import (
    OrderSyncCoordinator,
    SYNC_COVERAGE_FIELDS,
    DASHBOARD_ANALYTICS_STATUSES,
    VALID_ORDER_STATUSES,
    XianyuOrderListClient,
    new_order_sync_summary,
    normalize_order_status,
    parse_amount_fen,
    parse_order_time_utc,
    parse_trusted_order_quantity,
)
from api_routers import (
    accounts_router,
    admin_router,
    ai_router,
    auth_router,
    content_router,
    frontend_router,
    include_domain_routers,
    orders_router,
    settings_router,
    system_router,
)
from session_registry import (
    get_session_registry,
    sanitize_log_record,
    sanitize_runtime_error,
)
from official_login_sessions import OfficialLoginSessionCoordinator, OfficialLoginSessionRecord
from utils.xianyu_official_login import OfficialLoginResult, XianyuOfficialLoginService
from browser_extension_pairing import (
    MAX_COOKIE_COUNT,
    MAX_USER_AGENT_LENGTH,
    PAIRING_PROTOCOL_VERSION,
    PUBLIC_CONSOLE_ORIGIN,
    PUBLIC_IMPORT_URL,
    PairingError,
    browser_extension_pairings,
    is_loopback_host,
    normalize_structured_cookies,
)
from client_browser_login import (
    ClientBrowserError,
    client_login_sessions,
    device_challenges,
    normalize_device_id,
)
from utils.xianyu_session_probe import (
    PROBE_RETRYABLE_ERROR,
    cookies_to_string as session_cookies_to_string,
    detect_default_browser_user_agent,
    has_core_session_cookies,
    parse_cookie_string,
    probe_message_session_async,
)
from auth_registration_service import (
    RegistrationError,
    mask_email_for_log,
    normalize_email,
    resolve_client_ip,
)
from auth_email_service import (
    SMTP_CONFIGURATION_KEYS,
    SMTPConfigurationError,
    SMTPDeliveryError,
    SMTPEmailSender,
    registration_readiness,
    smtp_configuration_fingerprint,
    smtp_configuration_status,
)

from loguru import logger

logger.configure(patcher=sanitize_log_record)

# 简单的用户认证配置
ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"  # 系统初始化时的默认密码
SESSION_TOKENS = {}  # 存储会话token: {token: {'user_id': int, 'username': str, 'timestamp': float, 'expires_at': float}}
TOKEN_EXPIRE_TIME = 30 * 24 * 60 * 60  # token过期时间：30天
USER_BACKUP_MAX_BYTES = 25 * 1024 * 1024
IMAGE_UPLOAD_MAX_BYTES = 5 * 1024 * 1024
UPLOAD_READ_CHUNK_BYTES = 64 * 1024


async def _read_upload_with_limit(
    upload: UploadFile,
    *,
    max_bytes: int,
    label: str,
) -> bytes:
    """Read an upload incrementally and reject at the first over-limit byte."""
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await upload.read(
            min(UPLOAD_READ_CHUNK_BYTES, max_bytes + 1 - total)
        )
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"{label}超过{max_bytes // (1024 * 1024)}MB",
            )
        chunks.append(chunk)
    return b"".join(chunks)

# HTTP Bearer认证
security = HTTPBearer(auto_error=False)

AI_MODEL_NAME_MAX_LENGTH = 200
AI_MESSAGE_MAX_LENGTH = 4000
AI_ITEM_ID_MAX_LENGTH = 160
AI_ITEM_TITLE_MAX_LENGTH = 500
AI_ITEM_DESCRIPTION_MAX_LENGTH = 8000
AI_PROMPT_OVERRIDE_MAX_LENGTH = 8000
AI_SESSION_ID_MAX_LENGTH = 128
AI_TRAINING_RULE_MAX_COUNT = 50
AI_TRAINING_RULE_MAX_LENGTH = 2000
AI_TRAINING_RULE_SERIALIZED_MAX_BYTES = 4096
AI_LAB_MAX_SESSIONS_PER_USER = 8
AI_LAB_MAX_SESSIONS_GLOBAL = 128
AI_LAB_SESSION_TTL_SECONDS = 6 * 3600
AI_INTERACTIVE_GLOBAL_LIMIT = 4
AI_INTERACTIVE_PER_USER_LIMIT = 2
AI_INTERACTIVE_TIMEOUT_SECONDS = 45.0

ai_reply_lab_sessions: Dict[str, Dict[str, Any]] = {}
_ai_lab_sessions_lock = threading.Lock()
_ai_interactive_executor = ThreadPoolExecutor(
    max_workers=AI_INTERACTIVE_GLOBAL_LIMIT,
    thread_name_prefix="ai-interactive",
)
_ai_interactive_global_slots = threading.BoundedSemaphore(
    AI_INTERACTIVE_GLOBAL_LIMIT
)
_ai_interactive_user_slots: Dict[int, threading.BoundedSemaphore] = {}
_ai_interactive_user_slots_lock = threading.Lock()


def _ai_log_reference(value: Any, label: str) -> str:
    digest = hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()[:10]
    return f"{label}_{digest}"


def _get_ai_user_slot(user_id: int) -> threading.BoundedSemaphore:
    normalized_user_id = int(user_id)
    with _ai_interactive_user_slots_lock:
        return _ai_interactive_user_slots.setdefault(
            normalized_user_id,
            threading.BoundedSemaphore(AI_INTERACTIVE_PER_USER_LIMIT),
        )


def _run_bounded_ai_call(
    user_id: int,
    work: Callable[[], Any],
) -> Any:
    """Run expensive interactive AI work behind bounded global/user slots."""
    if not _ai_interactive_global_slots.acquire(blocking=False):
        raise HTTPException(status_code=429, detail="AI 请求繁忙，请稍后重试")
    user_slot = _get_ai_user_slot(user_id)
    if not user_slot.acquire(blocking=False):
        _ai_interactive_global_slots.release()
        raise HTTPException(status_code=429, detail="当前用户 AI 请求繁忙，请稍后重试")
    try:
        future = _ai_interactive_executor.submit(work)
    except Exception:
        user_slot.release()
        _ai_interactive_global_slots.release()
        raise

    def release_slots(_future: Any) -> None:
        user_slot.release()
        _ai_interactive_global_slots.release()

    future.add_done_callback(release_slots)
    try:
        return future.result(timeout=AI_INTERACTIVE_TIMEOUT_SECONDS)
    except FuturesTimeoutError as exc:
        raise HTTPException(status_code=504, detail="AI 请求超时，请稍后重试") from exc


def _prune_ai_lab_sessions(current_time: float, *, user_id: int) -> None:
    """Bound process-local training sessions by expiry, user and total size."""
    normalized_user_id = int(user_id)
    with _ai_lab_sessions_lock:
        expired = [
            session_id
            for session_id, session in ai_reply_lab_sessions.items()
            if current_time - float(session.get("timestamp") or 0)
            > AI_LAB_SESSION_TTL_SECONDS
        ]
        for session_id in expired:
            ai_reply_lab_sessions.pop(session_id, None)

        owned = sorted(
            (
                (session_id, float(session.get("timestamp") or 0))
                for session_id, session in ai_reply_lab_sessions.items()
                if int(session.get("user_id") or -1) == normalized_user_id
            ),
            key=lambda row: row[1],
        )
        for session_id, _timestamp in owned[:-AI_LAB_MAX_SESSIONS_PER_USER]:
            ai_reply_lab_sessions.pop(session_id, None)

        all_sessions = sorted(
            (
                (session_id, float(session.get("timestamp") or 0))
                for session_id, session in ai_reply_lab_sessions.items()
            ),
            key=lambda row: row[1],
        )
        for session_id, _timestamp in all_sessions[:-AI_LAB_MAX_SESSIONS_GLOBAL]:
            ai_reply_lab_sessions.pop(session_id, None)

# Direct API QR sessions are owner-scoped and persisted at most once.
qr_check_locks = defaultdict(lambda: asyncio.Lock())
qr_check_processed: Dict[str, Dict[str, Any]] = {}

# 不再需要单独的密码初始化，由数据库初始化时处理


def cleanup_qr_check_records() -> None:
    """Drop completed QR bookkeeping after one hour."""
    cutoff = time.time() - 3600
    for session_id, record in list(qr_check_processed.items()):
        if float(record.get("timestamp") or 0) >= cutoff:
            continue
        qr_check_processed.pop(session_id, None)
        qr_check_locks.pop(session_id, None)


# 认证相关模型
class LoginRequest(BaseModel):
    identifier: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    verification_code: Optional[str] = None


class LoginResponse(BaseModel):
    success: bool
    token: Optional[str] = None
    message: str
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class OrderSyncRequest(BaseModel):
    cookie_id: Optional[str] = None
    days: int = Field(90, ge=1, le=365)


class ItemMetricSyncRequest(BaseModel):
    cookie_id: str = Field(min_length=1, max_length=128)


class RegisterRequest(BaseModel):
    invite_code: str = ""
    username: str
    email: str
    password: str
    challenge_id: str
    verification_code: str
    terms_version: str
    terms_accepted: bool


class RegisterResponse(BaseModel):
    success: bool
    message: str
    token: Optional[str] = None
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None


class SendCodeRequest(BaseModel):
    email: str
    session_id: Optional[str] = None
    type: Optional[str] = 'register'  # 'register' 或 'login'


class SendCodeResponse(BaseModel):
    success: bool
    message: str


class EmailCodeRequest(BaseModel):
    purpose: Literal["register", "password_reset"]
    email: str
    invite_code: str = ""
    captcha_challenge_id: str
    captcha_code: str


class PasswordResetRequest(BaseModel):
    email: str
    new_password: str
    reset_grant_id: Optional[str] = None
    reset_grant_token: Optional[str] = None
    challenge_id: Optional[str] = None
    verification_code: Optional[str] = None


class PasswordResetVerifyCodeRequest(BaseModel):
    email: str
    challenge_id: str
    verification_code: str


class UserActiveUpdate(BaseModel):
    is_active: bool


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


class RegistrationLimitUpdate(BaseModel):
    limit: int


class SMTPVerificationConfirmRequest(BaseModel):
    challenge_id: str
    verification_code: str


class CaptchaRequest(BaseModel):
    session_id: str


class CaptchaResponse(BaseModel):
    success: bool
    captcha_image: str
    session_id: str
    message: str


class VerifyCaptchaRequest(BaseModel):
    session_id: str
    captcha_code: str


class VerifyCaptchaResponse(BaseModel):
    success: bool
    message: str


def generate_token() -> str:
    """生成随机token"""
    return secrets.token_urlsafe(32)


def create_login_session(user: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
    """创建并持久化后台登录会话"""
    token = generate_token()
    is_admin = user.get('is_admin', False) or user['username'] == ADMIN_USERNAME
    expires_at = time.time() + TOKEN_EXPIRE_TIME
    token_data = {
        'user_id': user['id'],
        'username': user['username'],
        'is_admin': is_admin,
        'timestamp': time.time(),
        'expires_at': expires_at
    }

    SESSION_TOKENS[token] = token_data
    db_manager.save_auth_session(
        token=token,
        user_id=user['id'],
        username=user['username'],
        is_admin=is_admin,
        expires_at=expires_at
    )
    db_manager.cleanup_expired_auth_sessions()
    return token, token_data


def _drop_user_sessions_from_memory(user_id: int) -> None:
    for token, data in list(SESSION_TOKENS.items()):
        if int(data.get('user_id') or 0) == int(user_id):
            SESSION_TOKENS.pop(token, None)


def _masked_identifier(identifier: str) -> str:
    value = str(identifier or '').strip()
    if '@' in value:
        return mask_email_for_log(value)
    return f"{value[:2]}***" if value else "[空账号]"


def _client_ip(request: Request) -> str:
    peer_ip = request.client.host if request.client else "0.0.0.0"
    trusted = db_manager.get_system_setting('auth_trusted_proxies') or ''
    return resolve_client_ip(peer_ip, request.headers, trusted)


def _registration_state() -> Dict[str, Any]:
    settings = db_manager.get_all_system_settings()
    capacity = db_manager.registration_service.registration_capacity()
    return registration_readiness(
        settings,
        db_path=db_manager.db_path,
        user_count=capacity['user_count'],
    )


def _require_registration_enabled() -> Dict[str, Any]:
    try:
        state = _registration_state()
    except Exception as exc:
        raise RegistrationError(
            "REGISTRATION_UNAVAILABLE",
            "注册服务暂不可用",
            http_status=503,
        ) from exc
    if not state['enabled']:
        raise RegistrationError(
            "REGISTRATION_CLOSED",
            "注册暂未开放",
            http_status=403,
        )
    return state


def _require_verified_smtp(settings: Dict[str, Any]) -> None:
    status = smtp_configuration_status(settings, db_path=db_manager.db_path)
    if not status['smtp_verified']:
        raise RegistrationError(
            "SMTP_NOT_READY",
            "邮件服务暂不可用",
            http_status=503,
        )


def verify_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Optional[Dict[str, Any]]:
    """验证token并返回用户信息"""
    if not credentials:
        return None

    token = credentials.credentials
    token_data = SESSION_TOKENS.get(token)
    if not token_data:
        token_data = db_manager.get_auth_session(token)
        if not token_data:
            return None
        SESSION_TOKENS[token] = token_data

    # 检查token是否过期
    expires_at = token_data.get('expires_at', token_data.get('timestamp', 0) + TOKEN_EXPIRE_TIME)
    if time.time() > expires_at:
        SESSION_TOKENS.pop(token, None)
        db_manager.delete_auth_session(token)
        return None

    current_user = db_manager.get_user_by_id(token_data.get('user_id'))
    if not current_user or not current_user.get('is_active'):
        SESSION_TOKENS.pop(token, None)
        db_manager.delete_auth_session(token)
        return None

    return token_data


def verify_admin_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    """验证管理员token"""
    user_info = verify_token(credentials)
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")

    # 检查是否是管理员
    if user_info['username'] != ADMIN_USERNAME:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    return user_info


def require_auth(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    """需要认证的依赖，返回用户信息"""
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")
    return user_info


def get_current_user(user_info: Dict[str, Any] = Depends(require_auth)) -> Dict[str, Any]:
    """获取当前登录用户信息"""
    return user_info


def get_current_user_optional(user_info: Optional[Dict[str, Any]] = Depends(verify_token)) -> Optional[Dict[str, Any]]:
    """获取当前用户信息（可选，不强制要求登录）"""
    return user_info


def get_user_log_prefix(user_info: Dict[str, Any] = None) -> str:
    """获取用户日志前缀"""
    if user_info:
        stable_id = str(user_info.get('user_id') or 'unknown')
        digest = hashlib.sha256(stable_id.encode('utf-8')).hexdigest()[:10]
        return f"【user_{digest}】"
    return "【系统】"


def require_admin(current_user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    """要求管理员权限"""
    if current_user['username'] != 'admin':
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return current_user


# 正式控制台域名与回环同等视为“本机”：单机自用形态下，公网域名经本机
# cloudflared/nginx 回环转发进入（client 地址仍是 127.0.0.1），仅 Host 头不同；
# 用户人就在这台 Mac 前，不应仅因地址栏写法被当成远程用户。
_SERVER_BROWSER_CONSOLE_HOSTS = {
    str(urlsplit(PUBLIC_CONSOLE_ORIGIN).hostname or "").lower(),
}


def _is_strict_loopback_request(request: Request) -> bool:
    """client 地址与 Host 头都必须是回环。

    旧版扩展导入（protocol v1）等保持最严格边界，不随服务端浏览器门禁
    对正式控制台域名的放宽而放宽。
    """
    client_host = request.client.host if request.client else ""
    host_header = str(request.headers.get("host") or "").strip()
    try:
        console_host = urlsplit(f"//{host_header}").hostname or ""
    except ValueError:
        return False
    return is_loopback_host(client_host) and is_loopback_host(console_host)


def _is_loopback_console_request(request: Request) -> bool:
    if _is_strict_loopback_request(request):
        return True
    client_host = request.client.host if request.client else ""
    host_header = str(request.headers.get("host") or "").strip()
    try:
        console_host = urlsplit(f"//{host_header}").hostname or ""
    except ValueError:
        return False
    return is_loopback_host(client_host) and (
        console_host.lower() in _SERVER_BROWSER_CONSOLE_HOSTS
    )


def _require_server_browser_access(
    request: Request,
    current_user: Dict[str, Any],
) -> None:
    # 单机自用形态：服务端 Chrome 登录的真正安全边界是控制台登录态——未登录请求
    # 在上游 get_current_user 处即被 401 拒绝，能走到这里的只有持有效会话的用户
    # 本人。网络来源（client 地址 / Host 头）不再作为拒绝条件：经 Cloudflare 隧道
    # 回流时 Host 头由远端 ingress 决定、不可控，曾两次把本人误判为远程用户。
    # 非白名单来源仅记录观测日志，供将来多用户化时重新收紧边界。
    del current_user
    if not _is_loopback_console_request(request):
        client_host = request.client.host if request.client else ""
        host_header = str(request.headers.get("host") or "").strip()
        logger.warning(
            "服务端浏览器入口收到非白名单来源请求（已按单机自用策略放行）："
            f"client={client_host or 'unknown'} host={host_header or 'unknown'}"
        )


def _has_server_browser_access(
    request: Request,
    current_user: Dict[str, Any],
) -> bool:
    try:
        _require_server_browser_access(request, current_user)
    except HTTPException:
        return False
    return True


def _client_browser_required_detail() -> Dict[str, str]:
    return {
        "code": "client_browser_required",
        "message": "请在当前设备的 Chrome 或 Edge 中继续登录",
        "action": "open_client_browser",
    }


def log_with_user(level: str, message: str, user_info: Dict[str, Any] = None):
    """带用户信息的日志记录"""
    prefix = get_user_log_prefix(user_info)
    full_message = f"{prefix} {message}"

    if level.lower() == 'info':
        logger.info(full_message)
    elif level.lower() == 'error':
        logger.error(full_message)
    elif level.lower() == 'warning':
        logger.warning(full_message)
    elif level.lower() == 'debug':
        logger.debug(full_message)
    else:
        logger.info(full_message)


app = FastAPI(
    title="Xianyu Auto Reply API",
    version="1.10.4",
    description="闲鱼自动回复系统API",
    docs_url="/docs",
    redoc_url="/redoc"
)

# 添加 CORS 中间件支持前端跨域请求
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
        "http://127.0.0.1:3002",
    ],  # 允许的前端开发服务器地址
    allow_origin_regex=r"^chrome-extension://[a-p]{32}$",
    allow_credentials=True,  # 允许携带凭证
    allow_methods=["*"],  # 允许所有HTTP方法
    allow_headers=["*"],  # 允许所有请求头
)

# 初始化文件日志收集器
setup_file_logging()

# 添加一条测试日志
from loguru import logger
logger.info("Web服务器启动，文件日志收集器已初始化")

# 添加请求日志中间件
def _request_log_path(request: Request) -> str:
    route = request.scope.get("route")
    route_path = getattr(route, "path_format", None) or getattr(route, "path", None)
    return str(route_path or "<unmatched>")


@app.middleware("http")
async def log_requests(request, call_next):
    start_time = time.time()
    supplied_request_id = request.headers.get("X-Request-ID", "")
    request_id = supplied_request_id if re.fullmatch(r"[A-Za-z0-9._-]{8,80}", supplied_request_id) else secrets.token_hex(8)
    request.state.request_id = request_id

    logger.info(f"🌐 API请求: {request.method} request_id={request_id}")

    response = await call_next(request)
    response.headers["X-Request-ID"] = request_id

    process_time = time.time() - start_time
    logger.info(
        f"✅ API响应: {request.method} {_request_log_path(request)} - "
        f"{response.status_code} ({process_time:.3f}s)"
    )

    return response


@app.exception_handler(RegistrationError)
async def registration_error_with_request_id(request: Request, exc: RegistrationError):
    headers = {}
    if exc.retry_after is not None:
        headers['Retry-After'] = str(exc.retry_after)
    return JSONResponse(
        status_code=exc.http_status,
        content={
            "success": False,
            "code": exc.code,
            "message": exc.message,
            "retry_after": exc.retry_after,
            "request_id": getattr(request.state, "request_id", ""),
        },
        headers=headers,
    )


@app.exception_handler(RequestValidationError)
async def validation_error_without_input(request: Request, exc: RequestValidationError):
    errors = [
        {
            "location": [str(part) for part in error.get("loc", ())],
            "message": error.get("msg", "输入无效"),
            "type": error.get("type", "validation_error"),
        }
        for error in exc.errors()
    ]
    return JSONResponse(
        status_code=422,
        content={
            "success": False,
            "code": "REQUEST_VALIDATION_FAILED",
            "message": "请求参数无效",
            "errors": errors,
            "request_id": getattr(request.state, "request_id", ""),
        },
    )


@app.exception_handler(HTTPException)
async def http_exception_with_request_id(request: Request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail, "request_id": getattr(request.state, "request_id", "")},
        headers=exc.headers,
    )


@app.exception_handler(Exception)
async def unhandled_exception_with_request_id(request: Request, exc: Exception):
    request_id = getattr(request.state, "request_id", "")
    logger.exception(f"未处理请求异常 request_id={request_id}: {exc}")
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal Server Error", "request_id": request_id},
    )

# 提供前端静态文件
import os
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir, exist_ok=True)

# 挂载静态文件目录
app.mount('/static', StaticFiles(directory=static_dir), name='static')

# 挂载 /assets 路径，指向 static/assets 目录
# 这样访问 /assets/xxx.js 时会正确映射到 static_dir/assets/xxx.js
assets_dir = os.path.join(static_dir, 'assets')
os.makedirs(assets_dir, exist_ok=True)
app.mount('/assets', StaticFiles(directory=assets_dir), name='assets')


_STATIC_NO_STORE_PREFIXES = ('/static/', '/assets/')


@app.middleware("http")
async def static_error_responses_no_store(request: Request, call_next):
    # 静态资源的 4xx/5xx 必须禁止一切缓存：部署窗口内的资产 404 一旦被 CDN 按默认
    # 规则缓存（Cloudflare 对 .js 等扩展名默认 Browser Cache TTL 4h），会把错误响应
    # 投毒到用户浏览器，服务端修复后用户普通刷新仍白屏（2026-08-28 生产事故加固）。
    response = await call_next(request)
    if response.status_code >= 400 and request.url.path.startswith(_STATIC_NO_STORE_PREFIXES):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        response.headers['CDN-Cache-Control'] = 'no-store'
    return response

# 确保图片上传目录存在
uploads_dir = os.path.join(static_dir, 'uploads', 'images')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir, exist_ok=True)
    logger.info(f"创建图片上传目录: {uploads_dir}")

# 健康检查端点
@system_router.get('/health')
async def health_check():
    """健康检查端点，用于Docker健康检查和负载均衡器"""
    try:
        # 检查Cookie管理器状态
        manager_status = "ok" if cookie_manager.manager is not None else "error"

        # 检查数据库连接
        from db_manager import db_manager
        try:
            db_manager.get_all_cookies()
            db_status = "ok"
        except Exception:
            db_status = "error"

        # 获取系统状态
        import psutil
        cpu_percent = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()

        status = {
            "status": "healthy" if manager_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": time.time(),
            "services": {
                "cookie_manager": manager_status,
                "database": db_status
            },
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory_info.percent,
                "memory_available": memory_info.available
            },
            "migration_version": getattr(db_manager, "schema_version", "legacy"),
            "runtime_sessions": get_session_registry().summary(),
        }
        if status["status"] == "unhealthy":
            raise HTTPException(status_code=503, detail=status)

        return status

    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e)
        }


@system_router.get('/health/live')
async def health_live():
    return {"status": "alive", "timestamp": time.time()}


@system_router.get('/health/ready')
async def health_ready():
    try:
        db_manager.conn.execute("SELECT 1").fetchone()
        database_ready = True
    except Exception:
        database_ready = False
    manager_ready = cookie_manager.manager is not None
    payload = {
        "status": "ready" if database_ready and manager_ready else "not_ready",
        "timestamp": time.time(),
        "services": {
            "database": "ok" if database_ready else "error",
            "cookie_manager": "ok" if manager_ready else "error",
        },
        "migration_version": getattr(db_manager, "schema_version", "legacy"),
        "runtime_sessions": get_session_registry().summary(),
    }
    return JSONResponse(status_code=200 if payload["status"] == "ready" else 503, content=payload)


# 服务 React 前端 SPA - 所有前端路由都返回 index.html
async def serve_frontend():
    """服务 React 前端 SPA"""
    index_path = os.path.join(static_dir, 'index.html')
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Frontend not found. Please build the frontend first.</h3>')

@frontend_router.get('/', response_class=HTMLResponse)
async def root():
    return await serve_frontend()


# 登录页面路由 - 重定向到 React 前端
@frontend_router.get('/login.html', response_class=HTMLResponse)
async def login_page():
    return await serve_frontend()

@frontend_router.get('/login', response_class=HTMLResponse)
async def login_route():
    return await serve_frontend()


# 注册页面路由
@frontend_router.get('/register.html', response_class=HTMLResponse)
async def register_page():
    return await serve_frontend()

@frontend_router.get('/register', response_class=HTMLResponse)
async def register_route():
    return await serve_frontend()


# 注意：不要在这里定义 /admin 或 /admin/{path} 路由
# 因为后端有 /admin/users, /admin/logs 等 API 路由
# 前端 SPA 通过根路由 / 加载，由 React Router 处理客户端路由
# 文件末尾的 catch-all 路由会处理前端页面的直接访问



# 登录接口
@auth_router.post('/login')
async def login(request: LoginRequest, http_request: Request):
    identifier = str(
        request.identifier or request.username or request.email or ''
    ).strip()
    if not identifier or not request.password:
        raise RegistrationError(
            "LOGIN_INPUT_REQUIRED",
            "请输入用户名或邮箱及密码",
        )

    client_ip = _client_ip(http_request)
    db_manager.auth_rate_limiter.check_login_limit(
        ip=client_ip,
        account=identifier,
    )
    verified = db_manager.verify_user_password(identifier, request.password)
    if not verified:
        logger.warning(f"登录失败 account={_masked_identifier(identifier)}")
        db_manager.auth_rate_limiter.record_login_result(
            ip=client_ip,
            account=identifier,
            success=False,
        )
        raise RegistrationError(
            "INVALID_CREDENTIALS",
            "账号或密码错误",
            http_status=401,
        )

    user = db_manager.user_repository.get_by_identifier(identifier)
    if not user or not user.get('is_active'):
        db_manager.auth_rate_limiter.record_login_result(
            ip=client_ip,
            account=identifier,
            success=False,
        )
        raise RegistrationError(
            "INVALID_CREDENTIALS",
            "账号或密码错误",
            http_status=401,
        )

    db_manager.auth_rate_limiter.record_login_result(
        ip=client_ip,
        account=identifier,
        success=True,
    )
    token, _ = create_login_session(user)
    logger.info(f"用户登录成功 user_id={user['id']}")
    return LoginResponse(
        success=True,
        token=token,
        message="登录成功",
        user_id=user['id'],
        username=user['username'],
        is_admin=(user['username'] == ADMIN_USERNAME),
    )


# 验证token接口
@auth_router.get('/verify')
async def verify(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    if user_info:
        return {
            "authenticated": True,
            "user_id": user_info['user_id'],
            "username": user_info['username'],
            "is_admin": user_info['username'] == ADMIN_USERNAME
        }
    return {"authenticated": False}


# 登出接口
@auth_router.post('/logout')
async def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if credentials:
        SESSION_TOKENS.pop(credentials.credentials, None)
        db_manager.delete_auth_session(credentials.credentials)
    return {"message": "已登出"}


# 修改管理员密码接口
@auth_router.post('/change-admin-password')
async def change_admin_password(request: ChangePasswordRequest, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    try:
        # 验证当前密码（使用用户表验证）
        if not db_manager.verify_user_password('admin', request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 密码与全部持久化会话在同一事务内更新
        changed_user_id = db_manager.update_user_password_and_revoke_sessions(
            'admin',
            request.new_password,
        )

        if changed_user_id is not None:
            _drop_user_sessions_from_memory(changed_user_id)
            logger.info(f"【admin#{admin_user['user_id']}】管理员密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改管理员密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 普通用户修改密码接口
@auth_router.post('/change-password')
async def change_user_password(request: ChangePasswordRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    try:
        username = current_user.get('username')
        user_id = current_user.get('user_id')

        if not username:
            return {"success": False, "message": "无法获取用户信息"}

        # 验证当前密码
        if not db_manager.verify_user_password(username, request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 密码与全部持久化会话在同一事务内更新
        changed_user_id = db_manager.update_user_password_and_revoke_sessions(
            username,
            request.new_password,
        )

        if changed_user_id is not None:
            _drop_user_sessions_from_memory(changed_user_id)
            logger.info(f"【{username}#{user_id}】用户密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改用户密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 检查是否使用默认密码
@auth_router.get('/api/check-default-password')
async def check_default_password(current_user: Dict[str, Any] = Depends(get_current_user)):
    from db_manager import db_manager

    try:
        username = current_user.get('username')
        is_admin = current_user.get('is_admin', False)

        logger.info(f"检查默认密码: username={username}, is_admin={is_admin}")

        # 只检查admin用户
        if not is_admin or username != 'admin':
            logger.info(f"非admin用户，跳过检查")
            return {"using_default": False}

        # 检查是否使用默认密码
        using_default = db_manager.verify_user_password('admin', DEFAULT_ADMIN_PASSWORD)
        logger.info(f"默认密码检查结果: {using_default}")

        return {"using_default": using_default}

    except Exception as e:
        logger.error(f"检查默认密码异常: {e}")
        return {"using_default": False}


# 生成图形验证码接口
@auth_router.post('/generate-captcha')
async def generate_captcha(request: CaptchaRequest):
    from db_manager import db_manager

    try:
        # 生成图形验证码
        captcha_text, captcha_image = db_manager.generate_captcha()

        if not captcha_image:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码生成失败"
            )

        # 保存验证码到数据库
        if db_manager.save_captcha(request.session_id, captcha_text):
            return CaptchaResponse(
                success=True,
                captcha_image=captcha_image,
                session_id=request.session_id,
                message="图形验证码生成成功"
            )
        else:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码保存失败"
            )

    except Exception as e:
        logger.error(f"生成图形验证码失败: {e}")
        return CaptchaResponse(
            success=False,
            captcha_image="",
            session_id=request.session_id,
            message="图形验证码生成失败"
        )


# 验证图形验证码接口
@auth_router.post('/verify-captcha')
async def verify_captcha(request: VerifyCaptchaRequest):
    from db_manager import db_manager

    try:
        if db_manager.verify_captcha(request.session_id, request.captcha_code):
            return VerifyCaptchaResponse(
                success=True,
                message="图形验证码验证成功"
            )
        else:
            return VerifyCaptchaResponse(
                success=False,
                message="图形验证码错误或已过期"
            )

    except Exception as e:
        logger.error(f"验证图形验证码失败: {e}")
        return VerifyCaptchaResponse(
            success=False,
            message="图形验证码验证失败"
        )


# ==================== 极验滑动验证码 ====================

# 极验验证状态只用于短期幂等响应。公开接口必须同时限制单键大小和
# 进程内条目数量，避免攻击者用任意 challenge 扩大常驻内存。
GEETEST_CHALLENGE_MAX_LENGTH = 64
GEETEST_PROOF_MAX_LENGTH = 2048
GEETEST_STATUS_MAX_ENTRIES = 2048
GEETEST_STATUS_TTL_SECONDS = 300
geetest_status_store: OrderedDict[str, Dict[str, float | int]] = OrderedDict()
_geetest_status_lock = threading.Lock()

# Lightweight, process-local diagnostics caches. They deliberately avoid account
# browser profiles and never call the Xianyu website.
_OPS_DB_PROBE_TTL_SECONDS = 60.0
_ops_db_probe_lock = threading.Lock()
_ops_db_probe_cache: Dict[str, Any] = {}
_BROWSER_PROBE_TTL_SECONDS = 600.0
_browser_probe_lock = threading.Lock()
_browser_probe_cache: Dict[str, Any] = {}
_browser_probe_checking = False


def _cleanup_expired_geetest_status_locked(current_time: float) -> None:
    """Remove expired rows while the caller holds ``_geetest_status_lock``."""
    expired_keys = [
        key
        for key, value in geetest_status_store.items()
        if float(value["expires_at"]) < current_time
    ]
    for key in expired_keys:
        geetest_status_store.pop(key, None)


def cleanup_expired_geetest_status() -> None:
    """清理过期的极验验证状态。"""
    current_time = time.time()
    with _geetest_status_lock:
        _cleanup_expired_geetest_status_locked(current_time)


def set_geetest_status(challenge: str, status: int) -> bool:
    """设置有界的极验验证状态；非法或过长键不会进入内存。"""
    normalized = str(challenge or "").strip()
    if not normalized or len(normalized) > GEETEST_CHALLENGE_MAX_LENGTH:
        return False
    if status not in {0, 1}:
        return False
    current_time = time.time()
    with _geetest_status_lock:
        _cleanup_expired_geetest_status_locked(current_time)
        geetest_status_store[normalized] = {
            "status": status,
            "expires_at": current_time + GEETEST_STATUS_TTL_SECONDS,
        }
        geetest_status_store.move_to_end(normalized)
        while len(geetest_status_store) > GEETEST_STATUS_MAX_ENTRIES:
            geetest_status_store.popitem(last=False)
    return True


def get_geetest_status(challenge: str) -> int:
    """获取极验验证状态，返回0表示未验证或已过期"""
    normalized = str(challenge or "").strip()
    if not normalized or len(normalized) > GEETEST_CHALLENGE_MAX_LENGTH:
        return 0
    current_time = time.time()
    with _geetest_status_lock:
        _cleanup_expired_geetest_status_locked(current_time)
        stored = geetest_status_store.get(normalized)
        if stored and float(stored["expires_at"]) > current_time:
            return int(stored["status"])
    return 0


class GeetestRegisterResponse(BaseModel):
    """极验验证码初始化响应"""
    success: bool
    code: int = 200
    message: str = ""
    data: Optional[dict] = None


class GeetestValidateRequest(BaseModel):
    """极验二次验证请求"""
    challenge: str = Field(
        ...,
        min_length=1,
        max_length=GEETEST_CHALLENGE_MAX_LENGTH,
    )
    validate_str: str = Field(
        ...,
        alias='validate',
        min_length=1,
        max_length=GEETEST_PROOF_MAX_LENGTH,
    )
    seccode: str = Field(..., min_length=1, max_length=GEETEST_PROOF_MAX_LENGTH)

    model_config = {'populate_by_name': True}


class GeetestValidateResponse(BaseModel):
    """极验二次验证响应"""
    success: bool
    code: int = 200
    message: str = ""


@auth_router.get('/geetest/register', response_model=GeetestRegisterResponse)
async def geetest_register():
    """
    获取极验验证码初始化参数

    前端调用此接口获取gt、challenge等参数，用于初始化验证码组件
    """
    try:
        from utils.geetest import GeetestLib

        gt_lib = GeetestLib()
        result = await gt_lib.register()

        data = result.to_dict()
        logger.info(
            f"极验初始化结果: status={result.status}, "
            f"fields={sorted(str(key) for key in data.keys())}"
        )

        # 记录初始状态
        challenge = data.get("challenge", "")
        if challenge:
            set_geetest_status(challenge, 0)

        return GeetestRegisterResponse(
            success=True,
            code=200,
            message="获取成功" if result.status == 1 else "宕机模式",
            data=data
        )

    except Exception as e:
        logger.error(f"极验初始化失败: {e}")
        # 返回本地初始化结果
        try:
            from utils.geetest import GeetestLib
            gt_lib = GeetestLib()
            result = gt_lib.local_init()
            data = result.to_dict()

            # 记录初始状态
            challenge = data.get("challenge", "")
            if challenge:
                set_geetest_status(challenge, 0)

            return GeetestRegisterResponse(
                success=True,
                code=200,
                message="本地初始化",
                data=data
            )
        except Exception as e2:
            logger.error(f"极验本地初始化也失败: {e2}")
            return GeetestRegisterResponse(
                success=False,
                code=500,
                message="验证码服务异常"
            )


@auth_router.post('/geetest/validate', response_model=GeetestValidateResponse)
async def geetest_validate(request: GeetestValidateRequest):
    """
    极验二次验证

    用户完成滑动验证后，前端调用此接口进行二次验证
    """
    try:
        # 检查是否已经验证过
        if get_geetest_status(request.challenge) == 1:
            return GeetestValidateResponse(
                success=True,
                code=200,
                message="验证通过"
            )

        from utils.geetest import GeetestLib

        gt_lib = GeetestLib()

        # 判断是正常模式还是宕机模式
        # 通过challenge长度判断：正常模式challenge是32位MD5，宕机模式是UUID
        is_normal_mode = len(request.challenge) == 32

        if is_normal_mode:
            result = await gt_lib.success_validate(
                request.challenge,
                request.validate_str,
                request.seccode
            )
        else:
            result = gt_lib.fail_validate(
                request.challenge,
                request.validate_str,
                request.seccode
            )

        if result.status == 1:
            # 记录验证通过状态
            set_geetest_status(request.challenge, 1)

            return GeetestValidateResponse(
                success=True,
                code=200,
                message="验证通过"
            )
        else:
            return GeetestValidateResponse(
                success=False,
                code=400,
                message=result.msg or "验证失败"
            )

    except Exception as e:
        logger.error(f"极验二次验证失败: {e}")
        return GeetestValidateResponse(
            success=False,
            code=500,
            message="验证服务异常"
        )


@auth_router.get('/api/auth/registration-config')
def get_registration_config():
    try:
        settings = db_manager.get_all_system_settings()
        state = _registration_state()
        support_email = str(settings.get('support_email') or '').strip()
        if '\r' in support_email or '\n' in support_email:
            support_email = ''
        return {
            "enabled": state['enabled'],
            "ready": state['ready'],
            "invite_required": False,
            "terms_version": state['terms_version'] or 'v2',
            "terms_url": "/terms",
            "privacy_url": "/privacy",
            "support_email": support_email,
            "message": "注册已开放" if state['enabled'] else "注册暂未开放",
        }
    except Exception:
        logger.warning("读取公开注册状态失败")
        return {
            "enabled": False,
            "ready": False,
            "invite_required": False,
            "terms_version": "v2",
            "terms_url": "/terms",
            "privacy_url": "/privacy",
            "support_email": "",
            "message": "注册暂未开放",
        }


@auth_router.post('/api/auth/captcha')
def create_auth_captcha(http_request: Request):
    client_ip = _client_ip(http_request)
    db_manager.auth_rate_limiter.enforce_captcha(client_ip)
    captcha_text, captcha_image = db_manager.generate_captcha()
    if not captcha_text or not captcha_image:
        raise RegistrationError(
            "CAPTCHA_UNAVAILABLE",
            "图形验证码暂不可用",
            http_status=503,
        )
    challenge = db_manager.registration_service.create_challenge(
        purpose="captcha",
        subject=client_ip,
        secret=captcha_text.upper(),
    )
    return {
        "success": True,
        "challenge_id": challenge['challenge_id'],
        "captcha_image": captcha_image,
        "expires_in": 600,
    }


@auth_router.post('/api/auth/email-code')
async def send_auth_email_code(request: EmailCodeRequest, http_request: Request):
    email = normalize_email(request.email).normalized
    client_ip = _client_ip(http_request)
    settings = db_manager.get_all_system_settings()
    if request.purpose == 'register':
        _require_registration_enabled()
        challenge_purpose = 'register_email'
        challenge_context = ''
        subject = "闲鱼监控台注册验证码"
    else:
        _require_verified_smtp(settings)
        challenge_purpose = 'password_reset_email'
        challenge_context = ''
        subject = "闲鱼监控台密码重置验证码"

    db_manager.registration_service.consume_challenge(
        challenge_id=request.captcha_challenge_id,
        purpose="captcha",
        subject=client_ip,
        secret=request.captcha_code.upper(),
    )
    db_manager.auth_rate_limiter.enforce_email_send(client_ip, email)
    _require_verified_smtp(settings)

    verification_code = f"{secrets.randbelow(1_000_000):06d}"
    user = db_manager.get_user_by_email_for_public_auth(email)
    if request.purpose == 'register':
        actionable_target = user is None
    else:
        actionable_target = bool(user and user.get('is_active'))
    decoy_secret = secrets.token_urlsafe(32)
    challenge_secret = (
        verification_code
        if actionable_target
        else decoy_secret
    )
    text_content = (
        f"您的验证码是 {verification_code}\n\n"
        "验证码在 10 分钟内有效，最多可尝试 5 次。请勿向任何人泄露。\n"
        "如非本人操作，请忽略此邮件。"
    )
    try:
        await asyncio.to_thread(
            SMTPEmailSender().send,
            settings,
            recipient=email,
            subject=subject,
            text=text_content,
        )
    except (SMTPConfigurationError, SMTPDeliveryError) as exc:
        logger.warning(
            f"认证邮件发送失败 type={type(exc).__name__} "
            f"email={mask_email_for_log(email)}"
        )
        raise RegistrationError(
            "EMAIL_SEND_FAILED",
            "验证码邮件发送失败，请稍后重试",
            http_status=502,
        ) from exc

    challenge = db_manager.registration_service.create_challenge(
        purpose=challenge_purpose,
        subject=email,
        context=challenge_context,
        secret=challenge_secret,
    )
    logger.info(f"认证验证码请求已处理 purpose={request.purpose}")
    return {
        "success": True,
        "challenge_id": challenge['challenge_id'],
        "expires_in": 600,
        "cooldown_seconds": 60,
        "message": "验证码已发送，请查收邮件",
    }


@auth_router.post('/send-verification-code')
async def send_verification_code(_request: SendCodeRequest):
    raise RegistrationError(
        "LEGACY_AUTH_ENDPOINT_REMOVED",
        "此接口已停用，请改用 /api/auth/captcha 和 /api/auth/email-code",
        http_status=410,
    )


@auth_router.post('/register')
async def register(request: RegisterRequest, http_request: Request):
    if not request.terms_accepted:
        raise RegistrationError("TERMS_NOT_ACCEPTED", "请先同意服务条款和隐私说明")
    _require_registration_enabled()
    client_ip = _client_ip(http_request)
    db_manager.auth_rate_limiter.check_registration_limit(client_ip)
    try:
        user = db_manager.registration_service.register_user(
            username=request.username,
            email=request.email,
            password=request.password,
            challenge_id=request.challenge_id,
            verification_code=request.verification_code,
            terms_version=request.terms_version,
            invite_code=request.invite_code,
        )
    except RegistrationError:
        db_manager.auth_rate_limiter.record_registration_failure(client_ip)
        raise
    except Exception as exc:
        logger.error(f"注册事务失败 type={type(exc).__name__}")
        raise RegistrationError(
            "REGISTRATION_FAILED",
            "注册失败，请稍后重试",
            http_status=503,
        ) from exc

    token, _ = create_login_session(user)
    logger.info(f"注册成功 user_id={user['id']}")
    return RegisterResponse(
        success=True,
        token=token,
        message="注册成功",
        user_id=user['id'],
        username=user['username'],
        is_admin=False,
    )


@auth_router.post('/api/auth/password-reset/verify-code')
async def verify_password_reset_code(request: PasswordResetVerifyCodeRequest):
    settings = db_manager.get_all_system_settings()
    _require_verified_smtp(settings)
    try:
        grant = db_manager.registration_service.verify_password_reset_code(
            email=request.email,
            challenge_id=request.challenge_id,
            verification_code=request.verification_code,
        )
    except RegistrationError:
        raise
    except Exception as exc:
        logger.error(f"密码重置邮箱校验失败 type={type(exc).__name__}")
        raise RegistrationError(
            "PASSWORD_RESET_VERIFICATION_FAILED",
            "邮箱验证失败，请稍后重试",
            http_status=503,
        ) from exc
    logger.info("密码重置邮箱校验成功")
    return {
        "success": True,
        "reset_grant_id": grant["grant_id"],
        "reset_grant_token": grant["grant_token"],
        "expires_in": max(0, int(grant["expires_at"] - time.time())),
        "message": "邮箱验证成功",
    }


@auth_router.post('/api/auth/password-reset')
async def reset_user_password(request: PasswordResetRequest):
    use_grant = bool(request.reset_grant_id or request.reset_grant_token)
    if use_grant and not (request.reset_grant_id and request.reset_grant_token):
        raise RegistrationError(
            "PASSWORD_RESET_GRANT_REQUIRED",
            "密码重置授权不完整，请重新验证邮箱",
        )
    if not use_grant and not (request.challenge_id and request.verification_code):
        raise RegistrationError(
            "PASSWORD_RESET_VERIFICATION_REQUIRED",
            "请先完成邮箱验证",
        )

    try:
        if use_grant:
            user_id = db_manager.registration_service.reset_password_with_grant(
                email=request.email,
                new_password=request.new_password,
                grant_id=request.reset_grant_id or "",
                grant_token=request.reset_grant_token or "",
            )
        else:
            settings = db_manager.get_all_system_settings()
            _require_verified_smtp(settings)
            user_id = db_manager.registration_service.reset_password(
                email=request.email,
                new_password=request.new_password,
                challenge_id=request.challenge_id or "",
                verification_code=request.verification_code or "",
            )
    except RegistrationError:
        raise
    except Exception as exc:
        logger.error(f"密码重置失败 type={type(exc).__name__}")
        raise RegistrationError(
            "PASSWORD_RESET_FAILED",
            "密码重置失败，请稍后重试",
            http_status=503,
        ) from exc
    _drop_user_sessions_from_memory(user_id)
    logger.info(f"用户密码重置成功 user_id={user_id}")
    return {
        "success": True,
        "message": "密码已重置，请重新登录",
    }


# ------------------------- 发送消息接口 -------------------------

# 兼容旧接口的后备秘钥，仅允许通过环境变量注入。
API_SECRET_KEY = os.getenv("XIANYU_REPLY_API_SECRET", "")

# 旧秘钥调用方必须显式绑定到一个用户，未绑定时该通道保持关闭。
SEND_MESSAGE_SECRET_OWNER_SETTING = 'qq_reply_secret_user_id'


class SendMessageRequest(BaseModel):
    api_key: str = ''
    cookie_id: str
    chat_id: str
    to_user_id: str
    message: str


class SendMessageResponse(BaseModel):
    success: bool
    message: str


def verify_api_key(api_key: str) -> bool:
    """验证API秘钥；秘钥未配置或校验异常一律判定失败（失败关闭）"""
    if not api_key:
        return False
    try:
        # 从系统设置中获取QQ回复消息秘钥
        qq_secret_key = db_manager.get_system_setting('qq_reply_secret_key')

        # 系统设置未配置时才回落到环境变量注入的后备秘钥
        if not qq_secret_key:
            qq_secret_key = API_SECRET_KEY

        if not qq_secret_key:
            return False

        return secrets.compare_digest(str(api_key), str(qq_secret_key))
    except Exception as e:
        logger.error(f"验证API秘钥时发生异常: {type(e).__name__}")
        return False


def resolve_send_message_caller(api_key: str,
                                current_user: Optional[Dict[str, Any]]) -> Optional[int]:
    """解析发信调用方的用户身份

    优先使用登录态；没有登录态时才走旧的共享秘钥通道，且该秘钥必须已在系统设置里
    绑定到一个存在且启用的用户。任何一步不成立都返回 None（失败关闭）。
    """
    if current_user and current_user.get('user_id') is not None:
        return int(current_user['user_id'])

    if not verify_api_key(api_key):
        return None

    try:
        bound_raw = db_manager.get_system_setting(SEND_MESSAGE_SECRET_OWNER_SETTING) or ''
        bound_user_id = int(str(bound_raw).strip())
    except (TypeError, ValueError):
        logger.warning("发信秘钥未绑定用户，已拒绝该调用")
        return None
    except Exception as e:
        logger.error(f"读取发信秘钥绑定用户失败: {type(e).__name__}")
        return None

    bound_user = db_manager.get_user_by_id(bound_user_id)
    if not bound_user or not bound_user.get('is_active'):
        logger.warning("发信秘钥绑定的用户不存在或已停用，已拒绝该调用")
        return None
    return bound_user_id


@system_router.post('/send-message', response_model=SendMessageResponse)
async def send_message_api(
    request: SendMessageRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional),
):
    """发送消息API接口（登录态优先，兼容已绑定用户的旧秘钥）"""
    try:
        # 清理所有参数中的换行符
        def clean_param(param_str):
            """清理参数中的换行符"""
            if isinstance(param_str, str):
                return param_str.replace('\\n', '').replace('\n', '')
            return param_str

        # 清理所有参数
        cleaned_api_key = clean_param(request.api_key)
        cleaned_cookie_id = clean_param(request.cookie_id)
        cleaned_chat_id = clean_param(request.chat_id)
        cleaned_to_user_id = clean_param(request.to_user_id)
        cleaned_message = clean_param(request.message)

        # 解析调用方身份：登录态优先，其次是已绑定用户的旧秘钥
        caller_user_id = resolve_send_message_caller(cleaned_api_key, current_user)
        if caller_user_id is None:
            logger.warning("发信调用方身份验证失败")
            return SendMessageResponse(
                success=False,
                message="身份验证失败"
            )

        # 验证必需参数不能为空
        required_params = {
            'cookie_id': cleaned_cookie_id,
            'chat_id': cleaned_chat_id,
            'to_user_id': cleaned_to_user_id,
            'message': cleaned_message
        }

        for param_name, param_value in required_params.items():
            if not param_value:
                logger.warning(f"必需参数 {param_name} 为空")
                return SendMessageResponse(
                    success=False,
                    message=f"参数 {param_name} 不能为空"
                )

        # 账号归属校验：只允许对调用方名下的闲鱼账号发信
        caller_cookies = db_manager.get_all_cookies(caller_user_id)
        if cleaned_cookie_id not in caller_cookies:
            logger.warning("发信目标账号不属于调用方，已拒绝")
            return SendMessageResponse(
                success=False,
                message="账号不存在或无权操作"
            )

        # 直接获取XianyuLive实例，跳过cookie_manager检查
        from XianyuAutoAsync import XianyuLive
        live_instance = XianyuLive.get_instance(cleaned_cookie_id)

        # 实例缺失与连接断开对外统一为“账号当前不可发信”，不暴露运行态细节
        if not live_instance or not live_instance.ws or live_instance.ws.closed:
            logger.warning(f"账号当前不可发信: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="账号当前不可发送消息，请稍后重试"
            )

        # 发送消息（使用清理后的所有参数）
        await live_instance.send_msg(
            live_instance.ws,
            cleaned_chat_id,
            cleaned_to_user_id,
            cleaned_message
        )

        logger.info(
            f"API消息发送成功: account={cleaned_cookie_id}, "
            f"message_length={len(cleaned_message)}"
        )

        return SendMessageResponse(
            success=True,
            message="消息发送成功"
        )

    except Exception as e:
        logger.error(f"API发送消息异常: error_type={type(e).__name__}")
        return SendMessageResponse(
            success=False,
            message="发送消息失败，请稍后重试"
        )


# ------------------------- 账号 / 关键字管理接口 -------------------------


class CookieIn(BaseModel):
    id: Optional[str] = None
    value: str = Field(..., min_length=1)


class OfficialWindowLoginIn(BaseModel):
    mode: str = Field('sms', pattern='^sms$')
    account: Optional[str] = Field(None, max_length=200)


class BrowserExtensionCookieIn(BaseModel):
    name: str = Field(..., min_length=1, max_length=256)
    value: str = Field(..., max_length=8192)
    domain: str = Field(..., min_length=1, max_length=255)
    path: str = Field("/", min_length=1, max_length=1024)
    secure: bool = False
    httpOnly: bool = False
    sameSite: Optional[str] = None
    expirationDate: Optional[float] = None
    storeId: Optional[str] = Field(None, max_length=128)
    partitionKey: Optional[Dict[str, Any]] = None


class BrowserExtensionImportIn(BaseModel):
    protocol_version: int = 1
    pairing_id: str = Field(..., min_length=8, max_length=80)
    pairing_code: Optional[str] = Field(None, min_length=6, max_length=128)
    pairing_token: Optional[str] = Field(None, min_length=32, max_length=128)
    cookies: List[BrowserExtensionCookieIn]
    user_agent: str = Field(..., min_length=1, max_length=MAX_USER_AGENT_LENGTH)


class ClientBrowserDeviceIn(BaseModel):
    device_id: str = Field(..., min_length=16, max_length=80)
    browser_family: Literal["chrome", "edge"]
    client_type: Literal["extension"] = "extension"
    display_name: str = Field("当前设备", min_length=1, max_length=80)
    signing_public_jwk: Dict[str, Any]
    encryption_public_jwk: Dict[str, Any]


class ClientBrowserSessionIn(BaseModel):
    device_id: str = Field(..., min_length=16, max_length=80)
    mode: Literal["qr", "sms", "password"]
    client_type: Literal["extension"] = "extension"


class ClientBrowserChallengeIn(BaseModel):
    device_id: str = Field(..., min_length=16, max_length=80)
    purpose: Literal[
        "login_import",
        "renewal_claim",
        "renewal_complete",
        "renewal_action_required",
    ]


class ClientBrowserLoginCookieIn(BrowserExtensionCookieIn):
    pass


class ClientBrowserLoginImportIn(BaseModel):
    session_id: str = Field(..., min_length=8, max_length=80)
    device_id: str = Field(..., min_length=16, max_length=80)
    mode: Literal["qr", "sms", "password"]
    challenge_id: str = Field(..., min_length=8, max_length=80)
    signature: str = Field(..., min_length=40, max_length=256)
    cookies: List[ClientBrowserLoginCookieIn]
    user_agent: str = Field(..., min_length=1, max_length=MAX_USER_AGENT_LENGTH)


class ClientBrowserSessionAuthorizeIn(BaseModel):
    device_id: str = Field(..., min_length=16, max_length=80)
    mode: Literal["qr", "sms", "password"]


class ClientBrowserConfirmIn(BaseModel):
    account_id: str = Field(..., min_length=1, max_length=200)


class ClientRenewalProofIn(BaseModel):
    device_id: str = Field(..., min_length=16, max_length=80)
    challenge_id: str = Field(..., min_length=8, max_length=80)
    signature: str = Field(..., min_length=40, max_length=256)


class ClientRenewalResultIn(ClientRenewalProofIn):
    outcome: Literal["completed", "action_required", "failed"]
    error_code: str = Field("", max_length=80)
    cookies: List[ClientBrowserLoginCookieIn] = Field(default_factory=list)
    user_agent: str = Field("", max_length=MAX_USER_AGENT_LENGTH)


class QRLoginCancelIn(BaseModel):
    ended_by: Literal[
        "user_cancelled",
        "switched_method",
        "switched_to_extension",
    ] = "user_cancelled"


class BrowserInteractionPointIn(BaseModel):
    x: float = Field(..., ge=0.0, le=1.0)
    y: float = Field(..., ge=0.0, le=1.0)


class BrowserInteractionIn(BaseModel):
    kind: Literal["gesture", "text", "key", "wheel"]
    frame_revision: int = Field(..., ge=1)
    points: List[BrowserInteractionPointIn] = Field(
        default_factory=list,
        max_length=80,
    )
    duration_ms: int = Field(0, ge=0, le=5000)
    text: str = Field("", max_length=128)
    key: Literal["", "Enter", "Backspace", "Tab", "Escape"] = ""
    delta_x: float = Field(0, ge=-2000, le=2000)
    delta_y: float = Field(0, ge=-2000, le=2000)


class CookieStatusIn(BaseModel):
    enabled: bool


class DefaultReplyIn(BaseModel):
    enabled: bool
    reply_content: Optional[str] = None
    reply_image_url: Optional[str] = None
    reply_once: bool = False


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    value: str
    description: Optional[str] = None


class SystemSettingsSectionIn(BaseModel):
    settings: Dict[str, Any] = Field(default_factory=dict)
    secret_actions: Dict[str, str] = Field(default_factory=dict)


class SystemSettingsVerifyIn(BaseModel):
    settings: Dict[str, Any] = Field(default_factory=dict)
    secret_actions: Dict[str, str] = Field(default_factory=dict)


class UserBasicSettingsIn(BaseModel):
    item_sync_enabled: Optional[bool] = None
    item_sync_interval: Optional[int] = Field(None, ge=60, le=86400)
    item_sync_max_pages: Optional[int] = Field(None, ge=1, le=50)


class SystemSettingCreateIn(BaseModel):
    key: str
    value: str
    description: Optional[str] = None


def _manual_cookie_identity(cookie_value: str) -> str:
    parsed = parse_cookie_string(cookie_value)
    if not has_core_session_cookies(parsed):
        raise HTTPException(
            status_code=400,
            detail={
                'code': 'invalid_cookie',
                'message': 'Cookie 必须包含 unb 和至少一个核心会话字段',
            },
        )
    return str(parsed.get('unb') or '').strip()


def _raise_account_identity_mismatch() -> None:
    raise HTTPException(
        status_code=409,
        detail={
            'code': 'account_identity_mismatch',
            'message': 'Cookie 中的闲鱼账号身份与当前账号不一致，未保存任何修改',
        },
    )


def _require_stable_cookie_identity(cookie_id: str, cookie_value: str) -> str:
    candidate_unb = _manual_cookie_identity(cookie_value)
    details = db_manager.get_cookie_details(cookie_id) or {}
    stable_unb = (
        str(details.get('xianyu_unb') or '').strip()
        or db_manager._extract_cookie_unb(str(details.get('value') or ''))
    )
    if stable_unb and candidate_unb != stable_unb:
        _raise_account_identity_mismatch()
    return candidate_unb





@accounts_router.get("/cookies")
def list_cookies(current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)
    return list(user_cookies.keys())


@accounts_router.get("/cookies/details")
def get_cookies_details(current_user: Dict[str, Any] = Depends(get_current_user)):
    """Return account capabilities without exposing Cookie or password material."""
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    result = []
    for cookie_id, cookie_value in user_cookies.items():
        cookie_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
        auto_confirm = db_manager.get_auto_confirm(cookie_id)
        auto_rate = db_manager.get_auto_rate_settings(cookie_id, user_id) or {}
        # 获取备注信息
        cookie_details = db_manager.get_cookie_details(cookie_id)
        remark = cookie_details.get('remark', '') if cookie_details else ''
        has_login_password = bool(cookie_details.get('password')) if cookie_details else False
        login_method = normalize_login_method(
            cookie_details.get('login_method') if cookie_details else 'unknown'
        )
        # 自动续期能力以真实的设备绑定 + 凭据为准（见 db_manager.get_cookie_refresh_settings）。
        # 旧的 supports_automatic_refresh 只表示“服务端 Playwright 密码续期”，已恒为关闭，
        # 不能用它回答“该账号能否自动续期”，否则前端会永远误报“不支持自动续期”。
        auto_refresh_supported = bool(
            db_manager.get_cookie_refresh_settings(cookie_id).get('auto_refresh_supported')
        )
        refresh_status = _current_session_refresh_status(cookie_id)
        reauth_required = refresh_status.get('state') == 'manual_reauth_required'
        search_context = db_manager.get_owned_cookie_search_context(user_id, cookie_id)
        search_state = str(search_context.get('state') or 'error')
        search_blocker_by_state = {
            'action_required': '账号身份不完整，请先完成登录恢复',
            'ownership_mismatch': '当前账号不属于此用户',
            'not_found': '账号记录不存在',
            'error': '账号身份状态读取失败，请稍后重试',
        }

        result.append({
            'id': cookie_id,
            'enabled': cookie_enabled,
            'auto_confirm': auto_confirm,
            'auto_rate_enabled': bool(auto_rate.get('enabled')),
            'auto_rate_enabled_at': auto_rate.get('enabled_at'),
            'auto_rate_pending_count': int(auto_rate.get('pending_count') or 0),
            'auto_rate_success_count': int(auto_rate.get('success_count') or 0),
            'auto_rate_failed_count': int(auto_rate.get('failed_count') or 0),
            'auto_rate_needs_reconcile_count': int(auto_rate.get('needs_reconcile_count') or 0),
            'remark': remark,
            'avatar_url': cookie_details.get('avatar_url', '') if cookie_details else '',
            'xianyu_nick': cookie_details.get('xianyu_nick', '') if cookie_details else '',
            'pause_duration': cookie_details.get('pause_duration', 10) if cookie_details else 10,
            'username': cookie_details.get('username', '') if cookie_details else '',
            'has_login_password': has_login_password,
            'login_credentials_valid': bool(
                cookie_details
                and cookie_details.get('password')
                and is_valid_account_login_username(cookie_details.get('username'))
            ),
            'show_browser': bool(cookie_details.get('show_browser')) if cookie_details else False,
            'cookie_refresh_enabled': bool(cookie_details.get('cookie_refresh_enabled')) if cookie_details else False,
            'cookie_refresh_interval_minutes': (
                cookie_details.get('cookie_refresh_interval_minutes', 1440) if cookie_details else 1440
            ),
            'login_method': login_method,
            'login_method_label': login_method_label(login_method),
            'auto_refresh_supported': auto_refresh_supported,
            'has_l3_memory': bool(
                cookie_details.get('has_l3_memory') if cookie_details else False
            ),
            'reauth_required': reauth_required,
            'reauth_action': reauth_action_for(login_method),
            'last_login_at': cookie_details.get('last_login_at') if cookie_details else None,
            'last_validated_at': cookie_details.get('last_validated_at') if cookie_details else None,
            'last_expired_at': cookie_details.get('last_expired_at') if cookie_details else None,
            'reauth_updated_at': refresh_status.get('updated_at') if reauth_required else None,
            'search_readiness': {
                'ready': search_state == 'ready',
                'state': search_state,
                'blockers': (
                    []
                    if search_state == 'ready'
                    else [search_blocker_by_state.get(
                        search_state,
                        '账号身份不完整，请先完成登录恢复',
                    )]
                ),
            },
        })
    return result


@accounts_router.post("/cookies")
def add_cookie(item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 添加cookie时绑定到当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager

        cookie_unb = _manual_cookie_identity(item.value)
        canonical_id = db_manager.find_cookie_id_by_unb(user_id, cookie_unb)
        target_id = canonical_id or cookie_unb

        log_with_user(
            'info',
            f"尝试添加 Cookie: {_masked_identifier(target_id)}, 当前用户ID: {user_id}",
            current_user,
        )

        # 检查cookie是否已存在且属于其他用户
        existing_cookies = db_manager.get_all_cookies()
        user_cookies = db_manager.get_all_cookies(user_id)
        if canonical_id is None and target_id in existing_cookies and target_id not in user_cookies:
            base_id = target_id
            suffix = 1
            while target_id in existing_cookies:
                target_id = f"{base_id}_{suffix}"
                suffix += 1

        # 保存到数据库时指定用户ID
        if not db_manager.save_cookie(
            target_id, item.value, user_id, login_method='manual_cookie'
        ):
            raise HTTPException(status_code=400, detail="Cookie 保存失败")

        # 添加到CookieManager，同时指定用户ID
        if target_id in cookie_manager.manager.cookies:
            cookie_manager.manager.update_cookie(target_id, item.value, save_to_db=False)
        else:
            cookie_manager.manager.add_cookie(target_id, item.value, user_id=user_id)
        log_with_user('info', f"Cookie 添加成功: {_masked_identifier(target_id)}", current_user)
        return {"msg": "success", "account_id": target_id, "matched_existing": bool(canonical_id)}
    except AccountIdentityMismatchError:
        _raise_account_identity_mismatch()
    except HTTPException:
        raise
    except Exception as e:
        log_with_user(
            'error',
            f"添加 Cookie 失败: {type(e).__name__}",
            current_user,
        )
        raise HTTPException(status_code=400, detail="Cookie 保存失败") from e


@accounts_router.post("/api/browser-extension/pairings")
def create_browser_extension_pairing(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Create a five-minute, owner-bound, single-use browser pairing."""
    try:
        status_info, pairing_token = browser_extension_pairings.create(current_user['user_id'])
    except PairingError as exc:
        raise HTTPException(status_code=exc.http_status, detail=str(exc)) from exc
    return {
        "success": True,
        "data": {
            **status_info,
            "pairing_token": pairing_token,
            # One release of loopback-only clients still reads this alias.
            "pairing_code": pairing_token,
            "import_url": PUBLIC_IMPORT_URL,
            "console_origin": PUBLIC_CONSOLE_ORIGIN,
            "local_import_url": "http://127.0.0.1:8091/api/browser-extension/import",
        },
    }


def _raise_client_browser_error(exc: ClientBrowserError) -> None:
    raise HTTPException(
        status_code=exc.http_status,
        detail={"code": exc.error_code, "message": str(exc)},
    ) from exc


@accounts_router.post("/api/client-browser/devices")
def register_client_browser_device(
    payload: ClientBrowserDeviceIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        device = db_manager.register_client_browser_device(
            user_id=current_user['user_id'],
            device_id=payload.device_id,
            browser_family=payload.browser_family,
            client_type=payload.client_type,
            display_name=payload.display_name,
            signing_public_jwk=payload.signing_public_jwk,
            encryption_public_jwk=payload.encryption_public_jwk,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": device}


@accounts_router.get("/api/client-browser/devices")
def list_client_browser_devices(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    return {
        "success": True,
        "data": db_manager.list_client_browser_devices(current_user['user_id']),
    }


@accounts_router.delete("/api/client-browser/devices/{device_id}")
def revoke_client_browser_device(
    device_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        revoked = db_manager.revoke_client_browser_device(
            user_id=current_user['user_id'],
            device_id=device_id,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    if not revoked:
        raise HTTPException(status_code=404, detail="当前设备不存在或已撤销")
    return {"success": True}


@accounts_router.post("/api/client-browser/sessions")
def create_client_browser_login_session(
    payload: ClientBrowserSessionIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        device = db_manager.get_client_browser_device(
            user_id=current_user['user_id'],
            device_id=payload.device_id,
        )
        if not device or device.get("revoked"):
            raise ClientBrowserError(
                "请先安装或刷新当前设备浏览器连接",
                error_code="client_device_missing",
                http_status=409,
            )
        if device.get("client_type") != payload.client_type:
            raise ClientBrowserError(
                "设备连接类型与登录入口不匹配",
                error_code="client_type_mismatch",
                http_status=409,
            )
        status_info = client_login_sessions.create(
            owner_user_id=current_user['user_id'],
            device_id=payload.device_id,
            mode=payload.mode,
            client_type=payload.client_type,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {
        "success": True,
        "data": status_info,
    }


@accounts_router.get("/api/client-browser/sessions/{session_id}")
def get_client_browser_login_session(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        status_info = client_login_sessions.get_for_owner(
            session_id, current_user['user_id']
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": status_info}


@accounts_router.post("/api/client-browser/sessions/{session_id}/confirm")
def confirm_client_browser_login_session(
    session_id: str,
    payload: ClientBrowserConfirmIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_owned_cookie(payload.account_id, current_user['user_id'])
    try:
        status_info = client_login_sessions.confirm(
            session_id=session_id,
            owner_user_id=current_user['user_id'],
            account_id=payload.account_id,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": status_info}


@accounts_router.post("/api/client-browser/sessions/{session_id}/cancel")
def cancel_client_browser_login_session(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        status_info = client_login_sessions.cancel(
            session_id, current_user['user_id']
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": status_info}


@accounts_router.post("/api/client-browser/challenges")
def create_client_browser_challenge(
    payload: ClientBrowserChallengeIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        device = db_manager.get_client_browser_device(
            user_id=current_user['user_id'],
            device_id=payload.device_id,
        )
        if not device or device.get("revoked"):
            raise ClientBrowserError(
                "当前设备连接不存在",
                error_code="client_device_missing",
                http_status=404,
            )
        challenge = device_challenges.create(
            device_id=payload.device_id,
            owner_user_id=current_user['user_id'],
            purpose=payload.purpose,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": challenge}


@accounts_router.post("/api/client-browser/sessions/{session_id}/challenge")
def create_client_browser_login_challenge(
    session_id: str,
    payload: ClientBrowserSessionAuthorizeIn,
):
    """Issue a proof challenge to the device bound to one login session."""
    try:
        record = client_login_sessions.get_for_device(
            session_id=session_id,
            device_id=payload.device_id,
            mode=payload.mode,
        )
        device = db_manager.get_client_browser_device(
            user_id=record.owner_user_id,
            device_id=payload.device_id,
        )
        if not device or device.get("revoked"):
            raise ClientBrowserError(
                "当前设备连接已失效",
                error_code="client_device_revoked",
                http_status=403,
            )
        if device.get("client_type") != record.client_type:
            raise ClientBrowserError(
                "设备连接类型与登录会话不匹配",
                error_code="client_type_mismatch",
                http_status=403,
            )
        challenge = device_challenges.create(
            device_id=payload.device_id,
            owner_user_id=record.owner_user_id,
            purpose="login_import",
        )
        client_login_sessions.mark_waiting_user(session_id)
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": challenge}


@accounts_router.post("/api/client-browser/import")
async def import_client_browser_login(
    payload: ClientBrowserLoginImportIn,
):
    """Accept one device-proved login only after real platform validation."""
    consumed = False
    try:
        record = client_login_sessions.get_for_device(
            session_id=payload.session_id,
            device_id=payload.device_id,
            mode=payload.mode,
        )
        device = db_manager.get_client_browser_device(
            user_id=record.owner_user_id,
            device_id=payload.device_id,
            include_public_keys=True,
        )
        if not device or device.get("revoked"):
            raise ClientBrowserError(
                "当前设备连接已失效",
                error_code="client_device_revoked",
                http_status=403,
            )
        if device.get("client_type") != record.client_type:
            raise ClientBrowserError(
                "设备连接类型与登录会话不匹配",
                error_code="client_type_mismatch",
                http_status=403,
            )
        binding = {
            "session_id": payload.session_id,
            "mode": payload.mode,
            "device_id": payload.device_id,
        }
        device_challenges.verify(
            challenge_id=payload.challenge_id,
            device_id=payload.device_id,
            purpose="login_import",
            public_jwk=device["signing_public_jwk"],
            signature=payload.signature,
            binding=binding,
            owner_user_id=record.owner_user_id,
        )
        imported_cookies = normalize_structured_cookies([
            cookie.model_dump() for cookie in payload.cookies
        ])
        imported_unb = str(imported_cookies.get("unb") or "").strip()
        probe = await probe_message_session_async(
            session_cookies_to_string(imported_cookies),
            payload.user_agent,
        )
        platform_unb = str((probe.cookies or {}).get("unb") or "").strip()
        if not probe.succeeded:
            if probe.status == PROBE_RETRYABLE_ERROR:
                raise ClientBrowserError(
                    "平台连接暂时异常，保持官方页面开启并自动重试",
                    error_code="session_probe_retryable",
                    http_status=503,
                )
            raise ClientBrowserError(
                probe.message or "平台未确认有效登录状态",
                error_code=probe.error_code or "session_validation_failed",
            )
        if not platform_unb or platform_unb != imported_unb:
            raise ClientBrowserError(
                "平台验证身份与当前设备登录结果不一致",
                error_code="account_mismatch",
            )
        record = client_login_sessions.consume_for_import(
            session_id=payload.session_id,
            device_id=payload.device_id,
            mode=payload.mode,
            client_type=device.get("client_type"),
        )
        consumed = True
        now = time.time()
        account_info = await _persist_validated_account_login(
            user_id=record.owner_user_id,
            cookies_str=session_cookies_to_string(probe.cookies),
            validated_unb=platform_unb,
            login_method={
                "qr": "chrome_extension",
                "sms": "sms_window",
                "password": "password",
            }[payload.mode],
            browser_user_agent=payload.user_agent,
            runtime_state={
                "current_token": probe.access_token,
                "last_token_refresh_time": now,
                "browser_user_agent": payload.user_agent,
                "cookie_refresh_anchor": now,
                "item_sync_anchor": now,
            },
        )
        db_manager.touch_client_browser_device(
            user_id=record.owner_user_id, device_id=payload.device_id
        )
        status_info = client_login_sessions.persisted(
            payload.session_id, account_id=account_info["account_id"]
        )
        return {"success": True, "data": status_info}
    except ClientBrowserError as exc:
        if consumed:
            client_login_sessions.fail(
                payload.session_id, message=str(exc), error_code=exc.error_code
            )
        _raise_client_browser_error(exc)
    except Exception as exc:
        if consumed:
            client_login_sessions.retryable(
                payload.session_id,
                message="账号保存暂时异常，保持官方页面开启并自动重试",
                error_code="cookie_persist_failed",
            )
        logger.error(f"当前设备登录导入失败: {type(exc).__name__}")
        raise HTTPException(
            status_code=503 if consumed else 400,
            detail={
                "code": (
                    "cookie_persist_failed" if consumed else "client_login_failed"
                ),
                "message": (
                    "账号保存暂时异常，保持官方页面开启并自动重试"
                    if consumed
                    else "当前设备登录处理失败"
                ),
            },
        ) from exc


@accounts_router.get("/api/browser-extension/pairings/{pairing_id}")
def get_browser_extension_pairing(
    pairing_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        return {
            "success": True,
            "data": browser_extension_pairings.get(pairing_id, current_user['user_id']),
        }
    except PairingError as exc:
        raise HTTPException(status_code=exc.http_status, detail=str(exc)) from exc


@accounts_router.post("/api/browser-extension/import")
async def import_browser_extension_cookies(
    payload: BrowserExtensionImportIn,
    request: Request,
):
    """Validate one single-use browser import without echoing sensitive material."""
    content_length = request.headers.get("content-length", "")
    try:
        parsed_content_length = int(content_length) if content_length else 0
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Content-Length 无效") from exc
    if parsed_content_length > 262_144:
        raise HTTPException(status_code=413, detail="导入请求体过大")
    if len(payload.cookies) > MAX_COOKIE_COUNT:
        raise HTTPException(status_code=413, detail="Cookie 数量超过限制")

    remote_host = request.client.host if request.client else ""
    locked = False
    try:
        protocol_version = int(payload.protocol_version or 1)
        if protocol_version == 1 and not _is_strict_loopback_request(request):
            raise PairingError(
                "旧版扩展导入仅接受本机回环请求",
                error_code="non_loopback_request",
                http_status=403,
            )
        pairing_secret = (
            payload.pairing_token
            if protocol_version == PAIRING_PROTOCOL_VERSION
            else payload.pairing_code
        )
        if not pairing_secret:
            raise PairingError(
                "配对凭据缺失",
                error_code="pairing_credential_missing",
                http_status=400,
            )
        record = browser_extension_pairings.begin_validation(
            payload.pairing_id,
            pairing_secret,
            protocol_version=protocol_version,
            remote_host=remote_host,
        )
        locked = True
        raw_records = [
            cookie.model_dump() if hasattr(cookie, "model_dump") else cookie.dict()
            for cookie in payload.cookies
        ]
        imported_cookies = normalize_structured_cookies(raw_records)
        imported_unb = str(imported_cookies.get("unb") or "").strip()
        probe_result = await probe_message_session_async(
            session_cookies_to_string(imported_cookies),
            payload.user_agent,
        )
        platform_unb = str((probe_result.cookies or {}).get("unb") or "").strip()

        if not probe_result.succeeded:
            retryable = probe_result.status == PROBE_RETRYABLE_ERROR
            probe_code = probe_result.error_code or (
                "session_probe_retryable" if retryable else "session_validation_failed"
            )
            error_code = "session_probe_retryable" if retryable else probe_code
            if retryable:
                browser_extension_pairings.restore_waiting(payload.pairing_id)
                locked = False
                raise PairingError(
                    f"平台状态检查出现临时异常（{probe_code}），未导入 Cookie，可重试",
                    error_code=error_code,
                    http_status=503,
                )
            raise PairingError(
                probe_result.message or f"平台未确认有效登录状态（{probe_code}）",
                error_code=error_code,
            )
        if not platform_unb or platform_unb != imported_unb:
            raise PairingError(
                "平台验证身份与导入 Cookie 不一致",
                error_code="account_mismatch",
            )

        now = time.time()
        account_info = await _persist_validated_account_login(
            user_id=record.owner_user_id,
            cookies_str=session_cookies_to_string(probe_result.cookies),
            validated_unb=platform_unb,
            login_method='chrome_extension',
            browser_user_agent=payload.user_agent,
            runtime_state={
                "current_token": probe_result.access_token,
                "last_token_refresh_time": now,
                "browser_user_agent": payload.user_agent,
                "cookie_refresh_anchor": now,
                "item_sync_anchor": now,
            },
        )
        safe_status = browser_extension_pairings.succeed(
            payload.pairing_id,
            account_id=account_info['account_id'],
        )
        locked = False
        return {
            "success": True,
            "status": safe_status['status'],
            "message": safe_status['message'],
            "data": safe_status,
        }
    except PairingError as exc:
        if locked and exc.error_code != "session_probe_retryable":
            browser_extension_pairings.fail(
                payload.pairing_id,
                message=str(exc),
                error_code=exc.error_code,
            )
        raise HTTPException(
            status_code=exc.http_status,
            detail={"code": exc.error_code, "message": str(exc)},
        ) from exc
    except Exception as exc:
        if locked:
            browser_extension_pairings.fail(
                payload.pairing_id,
                message="导入处理失败",
                error_code="import_failed",
            )
        logger.error(f"Chrome 扩展导入失败: {type(exc).__name__}")
        raise HTTPException(
            status_code=400,
            detail={"code": "import_failed", "message": "导入处理失败"},
        ) from exc


# ============ 带子路径的 /cookies/{cid}/xxx 路由必须在 /cookies/{cid} 之前定义 ============

class AccountLoginInfoUpdate(BaseModel):
    username: Optional[str] = None
    login_password: Optional[str] = None
    show_browser: Optional[bool] = None


class CookieRefreshSettingsUpdate(BaseModel):
    cookie_refresh_enabled: bool
    cookie_refresh_interval_minutes: int


class AccountRenewalBindingIn(BaseModel):
    login_session_id: str = Field(..., min_length=8, max_length=80)
    device_id: str = Field(..., min_length=16, max_length=80)
    username: str = Field(..., min_length=1, max_length=200)
    password: str = Field(..., min_length=1, max_length=500)
    authorized: Literal[True]
    authorized_at: float


@accounts_router.put("/cookies/{cid}/login-info")
def update_cookie_login_info(cid: str, update_data: AccountLoginInfoUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号登录信息（用户名、密码、是否显示浏览器）"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")
        if update_data.login_password:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "credential_authorization_required",
                    "message": "请在当前设备登录成功后再次授权保存密码",
                },
            )
        if update_data.show_browser is not None:
            raise HTTPException(
                status_code=403,
                detail="服务器浏览器设置仅在管理员本机运维入口中使用",
            )
        if update_data.username is not None and update_data.username and not is_valid_account_login_username(update_data.username):
            raise HTTPException(status_code=400, detail="闲鱼登录账号不能填写 API 地址，请填写手机号、邮箱或闲鱼登录名")

        # 使用现有的update_cookie_account_info方法更新登录信息
        success = db_manager.update_cookie_account_info(
            cid,
            username=update_data.username,
        )

        if success:
            return {"success": True, "message": "登录信息已更新"}
        else:
            raise HTTPException(status_code=500, detail="更新登录信息失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.post("/api/accounts/{cid}/renewal-binding")
def bind_account_renewal_device(
    cid: str,
    payload: AccountRenewalBindingIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_owned_cookie(cid, current_user['user_id'])
    if not payload.authorized:
        raise HTTPException(status_code=409, detail="需要明确授权后保存续期凭据")
    try:
        client_login_sessions.authorize_renewal(
            session_id=payload.login_session_id,
            owner_user_id=current_user['user_id'],
            device_id=payload.device_id,
            account_id=cid,
        )
        binding = db_manager.bind_account_renewal_device(
            user_id=current_user['user_id'],
            cookie_id=cid,
            device_id=payload.device_id,
            username=payload.username,
            password=payload.password,
            authorized_at=payload.authorized_at,
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": binding}


@accounts_router.post("/api/client-browser/renewal/challenge")
def create_client_renewal_challenge(payload: ClientBrowserChallengeIn):
    if payload.purpose not in {
        "renewal_claim", "renewal_complete", "renewal_action_required"
    }:
        raise HTTPException(status_code=400, detail="续期挑战用途无效")
    try:
        device = db_manager.find_active_client_browser_device(
            payload.device_id, include_public_keys=True
        )
        if not device:
            raise ClientBrowserError(
                "当前设备连接不存在",
                error_code="client_device_missing",
                http_status=404,
            )
        if device.get("client_type") != "extension":
            raise ClientBrowserError(
                "当前连接类型不支持扩展续期任务",
                error_code="renewal_client_type_unsupported",
                http_status=409,
            )
        challenge = device_challenges.create(
            device_id=payload.device_id,
            owner_user_id=device["user_id"],
            purpose=payload.purpose,
        )
        if payload.purpose == "renewal_claim":
            now = time.time()
            with db_manager.lock:
                due_accounts = db_manager.conn.execute(
                    """
                    SELECT c.id, c.user_id
                    FROM cookies AS c
                    JOIN account_renewal_bindings AS b
                      ON b.cookie_id = c.id AND b.user_id = c.user_id
                    WHERE b.device_id = ? AND b.revoked_at IS NULL
                      AND c.cookie_refresh_enabled = 1
                      AND NOT EXISTS (
                          SELECT 1 FROM client_renewal_tasks AS t
                          WHERE t.cookie_id = c.id AND t.state IN (
                              'pending', 'claimed', 'action_required', 'validating'
                          )
                      )
                      AND COALESCE((
                          SELECT COALESCE(s.last_success_at, s.last_attempt_at, 0)
                          FROM account_session_refresh_status AS s
                          WHERE s.cookie_id = c.id
                      ), 0) <= ? - 60 * CASE
                          WHEN c.cookie_refresh_interval_minutes BETWEEN 60 AND 10080
                          THEN c.cookie_refresh_interval_minutes
                          ELSE 1440
                      END
                    ORDER BY c.id
                    LIMIT 1
                    """,
                    (payload.device_id, now),
                ).fetchone()
            if due_accounts:
                try:
                    db_manager.create_client_renewal_task(
                        user_id=int(due_accounts[1]),
                        cookie_id=str(due_accounts[0]),
                        trigger="scheduled_client_device",
                        now=now,
                    )
                except ClientBrowserError as task_error:
                    if task_error.error_code != "renewal_task_exists":
                        raise
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": challenge}


def _verify_renewal_device_proof(
    payload: ClientRenewalProofIn,
    *,
    purpose: str,
    binding: Dict[str, Any],
) -> Dict[str, Any]:
    device = db_manager.find_active_client_browser_device(
        payload.device_id, include_public_keys=True
    )
    if not device:
        raise ClientBrowserError(
            "当前设备连接不存在",
            error_code="client_device_missing",
            http_status=404,
        )
    device_challenges.verify(
        challenge_id=payload.challenge_id,
        device_id=payload.device_id,
        purpose=purpose,
        public_jwk=device["signing_public_jwk"],
        signature=payload.signature,
        binding=binding,
        owner_user_id=device["user_id"],
    )
    return device


@accounts_router.post("/api/client-browser/renewal/claim")
def claim_client_renewal_task(payload: ClientRenewalProofIn):
    binding = {"device_id": payload.device_id, "operation": "claim_next"}
    try:
        device = _verify_renewal_device_proof(
            payload, purpose="renewal_claim", binding=binding
        )
        task = db_manager.claim_next_client_renewal_task(
            user_id=device["user_id"], device_id=payload.device_id
        )
        db_manager.touch_client_browser_device(
            user_id=device["user_id"], device_id=payload.device_id
        )
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)
    return {"success": True, "data": task}


@accounts_router.post("/api/client-browser/renewal/{task_id}/result")
async def complete_client_renewal_task(
    task_id: str,
    payload: ClientRenewalResultIn,
):
    try:
        device = db_manager.find_active_client_browser_device(
            payload.device_id, include_public_keys=True
        )
        if not device:
            raise ClientBrowserError(
                "当前设备连接不存在",
                error_code="client_device_missing",
                http_status=404,
            )
        task = db_manager.get_client_renewal_task(
            user_id=device["user_id"],
            device_id=payload.device_id,
            task_id=task_id,
        )
        if not task or task["state"] not in {"claimed", "action_required"}:
            raise ClientBrowserError(
                "续期任务状态无效",
                error_code="renewal_task_state_invalid",
                http_status=409,
            )
        purpose = (
            "renewal_action_required"
            if payload.outcome == "action_required"
            else "renewal_complete"
        )
        binding = {
            "device_id": payload.device_id,
            "task_id": task_id,
            "account_id": task["account_id"],
            "outcome": payload.outcome,
        }
        _verify_renewal_device_proof(payload, purpose=purpose, binding=binding)
        current_task_state = task["state"]
        if payload.outcome == "action_required":
            if current_task_state != "claimed":
                raise ClientBrowserError(
                    "续期任务已等待用户验证",
                    error_code="renewal_task_already_updated",
                    http_status=409,
                )
            updated = db_manager.set_client_renewal_task_state(
                user_id=device["user_id"], device_id=payload.device_id,
                task_id=task_id, expected_state=current_task_state,
                state="action_required",
                error_code=payload.error_code or "human_verification_required",
            )
            if not updated:
                raise ClientBrowserError(
                    "续期任务已更新",
                    error_code="renewal_task_already_updated",
                    http_status=409,
                )
            db_manager.update_account_session_refresh(
                task["account_id"],
                state="action_required",
                trigger="client_device_renewal",
                message="当前设备需要完成短信、滑块、人脸或其他官方验证",
                error_code=payload.error_code or "human_verification_required",
            )
            return {"success": True, "data": {"state": "action_required"}}
        if payload.outcome == "failed":
            updated = db_manager.set_client_renewal_task_state(
                user_id=device["user_id"], device_id=payload.device_id,
                task_id=task_id, expected_state=current_task_state, state="failed",
                error_code=payload.error_code or "client_renewal_failed",
            )
            if not updated:
                raise ClientBrowserError(
                    "续期任务已更新", error_code="renewal_task_already_updated", http_status=409
                )
            return {"success": True, "data": {"state": "failed"}}

        imported = normalize_structured_cookies([cookie.model_dump() for cookie in payload.cookies])
        expected_unb = str((db_manager.get_cookie_details(task["account_id"]) or {}).get("xianyu_unb") or "").strip()
        imported_unb = str(imported.get("unb") or "").strip()
        probe = await probe_message_session_async(
            session_cookies_to_string(imported), payload.user_agent
        )
        platform_unb = str((probe.cookies or {}).get("unb") or "").strip()
        if not probe.succeeded or not expected_unb or {expected_unb, imported_unb, platform_unb} != {expected_unb}:
            raise ClientBrowserError(
                "续期 Token 或账号身份验证失败",
                error_code=probe.error_code or "renewal_identity_mismatch",
            )
        account_info = await _persist_validated_account_login(
            user_id=device["user_id"],
            cookies_str=session_cookies_to_string(probe.cookies),
            validated_unb=platform_unb,
            login_method="password",
            browser_user_agent=payload.user_agent,
            runtime_state={
                "current_token": probe.access_token,
                "last_token_refresh_time": time.time(),
                "browser_user_agent": payload.user_agent,
            },
        )
        if account_info["account_id"] != task["account_id"]:
            raise ClientBrowserError(
                "续期账号落库结果不匹配", error_code="renewal_persist_mismatch"
            )
        updated = db_manager.set_client_renewal_task_state(
            user_id=device["user_id"], device_id=payload.device_id,
            task_id=task_id, expected_state=current_task_state, state="success",
        )
        if not updated:
            raise ClientBrowserError(
                "续期任务已更新", error_code="renewal_task_already_updated", http_status=409
            )
        return {"success": True, "data": {"state": "success"}}
    except ClientBrowserError as exc:
        _raise_client_browser_error(exc)


@accounts_router.put("/cookies/{cid}/cookie-refresh-settings")
def update_cookie_refresh_settings(
    cid: str,
    update_data: CookieRefreshSettingsUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """更新账号定时Cookie刷新设置"""
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_cookie_refresh_settings(
            cid,
            enabled=update_data.cookie_refresh_enabled,
            interval_minutes=update_data.cookie_refresh_interval_minutes,
        )
        if not success:
            raise HTTPException(status_code=500, detail="更新Cookie定时刷新设置失败")

        try:
            from XianyuAutoAsync import XianyuLive
            live_instance = XianyuLive.get_instance(cid)
            if live_instance:
                live_instance.configure_cookie_refresh(
                    update_data.cookie_refresh_enabled,
                    update_data.cookie_refresh_interval_minutes,
                )
        except Exception as sync_error:
            logger.warning(f"同步运行中Cookie刷新设置失败（数据库已保存）: {cid}, {sync_error}")

        return {"success": True, "message": "Cookie定时刷新设置已更新"}
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ 通用的 /cookies/{cid} 路由 ============

@accounts_router.put('/cookies/{cid}')
def update_cookie(cid: str, item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")
        _require_stable_cookie_identity(cid, item.value)

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None

        # 使用 update_cookie_account_info 更新（只更新cookie值，不覆盖其他字段）
        success = db_manager.update_cookie_account_info(
            cid,
            cookie_value=item.value,
            login_method='manual_cookie',
            login_validated=False,
        )

        if not success:
            raise HTTPException(status_code=400, detail="更新Cookie失败")

        # 只有当 cookie 值真的发生变化时才重启任务
        if item.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, item.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")

        return {'msg': 'updated'}
    except AccountIdentityMismatchError:
        _raise_account_identity_mismatch()
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新 Cookie 失败: {type(e).__name__}")
        raise HTTPException(status_code=400, detail="更新 Cookie 失败") from e


class CookieAccountInfo(BaseModel):
    """账号信息更新模型"""
    value: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    show_browser: Optional[bool] = None


@accounts_router.post("/cookie/{cid}/account-info")
def update_cookie_account_info(cid: str, info: CookieAccountInfo, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号信息（Cookie、用户名、密码、显示浏览器设置）"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")
        if info.value is not None:
            _require_stable_cookie_identity(cid, info.value)

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None

        # 更新数据库
        success = db_manager.update_cookie_account_info(
            cid,
            cookie_value=info.value,
            username=info.username,
            password=info.password,
            show_browser=info.show_browser,
            login_method='manual_cookie' if info.value is not None else None,
        )

        if not success:
            raise HTTPException(status_code=400, detail="更新账号信息失败")

        # 只有当 cookie 值真的发生变化时才重启任务
        if info.value is not None and info.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, info.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")

        return {'msg': 'updated', 'success': True}
    except AccountIdentityMismatchError:
        _raise_account_identity_mismatch()
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新账号信息失败: {type(e).__name__}")
        raise HTTPException(status_code=400, detail="更新账号信息失败") from e


@accounts_router.get("/cookie/{cid}/details")
def get_cookie_account_details(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号详细信息（包括用户名、密码、显示浏览器设置）"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取详细信息
        details = db_manager.get_cookie_details(cid)

        if not details:
            raise HTTPException(status_code=404, detail="账号不存在")

        safe_details = dict(details)
        safe_details['has_login_password'] = bool(safe_details.pop('password', ''))
        safe_details.pop('password_encrypted', None)
        safe_details.pop('value', None)
        safe_details.pop('browser_user_agent', None)
        safe_details['auto_refresh_supported'] = bool(
            db_manager.get_cookie_refresh_settings(cid).get('auto_refresh_supported')
        )
        safe_details['has_l3_memory'] = bool(details.get('has_l3_memory'))
        return safe_details
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取账号详情失败: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# ========================= 账号密码登录相关接口 =========================

async def _update_cookie_manager_after_official_login(
    account_id: str,
    cookies_str: str,
    user_id: int,
    *,
    is_new_account: bool,
    runtime_state: Optional[Dict[str, Any]] = None,
) -> None:
    """Apply one CookieManager mutation after the database transaction succeeds."""
    manager = cookie_manager.manager
    if manager is None:
        logger.warning(
            f"CookieManager 未初始化，账号 {_masked_identifier(account_id)} "
            "将在服务重启后启动监听"
        )
        return

    if is_new_account:
        operation = manager.add_cookie(
            account_id,
            cookies_str,
            user_id=user_id,
            runtime_state=runtime_state,
        )
    else:
        operation = manager.replace_cookie(
            account_id,
            cookies_str,
            save_to_db=False,
            runtime_state=runtime_state,
        )
    if asyncio.isfuture(operation) or asyncio.iscoroutine(operation):
        await operation


async def _persist_validated_account_login(
    *,
    user_id: int,
    cookies_str: str,
    validated_unb: str,
    login_method: str = 'unknown',
    browser_user_agent: str = "",
    runtime_state: Optional[Dict[str, Any]] = None,
    has_l3_memory: bool = False,
) -> Dict[str, Any]:
    """Persist one platform-validated identity and restart its listener once."""
    parsed = trans_cookies(cookies_str)
    cookie_unb = str(parsed.get("unb") or "").strip()
    expected_unb = str(validated_unb or "").strip()
    if not expected_unb or cookie_unb != expected_unb:
        raise ValueError("平台验证身份与 Cookie 身份不一致")

    canonical_account_id = await asyncio.to_thread(
        db_manager.find_cookie_id_by_unb,
        user_id,
        expected_unb,
    )
    account_id = canonical_account_id or expected_unb
    user_cookies = await asyncio.to_thread(db_manager.get_all_cookies, user_id)
    all_cookies = await asyncio.to_thread(db_manager.get_all_cookies)
    if canonical_account_id is None and account_id in all_cookies and account_id not in user_cookies:
        base_id = account_id
        suffix = 1
        while account_id in all_cookies:
            account_id = f"{base_id}_{suffix}"
            suffix += 1
    is_new_account = account_id not in user_cookies

    update_success = await asyncio.to_thread(
        db_manager.update_cookie_account_info,
        account_id,
        cookie_value=cookies_str,
        user_id=user_id,
        browser_user_agent=browser_user_agent or None,
        login_method=login_method,
        login_validated=True,
        has_l3_memory=True if has_l3_memory else None,
    )
    if not update_success:
        raise RuntimeError("已验证登录态保存失败")
    db_manager.update_account_session_refresh(
        account_id,
        state='success',
        trigger=f'{normalize_login_method(login_method)}_login',
        message='登录状态已验证',
    )

    await _update_cookie_manager_after_official_login(
        account_id,
        cookies_str,
        user_id,
        is_new_account=is_new_account,
        runtime_state=runtime_state,
    )
    return {
        "account_id": account_id,
        "is_new_account": is_new_account,
        "cookie_count": len(parsed),
    }


_official_login_completion_locks: Dict[tuple[int, int, str], asyncio.Lock] = {}


async def _complete_official_login_session(
    record: OfficialLoginSessionRecord,
    result: OfficialLoginResult,
    account: str,
    password: str,
) -> Dict[str, Any]:
    """Persist a successful official session once, then hand off its listener."""
    user_id = record.owner_user_id
    loop = asyncio.get_running_loop()
    completion_key = (id(loop), user_id, result.unb)
    completion_lock = _official_login_completion_locks.setdefault(
        completion_key,
        asyncio.Lock(),
    )
    async with completion_lock:
        canonical_account_id = await asyncio.to_thread(
            db_manager.find_cookie_id_by_unb,
            user_id,
            result.unb,
        )
        account_id = canonical_account_id or result.unb
        existing_cookies = await asyncio.to_thread(db_manager.get_all_cookies, user_id)
        is_new_account = account_id not in existing_cookies
        cookies_str = XianyuOfficialLoginService.cookies_to_string(result.cookies)

        update_kwargs: Dict[str, Any] = {
            "cookie_value": cookies_str,
            "user_id": user_id,
            "browser_user_agent": result.browser_user_agent,
            "login_method": (
                "password"
                if record.mode == "password"
                else "sms_window" if record.mode == "sms" else "qr"
            ),
            "login_validated": True,
            "has_l3_memory": True,
        }
        if record.mode == "password":
            update_kwargs.update({
                "username": account,
                "password": password,
                "show_browser": record.show_browser,
            })
        elif record.mode == "sms":
            update_kwargs.update({
                "username": account or None,
                "show_browser": record.show_browser,
            })
        update_success = await asyncio.to_thread(
            db_manager.update_cookie_account_info,
            account_id,
            **update_kwargs,
        )
        if not update_success:
            raise RuntimeError("官方登录成功，但保存账号信息失败")
        db_manager.update_account_session_refresh(
            account_id,
            state="success",
            trigger=f"{normalize_login_method(update_kwargs['login_method'])}_login",
            message="登录状态已验证",
        )

        official_login_coordinator.mark_restarting_listener(record)
        now = time.time()
        await _update_cookie_manager_after_official_login(
            account_id,
            cookies_str,
            user_id,
            is_new_account=is_new_account,
            runtime_state={
                "current_token": result.access_token,
                "last_token_refresh_time": now,
                "browser_user_agent": result.browser_user_agent,
                "cookie_refresh_anchor": now,
                "item_sync_anchor": now,
            },
        )
    logger.info(
        f"闲鱼官方登录成功，账号已按真实 unb 保存: {_masked_identifier(account_id)}, "
        f"cookie_fields={len(result.cookies)}"
    )
    return {"account_id": account_id, "is_new_account": is_new_account}


official_login_coordinator = OfficialLoginSessionCoordinator(
    completion_handler=_complete_official_login_session,
)


def _private_verification_image_response(path: str) -> FileResponse:
    resolved = resolve_private_verification_image(path)
    if resolved is None:
        raise HTTPException(status_code=404, detail="验证图片不存在或已过期")
    return FileResponse(
        str(resolved),
        headers={
            "Cache-Control": "no-store, private",
            "Pragma": "no-cache",
            "X-Content-Type-Options": "nosniff",
        },
    )


def _private_browser_frame_response(frame: bytes) -> Response:
    return Response(
        content=frame,
        media_type="image/png",
        headers={
            "Cache-Control": "no-store, private",
            "Pragma": "no-cache",
            "X-Content-Type-Options": "nosniff",
        },
    )


def _raise_browser_interaction_error(exc: Exception) -> None:
    if isinstance(exc, PermissionError):
        raise HTTPException(
            status_code=403,
            detail="无权限操作该登录会话",
        ) from exc
    if isinstance(exc, KeyError):
        raise HTTPException(
            status_code=404,
            detail="登录会话不存在或已过期",
        ) from exc
    if isinstance(exc, (InteractionRateLimited, InteractionQueueFull)):
        raise HTTPException(status_code=429, detail=str(exc)) from exc
    if isinstance(exc, InteractionValidationError):
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if isinstance(
        exc,
        (InteractionUnavailable, StaleFrameRevision, RuntimeError),
    ):
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    raise HTTPException(status_code=400, detail="浏览器操作无效") from exc


def _expected_owned_unb(user_id: int, account: str) -> str:
    value = str(account or "").strip()
    if not value:
        return ""
    for cookie_id in db_manager.get_all_cookies(user_id):
        details = db_manager.get_cookie_details(cookie_id) or {}
        if value in {
            str(cookie_id),
            str(details.get("xianyu_unb") or "").strip(),
            str(details.get("username") or "").strip(),
        }:
            return str(details.get("xianyu_unb") or "").strip()
    return ""


@accounts_router.post("/api/official-login/sessions")
async def create_official_login_session(
    request: Dict[str, Any],
    http_request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    mode = str(request.get("mode") or "qr").strip().lower()
    account = str(request.get("account") or "").strip()
    show_browser = bool(request.get("show_browser", False))
    # 服务端 Playwright 浏览器即使窗口隐藏也运行在本机，因此每次调用（不只是
    # show_browser=True）都要求回环访问；非回环请求回落到网页二维码或扩展导入。
    if show_browser:
        _require_server_browser_access(http_request, current_user)
    elif not _has_server_browser_access(http_request, current_user):
        raise HTTPException(
            status_code=409,
            detail=_client_browser_required_detail(),
        )
    try:
        session = await official_login_coordinator.start(
            owner_user_id=current_user["user_id"],
            mode=mode,
            account=account,
            expected_unb=(
                _expected_owned_unb(current_user["user_id"], account)
                if mode == "sms"
                else ""
            ),
            password=str(request.get("password") or ""),
            show_browser=show_browser,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"success": True, **session}


@accounts_router.get("/api/official-login/sessions/{session_id}")
async def get_official_login_session(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    session = await official_login_coordinator.get_status(session_id, current_user["user_id"])
    if session is not None:
        return session
    persisted = get_session_registry().get(session_id)
    if persisted and persisted.get("owner_user_id") != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="无权限访问该登录会话")
    if persisted and persisted.get("status") == "interrupted":
        return {
            "session_id": session_id,
            "state": "interrupted",
            "message": persisted.get("error_message") or "服务已重启，请重新发起登录",
            "error_code": persisted.get("error_code") or "interrupted",
        }
    raise HTTPException(status_code=404, detail="登录会话不存在或已过期")


@accounts_router.get("/api/official-login/sessions/{session_id}/image")
async def get_official_login_session_image(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    get_frame = getattr(
        official_login_coordinator,
        "get_interaction_frame",
        None,
    )
    if callable(get_frame):
        try:
            frame, _revision = await get_frame(
                session_id,
                current_user["user_id"],
            )
            return _private_browser_frame_response(frame)
        except PermissionError as exc:
            _raise_browser_interaction_error(exc)
        except (KeyError, RuntimeError):
            pass

    image_path = await official_login_coordinator.get_image_path(
        session_id,
        current_user["user_id"],
    )
    if image_path:
        return _private_verification_image_response(image_path)

    persisted = get_session_registry().get(session_id)
    if persisted and persisted.get("owner_user_id") != current_user["user_id"]:
        raise HTTPException(status_code=403, detail="无权限访问该登录会话")
    raise HTTPException(status_code=404, detail="验证图片不存在或已过期")


@accounts_router.post("/api/official-login/sessions/{session_id}/interact")
async def interact_with_official_login_session(
    session_id: str,
    payload: BrowserInteractionIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        result = await official_login_coordinator.interact(
            session_id,
            current_user["user_id"],
            payload.model_dump(),
        )
    except Exception as exc:
        _raise_browser_interaction_error(exc)
    return {
        "success": True,
        "accepted": bool(result.get("accepted")),
        "frame_revision": int(result.get("frame_revision") or 0),
    }


@accounts_router.post("/api/official-login/sessions/{session_id}/show-browser")
async def show_official_login_browser(
    session_id: str,
    http_request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_server_browser_access(http_request, current_user)
    shown = await official_login_coordinator.show_browser(session_id, current_user["user_id"])
    if not shown:
        raise HTTPException(status_code=404, detail="登录会话不存在、已结束或不属于当前用户")
    return {"success": True, "message": "已请求在本机显示闲鱼官方浏览器"}


@accounts_router.post("/api/official-login/sessions/{session_id}/cancel")
async def cancel_official_login_session(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    cancelled = await official_login_coordinator.cancel(session_id, current_user["user_id"])
    if not cancelled:
        raise HTTPException(status_code=404, detail="登录会话不存在、已结束或不属于当前用户")
    return {"success": True, "message": "登录会话已取消"}


@accounts_router.post("/official-window-login")
async def official_window_login(
    payload: OfficialWindowLoginIn,
    http_request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Open installed Chrome and wait for SMS login on the official page."""
    _require_server_browser_access(http_request, current_user)
    account = str(payload.account or "").strip()
    try:
        session = await official_login_coordinator.start(
            owner_user_id=current_user['user_id'],
            mode="sms",
            account=account,
            expected_unb=_expected_owned_unb(current_user['user_id'], account),
            password="",
            show_browser=True,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="验证码登录参数无效") from exc
    return {
        "success": True,
        "session_id": session["session_id"],
        "status": "processing",
        "message": "已打开闲鱼官方登录窗口，请在窗口内完成验证码登录",
    }


@accounts_router.get("/official-window-login/check/{session_id}")
async def check_official_window_login(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    try:
        session = await get_official_login_session(session_id, current_user)
    except HTTPException as exc:
        if exc.status_code == 404:
            return {"status": "not_found", "message": "会话不存在或已过期"}
        raise
    state = session.get("state", "failed")
    status_value = {
        "preparing": "processing",
        "waiting_user": "processing",
        "verification_required": "processing",
        "persisting": "processing",
        "restarting_listener": "processing",
        "expired": "timeout",
    }.get(state, state)
    response = {
        "status": status_value,
        "message": session.get("message") or "请在官方窗口完成验证码登录",
        "error_code": session.get("error_code") or "",
    }
    if status_value == "success":
        response.update({
            "account_id": session.get("account_id") or "",
            "is_new_account": bool(session.get("is_new_account")),
        })
    return response


@accounts_router.post("/official-window-login/cancel/{session_id}")
async def cancel_official_window_login(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    cancelled = await official_login_coordinator.cancel(
        session_id, current_user["user_id"]
    )
    if not cancelled:
        return {"success": True, "message": "登录会话已经结束"}
    return {"success": True, "message": "手机号验证码登录已取消"}


@accounts_router.post("/password-login")
async def password_login(
    request: Dict[str, Any],
    http_request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """兼容旧客户端的账号密码登录入口。"""
    try:
        account = request.get('account')
        password = request.get('password')
        show_browser = request.get('show_browser', False)

        if not _has_server_browser_access(http_request, current_user):
            return {
                'success': False,
                'error_code': 'client_browser_required',
                'message': '请在当前设备的 Chrome 或 Edge 中继续账号密码登录',
                'action': 'open_client_browser',
            }

        if not account or not password:
            return {'success': False, 'message': '登录账号和密码不能为空'}

        session = await official_login_coordinator.start(
            owner_user_id=current_user['user_id'],
            mode="password",
            account=str(account),
            password=str(password),
            show_browser=bool(show_browser),
        )

        return {
            'success': True,
            'session_id': session['session_id'],
            'status': 'processing',
            'message': session['message'],
        }

    except Exception as e:
        safe_error = sanitize_runtime_error(e)
        log_with_user('error', f"账号密码登录异常: {safe_error}", current_user)
        return {'success': False, 'message': f'登录失败: {safe_error}'}


@accounts_router.get("/password-login/check/{session_id}")
async def check_password_login_status(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """兼容旧客户端的账号密码登录状态。"""
    try:
        session = await get_official_login_session(session_id, current_user)
        state = session.get("state", "failed")
        status = {
            "preparing": "processing",
            "waiting_user": "processing",
            "persisting": "processing",
            "restarting_listener": "processing",
            "expired": "timeout",
        }.get(state, state)
        return {
            "status": status,
            "message": session.get("message", ""),
            "error": session.get("message", "") if status in {"failed", "timeout"} else "",
            "error_code": session.get("error_code", ""),
            "account_id": session.get("account_id", ""),
            "is_new_account": session.get("is_new_account", False),
            "screenshot_path": session.get("verification_image_url") or session.get("qr_image_url"),
        }

    except HTTPException:
        raise
    except Exception as e:
        safe_error = sanitize_runtime_error(e)
        log_with_user('error', f"检查账号密码登录状态异常: {safe_error}", current_user)
        return {'status': 'error', 'message': safe_error}


# ========================= 人脸验证截图相关接口 =========================

@accounts_router.get("/face-verification/screenshot/{account_id}")
async def get_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """Return the newest private verification image for one owned account."""
    _require_owned_cookie(account_id, current_user['user_id'])
    image_path = _latest_account_verification_image(account_id)
    if image_path is None:
        raise HTTPException(status_code=404, detail="验证图片不存在或已过期")
    return _private_verification_image_response(str(image_path))


@accounts_router.delete("/face-verification/screenshot/{account_id}")
async def delete_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """Delete private verification images for one owned account."""
    _require_owned_cookie(account_id, current_user['user_id'])
    image_paths = _account_verification_images(account_id)
    for image_path in image_paths:
        remove_verification_image(str(image_path))
    return {
        'success': True,
        'message': f'已删除 {len(image_paths)} 个验证截图',
        'deleted_count': len(image_paths),
    }


# ========================= 扫码登录相关接口 =========================


def _require_owned_qr_session(
    session_id: str,
    current_user: Dict[str, Any],
) -> tuple[Any, Dict[str, Any]]:
    registry = get_session_registry()
    registry.cleanup()
    persisted = registry.get(session_id)
    if not persisted or (
        persisted.get('expires_at')
        and float(persisted['expires_at']) <= time.time()
    ):
        raise HTTPException(status_code=404, detail='扫码会话不存在或已过期')
    if persisted.get('owner_user_id') != current_user['user_id']:
        raise HTTPException(status_code=403, detail='无权限访问该扫码会话')
    return registry, persisted


@accounts_router.post("/qr-login/generate")
async def generate_qr_code(current_user: Dict[str, Any] = Depends(get_current_user)):
    """Generate the default login QR without starting a browser."""
    try:
        result = await qr_login_manager.generate_qr_code()
        if result.get("success"):
            session_id = result["session_id"]
            get_session_registry().register(
                session_id,
                "qr_login",
                current_user['user_id'],
                status="processing",
                ttl_seconds=900,
                transient=qr_login_manager.sessions.get(session_id),
            )
            log_with_user('info', "接口二维码生成成功", current_user)
        return result
    except Exception as exc:
        logger.error(f"生成接口二维码失败: {type(exc).__name__}")
        return {
            'success': False,
            'error_code': 'qr_generation_failed',
            'retryable': True,
            'message': '生成二维码失败，请稍后重试',
        }


@accounts_router.get("/qr-login/verification-image/{session_id}")
async def get_qr_login_verification_image(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_owned_qr_session(session_id, current_user)
    get_frame = getattr(qr_login_manager, "get_interaction_frame", None)
    if callable(get_frame):
        frame_info = get_frame(session_id)
        if frame_info is not None:
            frame, _revision = frame_info
            return _private_browser_frame_response(frame)
    image_path = qr_login_manager.get_verification_image_path(session_id)
    if not image_path:
        raise HTTPException(status_code=404, detail="验证图片不存在或已过期")
    return _private_verification_image_response(image_path)


@accounts_router.post("/qr-login/interact/{session_id}")
async def interact_with_qr_login_session(
    session_id: str,
    payload: BrowserInteractionIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_owned_qr_session(session_id, current_user)
    try:
        qr_login_manager.submit_interaction(
            session_id,
            payload.model_dump(),
        )
    except Exception as exc:
        _raise_browser_interaction_error(exc)
    return {
        "success": True,
        "accepted": True,
        "frame_revision": payload.frame_revision,
    }


@accounts_router.get("/qr-login/check/{session_id}")
async def check_qr_code_status(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """Poll one owner-scoped API QR session and persist it once after validation."""
    try:
        registry, persisted = _require_owned_qr_session(session_id, current_user)
        if persisted.get('status') == 'interrupted':
            return {
                'status': 'interrupted',
                'message': persisted.get('error_message') or '服务已重启，请重新生成二维码',
            }

        cleanup_qr_check_records()
        completed = qr_check_processed.get(session_id)
        if completed and completed.get('processed'):
            return {
                'status': 'already_processed',
                'message': '该会话已处理完成',
                'account_info': completed.get('account_info'),
            }

        session_lock = qr_check_locks[session_id]
        if session_lock.locked():
            return {'status': 'processing', 'message': '正在处理中，请稍候'}

        async with session_lock:
            completed = qr_check_processed.get(session_id)
            if completed and completed.get('processed'):
                return {
                    'status': 'already_processed',
                    'message': '该会话已处理完成',
                    'account_info': completed.get('account_info'),
                }

            qr_login_manager.cleanup_expired_sessions()
            status_info = qr_login_manager.get_session_status(session_id)
            registry.update(
                session_id,
                status=status_info.get('status') or 'processing',
                error_code=(
                    'qr_login_error'
                    if status_info.get('status') in {'failed', 'error'}
                    else ''
                ),
                error_message=(
                    status_info.get('message', '')
                    if status_info.get('status') in {'failed', 'error'}
                    else ''
                ),
            )

            if status_info.get('status') == 'success':
                cookies_info = qr_login_manager.get_session_cookies(session_id)
                if not cookies_info:
                    return {
                        'status': 'error',
                        'message': '平台登录态尚未完成验证，请重新扫码',
                    }
                now = time.time()
                account_info = await process_qr_login_cookies(
                    cookies_info['cookies'],
                    cookies_info['unb'],
                    current_user,
                    runtime_state={
                        'browser_user_agent': detect_default_browser_user_agent(),
                        'cookie_refresh_anchor': now,
                        'item_sync_anchor': now,
                    },
                    has_l3_memory=bool(cookies_info.get('has_l3_memory')),
                )
                status_info['account_info'] = account_info
                qr_login_manager.mark_persisted(session_id)
                status_info['ended_by'] = 'validated_and_persisted'
                qr_check_processed[session_id] = {
                    'processed': True,
                    'timestamp': now,
                    'account_info': account_info,
                }
                qr_login_manager.remove_session(session_id)

            status_info.pop('cookies', None)
            status_info.pop('unb', None)
            return status_info
    except HTTPException:
        raise
    except Exception as exc:
        log_with_user(
            'error',
            f"检查扫码登录状态异常: {type(exc).__name__}",
            current_user,
        )
        return {'status': 'error', 'message': '扫码登录状态检查失败，请重新扫码'}


@accounts_router.post("/qr-login/continue/{session_id}")
async def continue_qr_code_after_verification(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """Start the dedicated Chrome window only after an explicit user action."""
    _require_owned_qr_session(session_id, current_user)
    qr_login_manager.continue_after_verification(session_id)
    return await check_qr_code_status(session_id, current_user)


@accounts_router.post("/qr-login/cancel/{session_id}")
async def cancel_qr_login_session(
    session_id: str,
    payload: QRLoginCancelIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    registry = get_session_registry()
    persisted = registry.get(session_id)
    if persisted and persisted.get('owner_user_id') != current_user['user_id']:
        raise HTTPException(status_code=403, detail='无权限访问该扫码会话')
    status_info = qr_login_manager.cancel_session(
        session_id,
        ended_by=payload.ended_by,
    )
    if status_info.get('status') == 'not_found':
        raise HTTPException(status_code=404, detail='二维码会话不存在或已过期')
    registry.update(
        session_id,
        status=status_info.get('status') or 'cancelled',
        error_code='cancelled',
        error_message=status_info.get('message') or '扫码登录已取消',
    )
    return status_info


async def process_qr_login_cookies(
    cookies: str,
    unb: str,
    current_user: Dict[str, Any],
    *,
    runtime_state: Optional[Dict[str, Any]] = None,
    has_l3_memory: bool = False,
) -> Dict[str, Any]:
    """Persist only a Cookie already validated by the API QR manager."""
    return await _persist_validated_account_login(
        user_id=current_user['user_id'],
        cookies_str=cookies,
        validated_unb=unb,
        login_method='qr',
        browser_user_agent=detect_default_browser_user_agent(),
        runtime_state=runtime_state,
        has_l3_memory=has_l3_memory,
    )


@accounts_router.put('/cookies/{cid}/status')
def update_cookie_status(cid: str, status_data: CookieStatusIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的启用/禁用状态"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.update_cookie_status(cid, status_data.enabled)
        return {'msg': 'status updated', 'enabled': status_data.enabled}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------- 默认回复管理接口 -------------------------

@content_router.get('/default-replies/{cid}')
def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        result = db_manager.get_default_reply(cid)
        if result is None:
            # 如果没有设置，返回默认值
            return {'enabled': False, 'reply_content': '', 'reply_once': False}
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.put('/default-replies/{cid}')
def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once, reply_data.reply_image_url)
        return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once, 'reply_image_url': reply_data.reply_image_url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.get('/default-replies')
def get_all_default_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的默认回复设置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_replies = db_manager.get_all_default_replies()
        # 过滤只属于当前用户的回复设置
        user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
        return user_replies
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.delete('/default-replies/{cid}')
def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_default_reply(cid)
        if success:
            return {'msg': 'default reply deleted'}
        else:
            raise HTTPException(status_code=400, detail='删除失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.post('/default-replies/{cid}/clear-records')
def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.clear_default_reply_records(cid)
        return {'msg': 'default reply records cleared'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 默认回复管理接口（单数形式兼容路由） -------------------------
# 兼容前端使用 /api/default-reply/ 的请求

@content_router.get('/api/default-reply/{cid}', deprecated=True)
def get_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置（兼容路由）"""
    return get_default_reply(cid, current_user)


@content_router.put('/api/default-reply/{cid}', deprecated=True)
def update_default_reply_compat(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置（兼容路由）"""
    return update_default_reply(cid, reply_data, current_user)


@content_router.delete('/api/default-reply/{cid}', deprecated=True)
def delete_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置（兼容路由）"""
    return delete_default_reply(cid, current_user)


@content_router.post('/api/default-reply/{cid}/clear-records', deprecated=True)
def clear_default_reply_records_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录（兼容路由）"""
    return clear_default_reply_records(cid, current_user)


# ------------------------- 通知渠道管理接口 -------------------------

@content_router.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知渠道"""
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建通知渠道"""
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@content_router.get('/notification-channels/{channel_id}')
def get_notification_channel(
    channel_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """获取指定通知渠道"""
    try:
        channel = db_manager.get_notification_channel(
            channel_id,
            current_user['user_id'],
        )
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.put('/notification-channels/{channel_id}')
def update_notification_channel(
    channel_id: int,
    channel_data: NotificationChannelUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """更新通知渠道"""
    try:
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled,
            current_user['user_id'],
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@content_router.delete('/notification-channels/{channel_id}')
def delete_notification_channel(
    channel_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """删除通知渠道"""
    try:
        success = db_manager.delete_notification_channel(
            channel_id,
            current_user['user_id'],
        )
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 消息通知配置接口 -------------------------

@content_router.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的消息通知配置"""
    try:
        # 只返回当前用户的消息通知配置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # 过滤只属于当前用户的通知配置
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的消息通知配置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """设置账号的消息通知"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查通知渠道是否存在
        channel = db_manager.get_notification_channel(
            notification_data.channel_id,
            user_id,
        )
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')

        success = db_manager.set_message_notification(
            cid,
            notification_data.channel_id,
            notification_data.enabled,
            user_id,
        )
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='设置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.delete('/message-notifications/account/{cid}')
def delete_account_notifications(
    cid: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """删除账号的所有消息通知配置"""
    try:
        success = db_manager.delete_account_notifications(
            cid,
            current_user['user_id'],
        )
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='账号通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.delete('/message-notifications/{notification_id}')
def delete_message_notification(
    notification_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """删除消息通知配置"""
    try:
        success = db_manager.delete_message_notification(
            notification_id,
            current_user['user_id'],
        )
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统设置接口 -------------------------

@settings_router.get('/system-settings/public')
def get_public_system_settings():
    """获取公开的系统设置（无需认证）"""
    try:
        all_settings = db_manager.get_all_system_settings()
        state = _registration_state()
        return {
            "registration_enabled": "true" if state['enabled'] else "false",
            "show_default_login_info": all_settings.get(
                "show_default_login_info", "false"
            ),
            "login_captcha_enabled": all_settings.get(
                "login_captcha_enabled", "true"
            ),
        }
    except Exception:
        logger.warning("获取公开系统设置失败")
        return {
            "registration_enabled": "false",
            "show_default_login_info": "false",
            "login_captcha_enabled": "true",
        }


@settings_router.get('/system-settings')
def get_system_settings(_: Dict[str, Any] = Depends(require_admin)):
    """获取类型化系统设置，不返回明文密钥。"""
    from db_manager import db_manager
    try:
        return normalize_system_settings(db_manager.get_all_system_settings())
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _prepare_settings_section(section: str, request: SystemSettingsSectionIn) -> Dict[str, Any]:
    if section not in SETTINGS_SECTION_KEYS:
        raise HTTPException(status_code=404, detail='配置分区不存在')
    unknown = set(request.settings) - SETTINGS_SECTION_KEYS[section]
    if unknown:
        raise HTTPException(status_code=400, detail=f"不支持的配置项: {', '.join(sorted(unknown))}")

    values = dict(request.settings)
    raw = db_manager.get_all_system_settings()
    secret_for_section = {'ai': 'ai_api_key', 'smtp': 'smtp_password'}.get(section)
    if secret_for_section:
        action = request.secret_actions.get(secret_for_section, 'keep')
        try:
            values[secret_for_section] = apply_secret_action(
                raw.get(secret_for_section, ''), action, str(values.get(secret_for_section, ''))
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    if section == 'basic':
        interval = int(values.get('item_sync_interval', raw.get('item_sync_interval', 600)) or 600)
        pages = int(values.get('item_sync_max_pages', raw.get('item_sync_max_pages', 5)) or 5)
        if not 60 <= interval <= 86400:
            raise HTTPException(status_code=400, detail='商品同步间隔必须在1分钟到24小时之间')
        if not 1 <= pages <= 50:
            raise HTTPException(status_code=400, detail='同步页数必须在1到50之间')
    elif section == 'ai':
        api_url = str(values.get('ai_api_url', raw.get('ai_api_url', '')) or '').strip()
        model = str(values.get('ai_model', raw.get('ai_model', '')) or '').strip()
        try:
            scheme, host, _port = parse_public_http_url(api_url)
        except OutboundRequestError as exc:
            raise HTTPException(status_code=400, detail='AI API地址格式无效') from exc
        if scheme != 'https':
            raise HTTPException(status_code=400, detail='AI API地址必须使用 HTTPS')
        if urlsplit(api_url).query:
            raise HTTPException(status_code=400, detail='AI API基础地址不能包含查询参数')
        if host == 'localhost' or host.endswith('.localhost'):
            raise HTTPException(status_code=400, detail='AI API地址必须使用公网主机')
        try:
            literal_host = ipaddress.ip_address(host)
        except ValueError:
            literal_host = None
        if literal_host is not None and not literal_host.is_global:
            raise HTTPException(status_code=400, detail='AI API地址必须使用公网主机')
        if not model:
            raise HTTPException(status_code=400, detail='AI模型不能为空')
    elif section == 'smtp':
        port = int(values.get('smtp_port', raw.get('smtp_port', 587)) or 587)
        if not 1 <= port <= 65535:
            raise HTTPException(status_code=400, detail='SMTP端口无效')
    return values


def _settings_summary() -> Dict[str, Any]:
    raw = db_manager.get_all_system_settings()
    settings = normalize_system_settings(raw)
    ai_configured = bool(raw.get('ai_api_url') and raw.get('ai_model') and raw.get('ai_api_key'))
    smtp_values = [raw.get('smtp_server'), raw.get('smtp_user'), raw.get('smtp_password')]
    smtp_status = smtp_configuration_status(raw, db_path=db_manager.db_path)
    smtp_configured = smtp_status['smtp_configured']
    smtp_verified = smtp_status['smtp_verified']
    smtp_partial = any(smtp_values) and not smtp_configured
    try:
        registration = _registration_state()
    except Exception:
        registration = {
            'enabled': False,
            'ready': False,
            'requested': False,
            'smtp_verified': False,
            'user_limit': 0,
            'user_count': 0,
            'remaining_slots': 0,
        }
    settings['smtp_verified'] = smtp_verified
    return {
        'settings': settings,
        'sections': {
            'basic': {'state': 'saved', 'label': '已保存', 'configured': True},
            'ai': {
                'state': 'ready' if ai_configured else 'missing',
                'label': '已配置' if ai_configured else '未配置',
                'configured': ai_configured,
                'model': settings.get('ai_model') or '',
            },
            'smtp': {
                'state': 'ready' if smtp_verified else ('warning' if smtp_configured or smtp_partial else 'missing'),
                'label': '已验证' if smtp_verified else ('待验证' if smtp_configured else ('配置不完整' if smtp_partial else '未配置')),
                'configured': smtp_configured,
                'verified': smtp_verified,
            },
        },
        'registration': registration,
        'runtime': {
            'cookie_manager': cookie_manager.manager is not None,
            'account_count': len(getattr(cookie_manager.manager, 'cookies', {}) or {}),
            'active_tasks': len(getattr(cookie_manager.manager, 'tasks', {}) or {}),
        },
    }


@settings_router.get('/api/settings/summary')
def get_settings_summary(_: Dict[str, Any] = Depends(require_admin)):
    return {'success': True, **_settings_summary()}


def _user_basic_settings_summary(user_id: int) -> Dict[str, Any]:
    resolved = resolve_user_basic_settings(
        db_manager.get_all_system_settings(),
        db_manager.get_user_settings(user_id),
    )
    return {"success": True, **resolved}


@settings_router.get('/api/settings/user-summary')
def get_user_settings_summary(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    return _user_basic_settings_summary(current_user['user_id'])


@settings_router.put('/api/settings/user-basic')
def save_user_basic_settings(
    request: UserBasicSettingsIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    values = (
        request.model_dump(exclude_none=True)
        if hasattr(request, "model_dump")
        else request.dict(exclude_none=True)
    )
    if not values:
        raise HTTPException(status_code=400, detail='至少提交一项个人设置')
    if not db_manager.set_user_settings(current_user['user_id'], values):
        raise HTTPException(status_code=500, detail='个人设置保存失败')
    return {
        **_user_basic_settings_summary(current_user['user_id']),
        "message": "个人同步设置已保存",
        "saved_at": datetime.now().isoformat(timespec='seconds'),
    }


@settings_router.put('/api/settings/sections/{section}')
def save_settings_section(section: str, request: SystemSettingsSectionIn,
                          _: Dict[str, Any] = Depends(require_admin)):
    values = _prepare_settings_section(section, request)
    if section == 'basic' and str(
        values.get('registration_enabled', '')
    ).strip().lower() in {'1', 'true', 'yes', 'on'}:
        try:
            ready = _registration_state()['ready']
        except Exception:
            ready = False
        if not ready:
            raise RegistrationError(
                "REGISTRATION_NOT_READY",
                "请先确认 SMTP、支持邮箱和注册容量",
                http_status=409,
            )
    if not db_manager.save_system_settings_section(values):
        raise HTTPException(status_code=500, detail='配置保存失败')
    return {
        'success': True,
        'message': '配置已保存',
        'saved_at': datetime.now().isoformat(timespec='seconds'),
        **_settings_summary(),
    }


@settings_router.post('/api/settings/verify/{section}')
def verify_settings_section(section: str, request: SystemSettingsVerifyIn,
                            _: Dict[str, Any] = Depends(require_admin)):
    if section not in {'ai', 'smtp'}:
        raise HTTPException(status_code=400, detail='该配置不需要连接验证')
    values = _prepare_settings_section(
        section,
        SystemSettingsSectionIn(settings=request.settings, secret_actions=request.secret_actions),
    )
    raw = db_manager.get_all_system_settings()
    effective = {**raw, **values}
    try:
        if section == 'ai':
            api_key = effective.get('ai_api_key') or ''
            if not api_key:
                raise ValueError('AI API Key未配置')
            base_url = str(effective.get('ai_api_url') or '').rstrip('/')
            body = {
                'model': effective.get('ai_model'),
                'messages': [{'role': 'user', 'content': '回复OK'}],
                'max_tokens': 8,
                'temperature': 0,
            }
            if (
                'deepseek' in str(effective.get('ai_model', '')).lower()
                or 'deepseek' in base_url.lower()
            ):
                body['thinking'] = {'type': 'disabled'}
            response = request_public_http_sync(
                'POST',
                f"{base_url}/chat/completions",
                headers={
                    'Authorization': f"Bearer {api_key}",
                    'Content-Type': 'application/json',
                },
                json_body=body,
                timeout_seconds=30,
                allowed_methods=('POST',),
                require_https=True,
            )
            response.raise_for_status()
            payload = response.json()
            choices = payload.get('choices') if isinstance(payload, dict) else None
            if not isinstance(choices, list) or not choices:
                raise ValueError('AI平台返回了无法识别的响应')
            return {'success': True, 'state': 'ready', 'message': 'AI连接可用'}
        recipient = normalize_email(
            str(effective.get('support_email') or '')
        ).normalized
        smtp_settings = {
            key: effective.get(key, '') for key in SMTP_CONFIGURATION_KEYS
        }
        if not db_manager.save_unverified_smtp_settings(smtp_settings):
            raise RegistrationError(
                "SMTP_VERIFICATION_SAVE_FAILED",
                "SMTP 待验证配置保存失败，请重试",
                http_status=503,
            )
        current = db_manager.get_all_system_settings()
        fingerprint = smtp_configuration_fingerprint(
            current,
            db_path=db_manager.db_path,
        )
        verification_code = f"{secrets.randbelow(1_000_000):06d}"
        SMTPEmailSender().send(
            current,
            recipient=recipient,
            subject='闲鱼监控台 SMTP 验证码',
            text=(
                f'您的 SMTP 验证码是 {verification_code}\n\n'
                '验证码在 10 分钟内有效，最多可尝试 5 次。'
            ),
        )
        challenge = db_manager.registration_service.create_challenge(
            purpose='smtp_verify_email',
            subject=recipient,
            context=fingerprint,
            secret=verification_code,
        )
        return {
            'success': True,
            'state': 'pending',
            'challenge_id': challenge['challenge_id'],
            'expires_in': 600,
            'masked_recipient': mask_email_for_log(recipient),
            'message': '验证邮件已发送',
        }
    except (SMTPConfigurationError, SMTPDeliveryError) as e:
        logger.warning(f"SMTP配置验证失败: {type(e).__name__}")
        raise RegistrationError(
            "SMTP_VERIFICATION_FAILED",
            "SMTP 验证邮件发送失败，请检查配置",
        ) from e
    except RegistrationError:
        raise
    except Exception as e:
        logger.warning(f"{section.upper()}配置验证失败: {type(e).__name__}")
        raise HTTPException(status_code=400, detail=f"验证失败: {str(e)}")


@settings_router.post('/api/settings/verify/smtp/confirm')
def confirm_smtp_verification(
    request: SMTPVerificationConfirmRequest,
    _: Dict[str, Any] = Depends(require_admin),
):
    verified_at = datetime.now().astimezone().isoformat(timespec='seconds')
    confirmation = db_manager.registration_service.confirm_smtp_verification(
        challenge_id=request.challenge_id,
        verification_code=request.verification_code,
        verified_at=verified_at,
    )
    return {
        'success': True,
        'state': 'ready',
        'verified_at': confirmation['verified_at'],
        'message': 'SMTP 配置已确认',
    }


@settings_router.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn,
                          _: Dict[str, Any] = Depends(require_admin)):
    """更新系统设置"""
    try:
        # 禁止直接修改密码哈希
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='请使用密码修改接口')
        if key in {'smtp_verified_fingerprint', 'smtp_verified_at'}:
            raise HTTPException(status_code=400, detail='该设置只能由 SMTP 验证流程更新')
        if key == 'registration_enabled' and setting_data.value.strip().lower() == 'true':
            if not _registration_state()['ready']:
                raise RegistrationError(
                    "REGISTRATION_NOT_READY",
                    "请先确认 SMTP、支持邮箱和注册容量",
                    http_status=409,
                )

        success = db_manager.set_system_setting(key, setting_data.value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except (HTTPException, RegistrationError):
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 注册设置接口 -------------------------


def _update_registration_enabled(
    enabled: bool,
    admin_user: Dict[str, Any],
) -> Dict[str, Any]:
    if enabled:
        try:
            ready = _registration_state()['ready']
        except Exception:
            ready = False
        if not ready:
            raise RegistrationError(
                "REGISTRATION_NOT_READY",
                "请先确认 SMTP、支持邮箱和注册容量",
                http_status=409,
            )
    if not db_manager.set_system_setting(
        'registration_enabled',
        'true' if enabled else 'false',
        '是否开启用户注册',
    ):
        raise RegistrationError(
            "REGISTRATION_SETTING_FAILED",
            "注册开关保存失败",
            http_status=503,
        )
    log_with_user(
        'info',
        f"更新注册设置: {'开启' if enabled else '关闭'}",
        admin_user,
    )
    return {
        'success': True,
        'enabled': enabled,
        'message': f"注册功能已{'开启' if enabled else '关闭'}",
    }


@admin_router.get('/api/admin/registration/status')
def get_registration_admin_status(
    _: Dict[str, Any] = Depends(require_admin),
):
    settings = db_manager.get_all_system_settings()
    state = _registration_state()
    support_email = str(
        settings.get('support_email') or settings.get('smtp_user') or ''
    ).strip()
    return {
        'success': True,
        'user_limit': state['user_limit'],
        'user_count': state['user_count'],
        'remaining_slots': state['remaining_slots'],
        'registration': {
            'enabled': state['enabled'],
            'ready': state['ready'],
            'requested': state['requested'],
            'terms_version': state['terms_version'],
        },
        'smtp': {
            'configured': state['smtp_configured'],
            'verified': state['smtp_verified'],
            'verified_at': settings.get('smtp_verified_at') or '',
            'support_email': mask_email_for_log(support_email)
            if support_email else '',
        },
    }


@admin_router.post('/api/admin/registration/invites')
def create_registration_invites(
    _request: Dict[str, Any] = Body(default_factory=dict),
    _: Dict[str, Any] = Depends(require_admin),
):
    raise RegistrationError(
        "INVITATION_REGISTRATION_REMOVED",
        "邀请注册已移除，请使用直接注册配置",
        http_status=410,
    )


@admin_router.get('/api/admin/registration/invites')
def list_registration_invites(
    _: Dict[str, Any] = Depends(require_admin),
):
    raise RegistrationError(
        "INVITATION_REGISTRATION_REMOVED",
        "邀请注册已移除，请使用直接注册配置",
        http_status=410,
    )


@admin_router.delete('/api/admin/registration/invites/{invite_id}')
def revoke_registration_invite(
    invite_id: int,
    _: Dict[str, Any] = Depends(require_admin),
):
    del invite_id
    raise RegistrationError(
        "INVITATION_REGISTRATION_REMOVED",
        "邀请注册已移除，请使用直接注册配置",
        http_status=410,
    )


@admin_router.put('/api/admin/registration/limit')
def update_registration_limit(
    request: RegistrationLimitUpdate,
    _: Dict[str, Any] = Depends(require_admin),
):
    capacity = db_manager.registration_service.update_registration_limit(
        request.limit
    )
    state = _registration_state()
    return {
        'success': True,
        **capacity,
        'enabled': state['enabled'],
        'requested': state['requested'],
        'message': '注册用户上限已更新',
    }


@admin_router.get('/api/admin/registration/users')
def list_registration_users(
    limit: int = Query(50, ge=1, le=200),
    _: Dict[str, Any] = Depends(require_admin),
):
    users = [
        user
        for user in db_manager.user_repository.list_recent(
            limit=min(200, limit + 1)
        )
        if str(user.get('username') or '').casefold()
        != ADMIN_USERNAME.casefold()
    ][:limit]
    return {
        'success': True,
        'users': [
            {
                key: user.get(key)
                for key in (
                    'id',
                    'username',
                    'email',
                    'is_active',
                    'created_at',
                    'terms_version',
                    'terms_accepted_at',
                )
            }
            for user in users
        ],
    }


@admin_router.put('/api/admin/registration/users/{user_id}')
def update_registration_user(
    user_id: int,
    request: UserActiveUpdate,
    _: Dict[str, Any] = Depends(require_admin),
):
    target = db_manager.get_user_by_id(user_id)
    if target and str(target.get('username') or '').casefold() == ADMIN_USERNAME.casefold():
        raise RegistrationError(
            "ADMIN_DEACTIVATION_FORBIDDEN",
            "管理员账号不能通过注册管理修改",
        )
    user = db_manager.auth_service.set_user_active(user_id, request.is_active)
    if not request.is_active:
        _drop_user_sessions_from_memory(user_id)
    return {'success': True, 'user': user}


@admin_router.put('/api/admin/registration/enabled')
def update_registration_enabled(
    request: RegistrationSettingUpdate,
    admin_user: Dict[str, Any] = Depends(require_admin),
):
    return _update_registration_enabled(request.enabled, admin_user)

@settings_router.get('/registration-status')
def get_registration_status():
    """兼容旧客户端的公开注册状态。"""
    try:
        state = _registration_state()
        return {
            'enabled': state['enabled'],
            'ready': state['ready'],
            'message': '注册功能已开启' if state['enabled'] else '注册暂未开放',
        }
    except Exception:
        logger.warning("获取注册状态失败")
        return {
            'enabled': False,
            'ready': False,
            'message': '注册暂未开放',
        }


@auth_router.get('/login-info-status')
def get_login_info_status():
    """获取默认登录信息显示状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('show_default_login_info')
        logger.debug(f"从数据库获取的登录信息显示设置值: '{enabled_str}'")

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录信息显示状态失败: {e}")
        # 出错时默认为开启
        return {"enabled": True}


class LoginInfoSettingUpdate(BaseModel):
    enabled: bool


@settings_router.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """兼容旧客户端的管理员注册开关。"""
    return _update_registration_enabled(setting_data.enabled, admin_user)

@auth_router.put('/login-info-settings')
def update_login_info_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新默认登录信息显示设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'show_default_login_info',
            'true' if enabled else 'false',
            '是否显示默认登录信息'
        )
        if success:
            log_with_user('info', f"更新登录信息显示设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"默认登录信息显示已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新登录信息显示设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新登录信息显示设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@accounts_router.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class AutoRateUpdate(BaseModel):
    auto_rate_enabled: bool


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@accounts_router.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_confirm设置
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动确认发货设置失败")

        # 通知CookieManager更新设置（如果账号正在运行）
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"自动确认发货已{'开启' if update_data.auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_confirm设置
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"自动确认发货当前{'开启' if auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.put("/cookies/{cid}/auto-rate")
def update_auto_rate(
    cid: str,
    update_data: AutoRateUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Enable seller reviews only for platform-rateable orders created afterward."""
    from db_manager import db_manager

    user_id = int(current_user['user_id'])
    if cid not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")
    if update_data.auto_rate_enabled:
        readiness = db_manager.get_owned_cookie_search_context(user_id, cid)
        if readiness.get('state') != 'ready':
            raise HTTPException(status_code=409, detail="账号身份未就绪，请先完成登录恢复")
    if not db_manager.update_auto_rate(cid, user_id, update_data.auto_rate_enabled):
        raise HTTPException(status_code=500, detail="更新自动好评设置失败")
    settings = db_manager.get_auto_rate_settings(cid, user_id) or {}
    return {
        "msg": "success",
        "auto_rate_enabled": bool(settings.get("enabled")),
        "auto_rate_enabled_at": settings.get("enabled_at"),
        "message": f"自动好评已{'开启' if update_data.auto_rate_enabled else '关闭'}",
    }


@accounts_router.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新备注
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"更新账号备注: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "备注更新成功",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="备注更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取Cookie详细信息（包含备注）
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "获取备注成功"
            }
        else:
            raise HTTPException(status_code=404, detail="账号不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证暂停时间范围（0-120分钟，0表示不暂停）
        if not (0 <= update_data.pause_duration <= 120):
            raise HTTPException(status_code=400, detail="暂停时间必须在0-120分钟之间（0表示不暂停）")

        # 更新暂停时间
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"更新账号自动回复暂停时间: {cid} -> {update_data.pause_duration}分钟", current_user)
            return {
                "message": "暂停时间更新成功",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="暂停时间更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@accounts_router.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取暂停时间
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "获取暂停时间成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@content_router.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@content_router.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@content_router.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@content_router.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        # 检查是否是图片关键词冲突
        if "已存在（图片关键词）" in error_msg:
            # 直接使用数据库管理器提供的友好错误信息
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            # 尝试从错误信息中提取具体的冲突关键词
            conflict_keyword = None
            conflict_type = None

            # 检查是否是数据库管理器抛出的详细错误
            if "关键词唯一约束冲突" in error_msg:
                # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                import re
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            # 构造用户友好的错误信息
            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@content_router.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的商品列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取该账号的所有商品
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, created_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY created_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or '未知商品',
                    'item_price': row[2] or '价格未知',
                    'created_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"获取商品列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@content_router.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")


@content_router.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件（解析是纯 CPU 阻塞，放到线程里跑，避免占住事件循环）
        contents = await file.read()
        df = await asyncio.to_thread(pd.read_excel, io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = await asyncio.to_thread(db_manager.get_keywords_with_type, cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        def clean_cell_value(value):
            """清理单元格值，处理数字转字符串时的 .0 后缀问题"""
            if pd.isna(value):
                return ''
            # 如果是数字类型，先转为整数（如果是整数值）再转字符串
            if isinstance(value, float) and value == int(value):
                return str(int(value)).strip()
            return str(value).strip()

        for index, row in df.iterrows():
            keyword = clean_cell_value(row['关键词'])
            item_id = clean_cell_value(row['商品ID']) or None
            reply = clean_cell_value(row['关键词内容'])

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        # 整批写入会长时间持有库锁，放到线程里执行，不阻塞事件循环
        success = await asyncio.to_thread(
            db_manager.save_text_keywords_only, cid, import_data
        )
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")


@content_router.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 分块读取，并在解码前执行硬大小门禁。
        image_data = await _read_upload_with_limit(
            image,
            max_bytes=IMAGE_UPLOAD_MAX_BYTES,
            label="图片文件",
        )
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")


@content_router.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """上传图片（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 分块读取，并在解码前执行硬大小门禁。
        image_data = await _read_upload_with_limit(
            image,
            max_bytes=IMAGE_UPLOAD_MAX_BYTES,
            label="图片文件",
        )
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")

        return {
            "message": "图片上传成功",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")


@content_router.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")


@content_router.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")


@content_router.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")


# 卡券管理API
@content_router.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的卡券列表"""
    try:
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新卡券"""
    try:
        user_id = current_user['user_id']
        card_name = str(card_data.get('name') or '').strip()
        if not card_name:
            raise HTTPException(status_code=400, detail="请输入资源名称")

        log_with_user('info', f"创建卡券: {card_name}", current_user)

        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec', False)
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_type = str(card_data.get('type') or '').strip()
        if card_type not in {'text', 'data', 'image', 'api'}:
            raise HTTPException(status_code=400, detail="资源类型无效")
        api_config = card_data.get('api_config')
        if card_type == 'api' and (
            not isinstance(api_config, dict)
            or api_config.get('protocol') != FULFILLMENT_API_PROTOCOL
        ):
            raise HTTPException(status_code=400, detail="新 API 资源只支持幂等 API v1")
        legacy_content = card_data.get('content')
        text_content = card_data.get('text_content')
        data_content = card_data.get('data_content')
        image_url = card_data.get('image_url')
        if legacy_content is not None:
            if card_type == 'text' and text_content is None:
                text_content = legacy_content
            elif card_type == 'data' and data_content is None:
                data_content = legacy_content
            elif card_type == 'image' and image_url is None:
                image_url = legacy_content

        if card_type == 'text':
            text_content = str(text_content or '').strip()
            if not text_content:
                raise HTTPException(status_code=400, detail="固定资料内容不能为空")
        elif card_type == 'data':
            raw_values = str(data_content or '').splitlines()
            if len(raw_values) > CARD_STOCK_IMPORT_MAX_ITEMS:
                raise HTTPException(status_code=400, detail="单批库存不能超过 10000 条")
            normalized_values: List[str] = []
            seen_values: set[str] = set()
            for raw_value in raw_values:
                value = str(raw_value or '').strip()
                if not value:
                    continue
                if len(value.encode('utf-8')) > CARD_STOCK_ITEM_MAX_BYTES:
                    raise HTTPException(status_code=400, detail="单条库存不能超过 2048 字节")
                if value not in seen_values:
                    seen_values.add(value)
                    normalized_values.append(value)
            if not normalized_values:
                raise HTTPException(status_code=400, detail="一次一密初始库存不能为空")
            data_content = '\n'.join(normalized_values)
        elif card_type == 'image':
            image_url = str(image_url or '').strip()
            if not image_url:
                raise HTTPException(status_code=400, detail="图片地址不能为空")
        elif card_type == 'api':
            api_token = str(card_data.get('api_token') or '').strip()
            if (
                not isinstance(api_config, dict)
                or not str(api_config.get('url') or '').lower().startswith('https://')
                or not api_token
            ):
                raise HTTPException(status_code=400, detail="请填写 HTTPS 地址和 API Token")

        card_id = db_manager.create_card(
            name=card_name,
            card_type=card_type,
            api_config=api_config,
            api_token=card_data.get('api_token'),
            text_content=text_content,
            data_content=data_content,
            image_url=image_url,
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            user_id=user_id,
            low_stock_threshold=card_data.get('low_stock_threshold', 5),
        )

        log_with_user('info', f"卡券创建成功: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "卡券创建成功"}
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        log_with_user('error', f"创建卡券失败: {card_data.get('name', '未知')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@content_router.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个卡券详情"""
    try:
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class CardStockImportRequest(BaseModel):
    format: Literal['lines', 'txt', 'csv'] = 'lines'
    content: str = Field(..., max_length=24 * 1024 * 1024)


class CardApiValidateRequest(BaseModel):
    api_token: Optional[str] = Field(default=None, max_length=8192)
    token: Optional[str] = Field(default=None, max_length=8192)


@content_router.post("/cards/{card_id}/stock/import")
def import_card_stock(
    card_id: int,
    payload: CardStockImportRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    values: List[str]
    if payload.format == 'csv':
        try:
            reader = csv.DictReader(io.StringIO(payload.content))
            if not reader.fieldnames or 'secret' not in reader.fieldnames:
                raise HTTPException(status_code=400, detail="CSV 必须包含 secret 列")
            values = [str(row.get('secret') or '') for row in reader]
        except csv.Error as exc:
            raise HTTPException(status_code=400, detail="CSV 内容格式无效") from exc
    else:
        values = payload.content.splitlines()
    try:
        return db_manager.import_card_stock(
            card_id,
            current_user['user_id'],
            values,
        )
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@content_router.post("/cards/{card_id}/api/validate")
def validate_card_api(
    card_id: int,
    payload: CardApiValidateRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = current_user['user_id']
    card = db_manager.get_card_by_id(card_id, user_id)
    if not card or card.get('type') != 'api':
        raise HTTPException(status_code=404, detail="API 资源不存在")
    runtime = db_manager.get_card_api_runtime_config(card_id, user_id) or {}
    supplied_token = payload.api_token if payload.api_token is not None else payload.token
    token = str(supplied_token if supplied_token is not None else runtime.get('api_token') or '').strip()
    if (
        runtime.get('protocol') != FULFILLMENT_API_PROTOCOL
        or not str(runtime.get('url') or '').lower().startswith('https://')
        or not token
    ):
        raise HTTPException(status_code=400, detail="请先填写 HTTPS 地址和 API Token")
    validation_key = hashlib.sha256(
        f"validate:{card_id}:{runtime.get('url')}:{token}".encode('utf-8')
    ).hexdigest()
    try:
        response = request_public_http_sync(
            'POST',
            runtime['url'],
            headers={
                'Authorization': f'Bearer {token}',
                'Idempotency-Key': validation_key,
            },
            json_body={
                'action': 'validate',
                'spec': runtime.get('spec') or {},
            },
            timeout_seconds=int(runtime.get('timeout') or 10),
            max_response_bytes=64 * 1024,
            allowed_methods=('POST',),
            require_https=True,
        )
        try:
            body = json.loads(response.text)
        except (TypeError, ValueError):
            body = None
        validated = (
            response.status == 200
            and isinstance(body, dict)
            and set(body) == {'status'}
            and body.get('status') == 'validated'
        )
        if not validated:
            db_manager.set_card_api_validation(card_id, user_id, 'failed')
            raise HTTPException(status_code=502, detail="供应方未通过幂等 API v1 校验")
        if not db_manager.set_card_api_validation(
            card_id,
            user_id,
            'validated',
            api_token=token,
        ):
            raise HTTPException(status_code=404, detail="API 资源不存在")
        return {"status": "validated", "message": "连接校验通过，可以绑定商品"}
    except HTTPException:
        raise
    except OutboundRequestError as exc:
        db_manager.set_card_api_validation(card_id, user_id, 'failed')
        raise HTTPException(status_code=502, detail="供应方连接校验失败") from exc


@content_router.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新卡券"""
    try:
        user_id = current_user['user_id']
        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec')
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_type = card_data.get('type')
        api_config = card_data.get('api_config')
        if card_type == 'api' and api_config is not None and (
            not isinstance(api_config, dict)
            or api_config.get('protocol') != FULFILLMENT_API_PROTOCOL
        ):
            raise HTTPException(status_code=400, detail="新 API 资源只支持幂等 API v1")
        for field_name, label in (
            ('text_content', '固定资料内容'),
            ('data_content', '一次一密库存'),
            ('image_url', '图片地址'),
        ):
            if field_name in card_data and not str(card_data.get(field_name) or '').strip():
                raise HTTPException(status_code=400, detail=f"{label}不能为空")
        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_type,
            api_config=api_config,
            api_token=card_data.get('api_token') if 'api_token' in card_data else None,
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled') if 'enabled' in card_data else None,
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value'),
            user_id=user_id,
            low_stock_threshold=card_data.get('low_stock_threshold'),
        )
        if success:
            return {"message": "卡券更新成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新带图片的卡券"""
    try:
        logger.info(f"接收到带图片的卡券更新请求: card_id={card_id}, name={name}, type={type}")

        # 必须先验证资源所有权，再读取 multipart 文件体。
        # 保留动态导入：该旧 multipart 路由的安全测试和插件调用方通过
        # db_manager 模块注入隔离数据库，普通 JSON 卡券路由则使用模块全局实例。
        from db_manager import db_manager
        existing_card = db_manager.get_card_by_id(
            card_id,
            current_user["user_id"],
        )
        if not existing_card:
            raise HTTPException(status_code=404, detail="卡券不存在")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 验证多规格字段
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        # 分块读取，并在解码前执行硬大小门禁。
        image_data = await _read_upload_with_limit(
            image,
            max_bytes=IMAGE_UPLOAD_MAX_BYTES,
            label="图片文件",
        )
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 更新卡券
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None,
            user_id=current_user['user_id']
        )

        if success:
            logger.info(f"卡券更新成功: {name} (ID: {card_id})")
            return {"message": "卡券更新成功", "image_url": image_url}
        else:
            # 如果数据库更新失败，删除已保存的图片
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="卡券不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新带图片的卡券失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 自动发货规则API
@content_router.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货规则列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "发货规则创建成功"}
    except ValueError as e:
        # 卡券归属校验失败（绑定了不存在或他人的卡券）
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个发货规则详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "发货规则更新成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except HTTPException:
        raise
    except ValueError as e:
        # 卡券归属校验失败（绑定了不存在或他人的卡券）
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.delete("/cards/{card_id}")
def delete_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除卡券"""
    try:
        user_id = current_user['user_id']
        blockers = db_manager.get_card_delete_blockers(card_id, user_id)
        if blockers == ['not_found']:
            raise HTTPException(status_code=404, detail="资源不存在")
        if blockers:
            raise HTTPException(
                status_code=409,
                detail="资源已有商品绑定或履约历史，请先停用并保留记录",
            )
        success = db_manager.delete_card(card_id, user_id=user_id)
        if success:
            return {"message": "资源删除成功"}
        else:
            raise HTTPException(status_code=409, detail="资源状态已变化，请刷新后重试")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@content_router.get("/fulfillment-records")
def get_fulfillment_records(
    state: Optional[Literal['all', 'succeeded', 'pending', 'failed', 'manual_review', 'ambiguous']] = None,
    limit: int = Query(100, ge=1, le=200),
    offset: int = Query(0, ge=0),
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    return db_manager.list_fulfillment_records(
        current_user['user_id'],
        state=state,
        limit=limit,
        offset=offset,
    )


@content_router.post("/fulfillment-records/{payload_id}/resend")
async def resend_fulfillment_record(
    payload_id: int,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = current_user['user_id']
    payload = db_manager.get_fulfillment_delivery_payload(payload_id, user_id)
    if not payload:
        raise HTTPException(status_code=404, detail="发货记录不存在")
    attempt = db_manager.get_fulfillment_attempt(payload['attempt_id'])
    if not attempt or int(attempt.get('user_id') or 0) != int(user_id):
        raise HTTPException(status_code=404, detail="发货记录不存在")
    if attempt.get('state') != 'committed':
        raise HTTPException(status_code=409, detail="只有已成功提交的原始内容可以重发")
    from XianyuAutoAsync import XianyuLive

    live = XianyuLive.get_instance(str(attempt.get('cookie_id') or ''))
    if not live or not live.ws or getattr(live.ws, 'closed', False):
        raise HTTPException(status_code=409, detail="账号当前不在线，请等待重连后再试")
    result = await live.resend_fulfillment_payload(
        payload_id=payload_id,
        user_id=user_id,
        database=db_manager,
    )
    if not result:
        raise HTTPException(status_code=409, detail="重发未完成，请查看记录状态")
    return result


@content_router.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "发货规则删除成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 备份和恢复 API ====================

@admin_router.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出用户备份"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # 导出当前用户的数据
        backup_data = db_manager.export_backup(user_id)

        # 生成文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # 返回JSON响应，设置下载头
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")


@admin_router.post("/backup/import")
async def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入用户备份"""
    try:
        # 验证文件类型
        if not str(file.filename or '').lower().endswith('.json'):
            raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

        # 有界读取，避免在解析前把任意大上传全部载入内存。
        content = file.file.read(USER_BACKUP_MAX_BYTES + 1)
        if len(content) > USER_BACKUP_MAX_BYTES:
            raise HTTPException(status_code=413, detail="备份文件超过25MB")
        backup_data = json.loads(content.decode('utf-8'))

        # 导入备份到当前用户
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            import cookie_manager

            manager = cookie_manager.manager
            if manager is not None:
                reconcile = await manager.reconcile_from_db()
                if not reconcile.get("success"):
                    logger.error(
                        "备份导入后运行态对账未完成: "
                        f"failed={reconcile.get('failed', 0)}"
                    )
                    raise HTTPException(
                        status_code=503,
                        detail="备份已写入，但运行态对账未完成，请人工检查",
                    )
                logger.info("备份导入后已完成 CookieManager 运行态对账")

            return {"message": "备份导入成功"}
        else:
            raise HTTPException(status_code=400, detail="备份导入失败")

    except HTTPException:
        raise
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise HTTPException(status_code=400, detail="备份文件格式无效")
    except Exception as e:
        logger.error(f"导入备份失败: error_type={type(e).__name__}")
        raise HTTPException(status_code=500, detail="导入备份失败") from e


@admin_router.post("/system/reload-cache")
def reload_cache(_: Dict[str, Any] = Depends(require_admin)):
    """重新加载系统缓存（用于手动刷新数据）"""
    try:
        import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "系统缓存已刷新", "success": True}
            else:
                raise HTTPException(status_code=500, detail="缓存刷新失败")
        else:
            raise HTTPException(status_code=500, detail="CookieManager 未初始化")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")


# ==================== 商品管理 API ====================

def _attach_item_knowledge_status(cookie_id: str, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """给商品列表附加知识档案状态，用于前端标识与复制目标提示。"""
    from db_manager import db_manager
    status_map = db_manager.get_ai_item_knowledge_status_by_cookie(cookie_id)
    for entry in items:
        info = status_map.get(str(entry.get('item_id') or ''))
        entry['knowledge_has_draft'] = bool(info and info.get('has_draft'))
        entry['knowledge_published_version'] = int(info.get('published_version') or 0) if info else 0
    return items


@content_router.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id, include_inactive=False)
            _attach_item_knowledge_status(cookie_id, items)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@accounts_router.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """检查当前用户是否有有效的cookies账户（必须是启用状态）"""
    try:
        # 匿名访问或管理器未就绪：一律返回 0，不泄露全站账号计数
        if current_user is None or cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        # 仅统计当前用户名下的账号，防止跨租户计数泄露
        all_cookies = db_manager.get_all_cookies(current_user['user_id'])

        # 检查启用状态和有效性
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # 检查是否启用
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # 检查是否有效（长度大于50）
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"检查cookies失败: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }


@content_router.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_items_by_cookie(cookie_id, include_inactive=False)
        _attach_item_knowledge_status(cookie_id, items)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@content_router.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="商品不存在")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_detail: str


@content_router.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_item_detail(cookie_id, item_id, update_data.item_detail)
        if success:
            return {"message": "商品详情更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品详情失败: {str(e)}")


@content_router.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "商品信息删除成功"}
        else:
            raise HTTPException(status_code=404, detail="商品信息不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    provider_profile_id: Optional[int] = None
    model_name: str = "deepseek-v4-flash"
    api_key: str = ""
    base_url: str = "https://api.deepseek.com"
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""
    api_key_action: str = "keep"
    provider_test_token: str = ""


class AIProviderProfileCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    provider_type: str = Field(default="openai_compatible", max_length=50)
    preset: str = Field(default="custom", max_length=50)
    base_url: str = Field(default="", max_length=2048)
    api_key: str = Field(default="", max_length=8192)
    default_model: str = Field(default="", max_length=AI_MODEL_NAME_MAX_LENGTH)
    models: List[str] = Field(default_factory=list, max_length=500)
    is_default: bool = False


class AIProviderProfileUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=100)
    provider_type: Optional[str] = Field(default=None, max_length=50)
    preset: Optional[str] = Field(default=None, max_length=50)
    base_url: Optional[str] = Field(default=None, max_length=2048)
    api_key: str = Field(default="", max_length=8192)
    api_key_action: str = Field(default="keep", max_length=20)
    default_model: Optional[str] = Field(
        default=None,
        max_length=AI_MODEL_NAME_MAX_LENGTH,
    )
    models: Optional[List[str]] = Field(default=None, max_length=500)
    is_default: Optional[bool] = None


class AIProviderModelDiscoveryRequest(BaseModel):
    profile_id: Optional[int] = Field(default=None, ge=1)
    provider_type: Optional[str] = Field(default=None, max_length=50)
    preset: Optional[str] = Field(default=None, max_length=50)
    base_url: Optional[str] = Field(default=None, max_length=2048)
    api_key: str = Field(default="", max_length=8192)


class AIProviderTestRequest(BaseModel):
    model_name: str = Field(
        ...,
        min_length=1,
        max_length=AI_MODEL_NAME_MAX_LENGTH,
    )


class AIReplyTestRequest(BaseModel):
    message: str = Field(default="你好", max_length=AI_MESSAGE_MAX_LENGTH)
    item_title: str = Field(
        default="测试商品",
        max_length=AI_ITEM_TITLE_MAX_LENGTH,
    )
    item_price: float = Field(default=100, ge=0, le=1_000_000_000)
    item_desc: str = Field(
        default="这是一个测试商品",
        max_length=AI_ITEM_DESCRIPTION_MAX_LENGTH,
    )


class AIReplyLabRequest(BaseModel):
    session_id: Optional[str] = Field(
        default=None,
        max_length=AI_SESSION_ID_MAX_LENGTH,
    )
    message: str = Field(..., min_length=1, max_length=AI_MESSAGE_MAX_LENGTH)
    item_id: Optional[str] = Field(
        default=None,
        max_length=AI_ITEM_ID_MAX_LENGTH,
    )
    item_title: str = Field(
        default="测试商品",
        max_length=AI_ITEM_TITLE_MAX_LENGTH,
    )
    item_price: float = Field(default=100, ge=0, le=1_000_000_000)
    item_desc: str = Field(
        default="这是一个测试商品",
        max_length=AI_ITEM_DESCRIPTION_MAX_LENGTH,
    )
    training_rules: List[Any] = Field(
        default_factory=list,
        max_length=AI_TRAINING_RULE_MAX_COUNT,
    )
    prompt_override: str = Field(
        default="",
        max_length=AI_PROMPT_OVERRIDE_MAX_LENGTH,
    )

    @field_validator("training_rules")
    @classmethod
    def validate_training_rules(cls, rules: List[Any]) -> List[Any]:
        for rule in rules:
            text = rule.get("text") if isinstance(rule, dict) else rule
            if len(str(text or "")) > AI_TRAINING_RULE_MAX_LENGTH:
                raise ValueError("训练规则过长")
            try:
                serialized = json.dumps(
                    rule,
                    ensure_ascii=False,
                    separators=(",", ":"),
                    default=str,
                ).encode("utf-8")
            except (TypeError, ValueError) as exc:
                raise ValueError("训练规则格式无效") from exc
            if len(serialized) > AI_TRAINING_RULE_SERIALIZED_MAX_BYTES:
                raise ValueError("训练规则负载过大")
        return rules


class AIReplyLabSaveRequest(BaseModel):
    item_id: str = Field(default="", max_length=AI_ITEM_ID_MAX_LENGTH)
    training_rules: List[Any] = Field(
        default_factory=list,
        max_length=AI_TRAINING_RULE_MAX_COUNT,
    )

    @field_validator("training_rules")
    @classmethod
    def validate_training_rules(cls, rules: List[Any]) -> List[Any]:
        return AIReplyLabRequest.validate_training_rules(rules)


class AITrainingRuleStatusRequest(BaseModel):
    enabled: bool


class AIItemKnowledgeDraftRequest(BaseModel):
    profile: Dict[str, Any] = Field(default_factory=dict)


class AIItemKnowledgeGenerateRequest(BaseModel):
    overview: str = ""
    # Backward-compatible input only; generation now replaces the draft.
    profile: Dict[str, Any] = Field(default_factory=dict)


class AIItemKnowledgeCopyRequest(BaseModel):
    target_item_ids: List[str] = Field(default_factory=list)
    # Kept for older clients. Copying always replaces target drafts.
    overwrite: bool = True


def _mask_secret(value: str) -> str:
    """Return a display-safe secret preview without exposing the stored value."""
    return mask_secret_preview(value)


@content_router.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """批量删除商品信息"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="删除列表不能为空")
        if len(request.items) > 500:
            raise HTTPException(status_code=400, detail="单次最多删除500个商品")

        normalized_items = []
        for item in request.items:
            if not isinstance(item, dict):
                raise HTTPException(status_code=400, detail="删除项格式无效")
            cookie_id = str(item.get('cookie_id') or '').strip()
            item_id = str(item.get('item_id') or '').strip()
            if not cookie_id or not item_id:
                raise HTTPException(status_code=400, detail="删除项缺少账号或商品ID")
            normalized_items.append({'cookie_id': cookie_id, 'item_id': item_id})

        user_cookies = db_manager.get_all_cookies(current_user['user_id'])
        if any(item['cookie_id'] not in user_cookies for item in normalized_items):
            raise HTTPException(status_code=403, detail="无权限删除其他账号的商品")

        success_count = db_manager.batch_delete_item_info(
            normalized_items,
            current_user['user_id'],
        )
        total_count = len(normalized_items)

        return {
            "message": f"批量删除完成",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"批量删除商品信息异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="批量删除商品失败")


# ==================== AI回复管理API ====================

AI_TRAINING_SECTION_TITLE = "AI训练修正"
AI_TRAINING_MARKER = f"【{AI_TRAINING_SECTION_TITLE}】"
AI_REPLY_RISK_PHRASES = [
    "登录我发给你的邮箱",
    "登录我发给您的邮箱",
    "登录我发的邮箱",
    "登录卖家的邮箱",
    "买邮箱",
    "买的是邮箱",
    "发密码",
    "验证码发我",
    "我帮你登录",
    "我帮您登录",
    "我帮你查询账号",
    "我帮您查询账号",
]


def _dedupe_rules(rules: List[str]) -> List[str]:
    seen = set()
    cleaned = []
    for rule in rules or []:
        text = str(rule or '').strip()
        if not text or text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return cleaned


def _normalize_scoped_rules(rules: List[Any], default_scope: str = 'item') -> List[Dict[str, str]]:
    seen = set()
    cleaned = []
    for rule in (rules or [])[:AI_TRAINING_RULE_MAX_COUNT]:
        if isinstance(rule, dict):
            scope = str(rule.get('scope') or default_scope).strip().lower()
            text = str(rule.get('text') or '').strip()
        else:
            scope = default_scope
            text = str(rule or '').strip()
        if (
            scope not in {'global', 'item'}
            or not text
            or len(text) > AI_TRAINING_RULE_MAX_LENGTH
        ):
            continue
        key = (scope, text)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append({'scope': scope, 'text': text})
    return cleaned


def _detect_ai_reply_warnings(reply: str) -> List[str]:
    return [phrase for phrase in AI_REPLY_RISK_PHRASES if phrase in (reply or "")]


def _extract_custom_prompt_text(raw: str) -> str:
    raw = (raw or '').strip()
    if not raw:
        return ""
    try:
        parsed = json.loads(raw)
    except Exception:
        return raw

    if isinstance(parsed, str):
        return parsed.strip()
    if isinstance(parsed, dict):
        for key in ('default', 'price', 'tech'):
            value = parsed.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return raw


def _merge_training_rules_into_prompt(raw_prompt: str, rules: List[str]) -> str:
    rules = _dedupe_rules(rules)
    base_text = _extract_custom_prompt_text(raw_prompt)
    existing_rules: List[str] = []

    if AI_TRAINING_MARKER in base_text:
        before, after = base_text.split(AI_TRAINING_MARKER, 1)
        base_text = before.strip()
        existing_rules = [
            line.strip().lstrip('-').strip()
            for line in after.splitlines()
            if line.strip().lstrip('-').strip()
        ]

    merged_rules = _dedupe_rules(existing_rules + rules)
    training_block = ""
    if merged_rules:
        training_block = f"\n\n{AI_TRAINING_MARKER}\n" + "\n".join([f"- {rule}" for rule in merged_rules])

    merged_text = f"{base_text}{training_block}".strip()
    return json.dumps({
        "default": merged_text,
        "price": merged_text,
        "tech": merged_text,
    }, ensure_ascii=False)


def _ensure_ai_cookie_access(cookie_id: str, current_user: Dict[str, Any]):
    user_id = current_user['user_id']
    user_cookies = db_manager.get_all_cookies(user_id)
    if cookie_id not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    if cookie_id not in cookie_manager.manager.cookies:
        raise HTTPException(status_code=404, detail='账号不存在')


def _get_ai_knowledge_item(cookie_id: str, item_id: str, current_user: Dict[str, Any]) -> Dict[str, Any]:
    _ensure_ai_cookie_access(cookie_id, current_user)
    item = db_manager.get_item_info(cookie_id, item_id)
    if not item:
        raise HTTPException(status_code=404, detail='当前账号中找不到这个商品，请先同步商品')
    return item


def _item_knowledge_source_hash(item: Dict[str, Any]) -> str:
    source = json.dumps({
        'title': item.get('item_title') or '',
        'price': item.get('item_price') or '',
        'detail': item.get('item_detail') or item.get('item_description') or '',
    }, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(source.encode('utf-8')).hexdigest()


def _item_knowledge_payload(cookie_id: str, item_id: str, item: Dict[str, Any]) -> Dict[str, Any]:
    profile = db_manager.get_ai_item_knowledge_profile(cookie_id, item_id)
    current_hash = _item_knowledge_source_hash(item)
    return {
        **profile,
        'item': {
            'item_id': item_id,
            'title': item.get('item_title') or '',
            'price': item.get('item_price') or '',
            'detail': item.get('item_detail') or item.get('item_description') or '',
            'updated_at': item.get('updated_at'),
        },
        'current_source_hash': current_hash,
        'source_changed': bool(profile.get('source_detail_hash') and profile.get('source_detail_hash') != current_hash),
    }


def _normalize_provider_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    preset = str(data.get('preset') or 'custom').strip().lower()
    if preset not in PROVIDER_PRESETS:
        preset = 'custom'
    preset_data = PROVIDER_PRESETS.get(preset, PROVIDER_PRESETS['custom'])
    provider_type = str(data.get('provider_type') or preset_data['provider_type']).strip()
    if provider_type not in {'openai_compatible', 'gemini'}:
        raise HTTPException(status_code=400, detail='平台类型仅支持 OpenAI 兼容接口或 Gemini')
    name = str(data.get('name') or preset_data['label']).strip()
    base_url = str(data.get('base_url') or preset_data['base_url']).strip().rstrip('/')
    default_model = str(data.get('default_model') or preset_data['default_model']).strip()
    if not name:
        raise HTTPException(status_code=400, detail='平台名称不能为空')
    if not base_url:
        raise HTTPException(status_code=400, detail='API 地址不能为空')
    try:
        scheme, _host, _port = parse_public_http_url(base_url)
    except OutboundRequestError as exc:
        raise HTTPException(status_code=400, detail='API 地址格式无效') from exc
    if scheme != 'https':
        raise HTTPException(status_code=400, detail='AI 平台 API 地址必须使用 HTTPS')
    if urlsplit(base_url).query:
        raise HTTPException(status_code=400, detail='AI 平台 API 基础地址不能包含查询参数')
    normalized = {
        **data,
        'name': name,
        'preset': preset,
        'provider_type': provider_type,
        'base_url': base_url,
        'default_model': default_model,
    }
    if 'models' in data:
        normalized['models'] = normalize_provider_models(data.get('models'))
    return normalized


def _provider_public_payload(profile: Dict[str, Any]) -> Dict[str, Any]:
    cached_at = profile.get('models_cached_at')
    return {
        **profile,
        'models_cache_fresh': bool(cached_at and time.time() - float(cached_at) < 86400),
    }


def _provider_discovery_profile(
    payload: AIProviderModelDiscoveryRequest,
    user_id: int,
) -> Dict[str, Any]:
    existing = None
    if payload.profile_id:
        existing = db_manager.get_ai_provider_profile(
            payload.profile_id,
            user_id,
            include_secret=True,
        )
        if not existing:
            raise HTTPException(status_code=404, detail='平台配置不存在')

    profile = _normalize_provider_payload({
        'name': existing['name'] if existing else '模型发现',
        'preset': payload.preset or (existing['preset'] if existing else 'custom'),
        'provider_type': payload.provider_type or (
            existing['provider_type'] if existing else 'openai_compatible'
        ),
        'base_url': payload.base_url or (existing['base_url'] if existing else ''),
        'default_model': existing['default_model'] if existing else '',
        'api_key': payload.api_key or (existing.get('api_key', '') if existing else ''),
    })
    if not profile.get('api_key'):
        raise HTTPException(status_code=400, detail='请先配置 API Key')
    return profile


@ai_router.get('/api/ai/providers')
def list_ai_providers(current_user: Dict[str, Any] = Depends(get_current_user)):
    user_id = current_user['user_id']
    db_manager.ensure_legacy_ai_provider_profiles(user_id)
    return {
        'providers': [_provider_public_payload(item) for item in db_manager.list_ai_provider_profiles(user_id)],
        'presets': PROVIDER_PRESETS,
    }


@ai_router.post('/api/ai/providers/discover-models')
def discover_ai_provider_models(
    payload: AIProviderModelDiscoveryRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = current_user['user_id']
    profile = _provider_discovery_profile(payload, user_id)
    try:
        models = _run_bounded_ai_call(
            user_id,
            lambda: discover_provider_models(profile),
        )
        return {'models': normalize_provider_models(models)}
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(
            f'平台模型发现失败 profile={payload.profile_id or "draft"}: '
            f'{type(e).__name__}'
        )
        raise HTTPException(status_code=400, detail='模型列表读取失败，请检查平台、Key 和地址')


@ai_router.post('/api/ai/providers')
def create_ai_provider(payload: AIProviderProfileCreate, current_user: Dict[str, Any] = Depends(get_current_user)):
    user_id = current_user['user_id']
    data = _normalize_provider_payload(payload.dict())
    if not data.get('api_key'):
        raise HTTPException(status_code=400, detail='新平台必须填写 API Key')
    if not db_manager.list_ai_provider_profiles(user_id):
        data['is_default'] = True
    try:
        profile_id = db_manager.create_ai_provider_profile(user_id, data)
    except Exception as e:
        if 'UNIQUE constraint failed' in str(e):
            raise HTTPException(status_code=409, detail='平台名称已存在')
        raise HTTPException(status_code=400, detail='平台配置创建失败')
    return _provider_public_payload(db_manager.get_ai_provider_profile(profile_id, user_id))


@ai_router.put('/api/ai/providers/{profile_id}')
def update_ai_provider(profile_id: int, payload: AIProviderProfileUpdate,
                       current_user: Dict[str, Any] = Depends(get_current_user)):
    user_id = current_user['user_id']
    current = db_manager.get_ai_provider_profile(profile_id, user_id)
    if not current:
        raise HTTPException(status_code=404, detail='平台配置不存在')
    merged = _normalize_provider_payload({**current, **payload.dict(exclude_none=True)})
    try:
        return _provider_public_payload(db_manager.update_ai_provider_profile(profile_id, user_id, merged))
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))


@ai_router.delete('/api/ai/providers/{profile_id}')
def delete_ai_provider(profile_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    try:
        if not db_manager.delete_ai_provider_profile(profile_id, current_user['user_id']):
            raise HTTPException(status_code=404, detail='平台配置不存在')
        return {'message': '平台配置已删除'}
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))


@ai_router.post('/api/ai/providers/{profile_id}/models/refresh')
def refresh_ai_provider_models(profile_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    user_id = current_user['user_id']
    profile = db_manager.get_ai_provider_profile(profile_id, user_id, include_secret=True)
    if not profile:
        raise HTTPException(status_code=404, detail='平台配置不存在')
    try:
        models = _run_bounded_ai_call(
            user_id,
            lambda: discover_provider_models(profile),
        )
        models = normalize_provider_models(models)
        if models:
            db_manager.update_ai_provider_models(profile_id, user_id, models)
            cached_at = time.time()
        else:
            models = normalize_provider_models(profile.get('models'))
            cached_at = profile.get('models_cached_at')
        return {'models': models, 'cached_at': cached_at}
    except HTTPException:
        raise
    except Exception as e:
        error_code = e.code if isinstance(e, OutboundRequestError) else type(e).__name__
        logger.warning(
            f'平台模型列表刷新失败 profile={profile_id}: error_code={error_code}'
        )
        raise HTTPException(status_code=400, detail='模型列表读取失败，可手动填写模型 ID 后测试')


@ai_router.post('/api/ai/providers/{profile_id}/test')
def test_ai_provider(profile_id: int, payload: AIProviderTestRequest,
                     current_user: Dict[str, Any] = Depends(get_current_user)):
    user_id = current_user['user_id']
    profile = db_manager.get_ai_provider_profile(profile_id, user_id, include_secret=True)
    if not profile:
        raise HTTPException(status_code=404, detail='平台配置不存在')
    model_name = payload.model_name.strip()
    try:
        reply = _run_bounded_ai_call(
            user_id,
            lambda: test_provider_reply(profile, model_name),
        )
        if not reply:
            raise ValueError('模型返回空内容')
        db_manager.update_ai_provider_verification(profile_id, user_id, 'verified', '测试回复生成成功')
        token = provider_test_tokens.issue(user_id, profile_id, model_name)
        return {
            'message': '测试回复生成成功，可以应用到账号',
            'reply': reply,
            'test_token': token,
            'model_name': model_name,
        }
    except HTTPException:
        raise
    except Exception as e:
        db_manager.update_ai_provider_verification(profile_id, user_id, 'failed', '测试回复生成失败')
        error_code = e.code if isinstance(e, OutboundRequestError) else type(e).__name__
        logger.warning(
            f'AI平台测试失败 profile={profile_id} model={model_name}: '
            f'error_code={error_code}'
        )
        raise HTTPException(status_code=400, detail='测试回复生成失败，请检查平台、Key、地址和模型 ID')


# ===== 高级回复策略的底层提示词存储（原技能中心 agent prompts）=====
SKILL_AGENT_PROMPT_TITLES = {
    'classify': '意图分类专家',
    'price': '议价专家',
    'tech': '技术专家',
    'default': '默认客服',
}


def _default_skill_agent_prompts() -> Dict[str, str]:
    prompts = {
        key: ai_reply_engine.default_prompts.get(key, '')
        for key in SKILL_AGENT_PROMPT_TITLES.keys()
    }
    prompts['classify'] = '''你是一个闲鱼客服意图分类专家。
只输出一个类别：price、tech、default。
price：砍价、优惠、包邮、最低价。
tech：参数、规格、功能、安装、兼容、使用方法。
default：其它售前、物流、售后和普通咨询。'''
    return prompts


def _ensure_skill_agent_prompts(user_id: int) -> Dict[str, Dict[str, Any]]:
    prompts = db_manager.get_skill_agent_prompts(user_id)
    defaults = _default_skill_agent_prompts()
    for prompt_type, content in defaults.items():
        if prompt_type not in prompts:
            db_manager.upsert_skill_agent_prompt(
                user_id=user_id,
                prompt_type=prompt_type,
                title=SKILL_AGENT_PROMPT_TITLES[prompt_type],
                content=content,
                enabled=True
            )
        elif prompt_type == 'classify' and '已不再被 detect_intent 使用' in prompts[prompt_type].get('content', ''):
            db_manager.upsert_skill_agent_prompt(
                user_id=user_id,
                prompt_type=prompt_type,
                title=SKILL_AGENT_PROMPT_TITLES[prompt_type],
                content=content,
                enabled=True
            )
    return db_manager.get_skill_agent_prompts(user_id)


# ===== 高级回复策略（用户级，跨账号共享）=====
# 归并自原“AI 专家客服”。底层复用 skill_agent_prompts 存储，仅暴露 price/tech/default，
# classify（意图分类）保留在库中但不通过本接口暴露。优先级低于产品事实/硬性价格规则/商品训练规则。
REPLY_STRATEGY_TYPES = ['price', 'tech', 'default']


class ReplyStrategyIn(BaseModel):
    content: str
    enabled: bool = True


class ReplyStrategiesUpdateIn(BaseModel):
    price: ReplyStrategyIn
    tech: ReplyStrategyIn
    default: ReplyStrategyIn

    model_config = {'extra': 'forbid'}


def _reply_strategy_payload(user_id: int) -> List[Dict[str, Any]]:
    prompts = _ensure_skill_agent_prompts(user_id)
    payload: List[Dict[str, Any]] = []
    for prompt_type in REPLY_STRATEGY_TYPES:
        item = prompts.get(prompt_type)
        if not item:
            continue
        payload.append({
            'prompt_type': prompt_type,
            'title': item.get('title') or SKILL_AGENT_PROMPT_TITLES.get(prompt_type, prompt_type),
            'content': item.get('content', ''),
            'enabled': bool(item.get('enabled', True)),
            'updated_at': item.get('updated_at'),
        })
    return payload


@ai_router.get('/api/ai/reply-strategies')
def get_ai_reply_strategies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取高级回复策略（price/tech/default，当前用户所有账号共享）"""
    return {
        'success': True,
        'data': _reply_strategy_payload(current_user['user_id']),
        'shared_scope': 'user',
        'note': '当前用户所有账号共享；优先级低于产品事实、硬性价格规则与商品训练规则',
    }


@ai_router.put('/api/ai/reply-strategies')
def update_ai_reply_strategies(
    payload: ReplyStrategiesUpdateIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Atomically replace price/tech/default reply strategies."""
    strategy_values = {
        prompt_type: getattr(payload, prompt_type)
        for prompt_type in REPLY_STRATEGY_TYPES
    }
    empty_types = [
        prompt_type
        for prompt_type, strategy in strategy_values.items()
        if not strategy.content.strip()
    ]
    if empty_types:
        raise HTTPException(status_code=400, detail="三类回复策略内容均不能为空")

    user_id = int(current_user['user_id'])
    _ensure_skill_agent_prompts(user_id)
    prompts = {
        prompt_type: {
            'title': SKILL_AGENT_PROMPT_TITLES[prompt_type],
            'content': strategy.content.strip(),
            'enabled': strategy.enabled,
        }
        for prompt_type, strategy in strategy_values.items()
    }
    if not db_manager.upsert_skill_agent_prompts_transaction(user_id, prompts):
        raise HTTPException(status_code=400, detail="保存回复策略失败，未修改任何策略")

    db_manager.log_skill_event(
        user_id,
        'agent',
        '批量更新高级回复策略',
        payload={'prompt_types': REPLY_STRATEGY_TYPES},
    )
    return {
        'success': True,
        'message': '三类回复策略已保存',
        'data': _reply_strategy_payload(user_id),
    }


@ai_router.put('/api/ai/reply-strategies/{prompt_type}')
def update_ai_reply_strategy(
    prompt_type: str,
    payload: ReplyStrategyIn,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """更新高级回复策略（仅 price/tech/default）"""
    if prompt_type not in REPLY_STRATEGY_TYPES:
        raise HTTPException(status_code=400, detail="不支持的回复策略类型")
    if not payload.content.strip():
        raise HTTPException(status_code=400, detail="策略内容不能为空")

    user_id = current_user['user_id']
    # 确保底层记录已初始化（保留已有的 8 条 prompt 记录，不重置）
    _ensure_skill_agent_prompts(user_id)
    success = db_manager.upsert_skill_agent_prompt(
        user_id,
        prompt_type,
        SKILL_AGENT_PROMPT_TITLES.get(prompt_type, prompt_type),
        payload.content,
        payload.enabled,
    )
    if not success:
        raise HTTPException(status_code=400, detail="保存回复策略失败")

    db_manager.log_skill_event(
        user_id,
        'agent',
        f"更新高级回复策略: {prompt_type}",
        payload={'prompt_type': prompt_type},
    )
    return {"success": True, "message": "回复策略已保存"}


@ai_router.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        db_manager.ensure_legacy_ai_provider_profiles(user_id)
        settings = db_manager.get_ai_reply_settings(cookie_id)
        account_api_key = ''
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute("SELECT api_key FROM ai_reply_settings WHERE cookie_id = ?", (cookie_id,))
            row = cursor.fetchone()
            account_api_key = row[0] if row and row[0] else ''

        system_api_key = db_manager.get_system_setting('ai_api_key') or ''
        profile = db_manager.get_ai_provider_profile(settings.get('provider_profile_id'), user_id)
        effective_key = settings.get('api_key') or account_api_key or system_api_key
        if profile:
            api_key_source = 'provider'
            api_key_masked = profile.get('api_key_masked', '')
        elif account_api_key:
            api_key_source = 'account'
            api_key_masked = _mask_secret(effective_key)
        elif system_api_key:
            api_key_source = 'global'
            api_key_masked = _mask_secret(effective_key)
        else:
            api_key_source = 'missing'
            api_key_masked = ''

        settings.update({
            'api_key': '',
            'api_key_source': api_key_source,
            'api_key_masked': api_key_masked,
            'has_effective_api_key': bool(effective_key),
        })
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取AI回复设置异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="获取AI回复设置失败")


@ai_router.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        db_manager.ensure_legacy_ai_provider_profiles(user_id)
        current_settings = db_manager.get_ai_reply_settings(cookie_id)
        requested_profile_id = settings.provider_profile_id or current_settings.get('provider_profile_id')
        if requested_profile_id is not None:
            profile = db_manager.get_ai_provider_profile(requested_profile_id, user_id)
            if not profile:
                raise HTTPException(status_code=404, detail='所选 AI 平台不存在')
            provider_changed = int(current_settings.get('provider_profile_id') or 0) != int(requested_profile_id)
            model_changed = str(current_settings.get('model_name') or '') != settings.model_name
            if provider_changed or model_changed:
                valid_test = provider_test_tokens.consume(
                    settings.provider_test_token, user_id, requested_profile_id, settings.model_name
                )
                if not valid_test:
                    raise HTTPException(status_code=409, detail='请先用所选平台和模型生成测试回复，成功后再应用')

        # 明确处理旧版账号专属Key：空值默认保留，只有clear才删除。
        settings_dict = settings.dict()
        settings_dict['provider_profile_id'] = requested_profile_id
        with db_manager.lock:
            row = db_manager.conn.execute(
                "SELECT api_key FROM ai_reply_settings WHERE cookie_id = ?", (cookie_id,)
            ).fetchone()
        existing_api_key = row[0] if row and row[0] else ''
        try:
            settings_dict['api_key'] = apply_secret_action(
                existing_api_key, settings.api_key_action, settings.api_key
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        settings_dict.pop('api_key_action', None)
        settings_dict.pop('provider_test_token', None)
        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:

            # 如果启用了AI回复，记录日志
            if settings.ai_enabled:
                logger.info(
                    f"{_ai_log_reference(cookie_id, 'account')} 启用AI回复"
                )
            else:
                logger.info(
                    f"{_ai_log_reference(cookie_id, 'account')} 禁用AI回复"
                )

            return {"message": "AI回复设置更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新AI回复设置异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="更新AI回复设置失败")


@ai_router.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的AI回复设置

    归属过滤在 SQL 层完成，设置解析一次性批量完成，循环内不再逐账号查库。
    """
    try:
        user_id = current_user['user_id']
        db_manager.ensure_legacy_ai_provider_profiles(user_id)
        return db_manager.get_ai_reply_settings_for_user(user_id)
    except Exception as e:
        logger.error(f"获取所有AI回复设置异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="获取AI回复设置失败")


@ai_router.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(
    cookie_id: str,
    test_data: AIReplyTestRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """测试AI回复功能"""
    try:
        _ensure_ai_cookie_access(cookie_id, current_user)
        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='账号不存在')

        # 检查是否启用AI回复
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')

        # 检查AI设置是否完整
        settings = db_manager.get_ai_reply_settings(cookie_id)
        if not settings.get('api_key'):
            raise HTTPException(status_code=400, detail='未配置API Key，请先在AI设置中配置API Key')
        if not settings.get('base_url'):
            raise HTTPException(status_code=400, detail='未配置API地址，请先在AI设置中配置API地址')

        # 构造测试数据
        test_message = test_data.message
        test_item_info = {
            'title': test_data.item_title,
            'price': test_data.item_price,
            'desc': test_data.item_desc,
        }

        # 生成测试回复（跳过等待时间）
        reply = _run_bounded_ai_call(
            current_user['user_id'],
            lambda: ai_reply_engine.generate_reply(
                message=test_message,
                item_info=test_item_info,
                chat_id=f"test_{int(time.time())}",
                cookie_id=cookie_id,
                user_id="test_user",
                item_id="test_item",
                skip_wait=True,
            ),
        )

        if reply:
            return {"message": "测试成功", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AI回复生成失败，请检查API Key是否正确、API地址是否可访问")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"测试AI回复异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="测试AI回复失败")


@ai_router.post("/ai-reply-lab/reply/{cookie_id}")
def ai_reply_lab_reply(cookie_id: str, request: AIReplyLabRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    """AI训练实验室回复，不污染正式对话记录和线上提示词。"""
    try:
        _ensure_ai_cookie_access(cookie_id, current_user)

        settings = db_manager.get_ai_reply_settings(cookie_id)
        if not settings.get('ai_enabled'):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')
        if not settings.get('api_key'):
            raise HTTPException(status_code=400, detail='未配置API Key，请先在AI设置中配置API Key')
        if not settings.get('base_url'):
            raise HTTPException(status_code=400, detail='未配置API地址，请先在AI设置中配置API地址')

        message = (request.message or '').strip()
        if not message:
            raise HTTPException(status_code=400, detail='买家消息不能为空')

        item_title = request.item_title or '测试商品'
        item_price = request.item_price if request.item_price not in (None, '') else 100
        item_desc = request.item_desc or '暂无商品描述'

        if request.item_id:
            db_item = db_manager.get_item_info(cookie_id, request.item_id)
            if not db_item:
                raise HTTPException(status_code=404, detail='当前账号中找不到这个商品，请先同步商品')
            item_title = db_item.get('item_title') or item_title
            item_price = db_item.get('item_price') or item_price
            item_desc = db_item.get('item_detail') or db_item.get('item_description') or item_desc

        current_time = time.time()
        _prune_ai_lab_sessions(
            current_time,
            user_id=current_user['user_id'],
        )

        session_id = request.session_id or secrets.token_urlsafe(16)
        registry = get_session_registry()
        persisted = registry.get(session_id)
        if persisted and persisted.get('owner_user_id') != current_user['user_id']:
            raise HTTPException(status_code=403, detail='无权限访问该训练会话')
        normalized_item_id = str(request.item_id or '')
        with _ai_lab_sessions_lock:
            session = ai_reply_lab_sessions.get(session_id)
            if (not session or session.get('cookie_id') != cookie_id
                    or session.get('user_id') != current_user['user_id']
                    or session.get('item_id') != normalized_item_id):
                session = {
                    'cookie_id': cookie_id,
                    'user_id': current_user['user_id'],
                    'item_id': normalized_item_id,
                    'history': [],
                    'timestamp': current_time,
                }
                ai_reply_lab_sessions[session_id] = session
            history = list(session.get('history', []))
        _prune_ai_lab_sessions(
            current_time,
            user_id=current_user['user_id'],
        )

        if not persisted and not registry.get(session_id):
            registry.register(
                session_id,
                "ai_training",
                current_user['user_id'],
                account_id=cookie_id,
                status="processing",
                ttl_seconds=6 * 3600,
                transient=session,
            )
        else:
            registry.update(session_id, status="processing", ttl_seconds=6 * 3600)

        reply_result = _run_bounded_ai_call(
            current_user['user_id'],
            lambda: ai_reply_engine.generate_lab_reply(
                message=message,
                item_info={
                    'title': item_title,
                    'price': item_price,
                    'desc': item_desc,
                },
                cookie_id=cookie_id,
                context=history,
                training_rules=_normalize_scoped_rules(request.training_rules),
                item_id=normalized_item_id,
                prompt_override=request.prompt_override,
                return_metadata=True,
            ),
        )

        if not reply_result or not reply_result.get('reply'):
            raise HTTPException(status_code=400, detail="AI回复生成失败，请检查API Key、API地址或训练规则")
        reply = reply_result['reply']

        history.extend([
            {'role': 'user', 'content': message},
            {'role': 'assistant', 'content': reply},
        ])
        with _ai_lab_sessions_lock:
            session['history'] = history[-24:]
            session['timestamp'] = current_time
            response_history = list(session['history'])
        registry.update(session_id, status="success", ttl_seconds=6 * 3600)

        return {
            "session_id": session_id,
            "reply": reply,
            "warnings": _detect_ai_reply_warnings(reply),
            "history": response_history,
            "rule_context": reply_result.get('rule_context', {}),
            "rule_audit": reply_result.get('audit', {}),
            "regenerated": bool(reply_result.get('regenerated')),
            "guarded_by_rule": bool(reply_result.get('guarded_by_rule')),
            "guard_reason": reply_result.get('guard_reason', ''),
            "guarded_rule_ids": reply_result.get('guarded_rule_ids', []),
            "knowledge_source": reply_result.get('knowledge_source', 'none'),
            "knowledge_version": reply_result.get('knowledge_version', 0),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI训练实验室回复异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="AI训练实验室回复失败")


@ai_router.post("/ai-reply-lab/save/{cookie_id}")
def save_ai_reply_lab_rules(cookie_id: str, request: AIReplyLabSaveRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    """兼容旧前端：训练规则写入分层规则表，不再污染账号提示词。"""
    try:
        _ensure_ai_cookie_access(cookie_id, current_user)
        rules = _normalize_scoped_rules(request.training_rules)
        if not rules:
            raise HTTPException(status_code=400, detail='没有可保存的训练规则')
        saved = db_manager.save_ai_training_rules(cookie_id, request.item_id, rules)

        return {
            "message": "训练规则已按范围保存",
            "rules": saved,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存AI训练规则异常: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="保存AI训练规则失败")


@ai_router.get("/ai-training-rules/{cookie_id}")
def get_ai_training_rules(cookie_id: str, item_id: str = Query(default=''), current_user: Dict[str, Any] = Depends(get_current_user)):
    _ensure_ai_cookie_access(cookie_id, current_user)
    rules = db_manager.get_ai_training_rules(cookie_id, item_id, include_disabled=True)
    return {**rules, 'context': db_manager.get_ai_training_rule_context(cookie_id, item_id)}


@ai_router.post("/ai-training-rules/{cookie_id}")
def save_ai_training_rules(cookie_id: str, request: AIReplyLabSaveRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    _ensure_ai_cookie_access(cookie_id, current_user)
    rules = _normalize_scoped_rules(request.training_rules)
    if not rules:
        raise HTTPException(status_code=400, detail='没有可保存的训练规则')
    saved = db_manager.save_ai_training_rules(cookie_id, request.item_id, rules)
    return {"message": "训练规则已保存", "rules": saved}


@ai_router.delete("/ai-training-rules/{cookie_id}/{rule_id}")
def delete_ai_training_rule(cookie_id: str, rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    _ensure_ai_cookie_access(cookie_id, current_user)
    if not db_manager.delete_ai_training_rule(cookie_id, rule_id):
        raise HTTPException(status_code=404, detail='训练规则不存在')
    return {"message": "训练规则已删除"}


@ai_router.patch("/ai-training-rules/{cookie_id}/{rule_id}")
def set_ai_training_rule_status(cookie_id: str, rule_id: int, request: AITrainingRuleStatusRequest,
                                current_user: Dict[str, Any] = Depends(get_current_user)):
    _ensure_ai_cookie_access(cookie_id, current_user)
    if not db_manager.set_ai_training_rule_enabled(cookie_id, rule_id, request.enabled):
        raise HTTPException(status_code=404, detail='训练规则不存在')
    return {"message": "训练规则状态已更新"}


@ai_router.get("/ai-item-knowledge/{cookie_id}/{item_id}")
def get_ai_item_knowledge(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    item = _get_ai_knowledge_item(cookie_id, item_id, current_user)
    return _item_knowledge_payload(cookie_id, item_id, item)


@ai_router.post("/ai-item-knowledge/{cookie_id}/{item_id}/generate")
def generate_ai_item_knowledge(cookie_id: str, item_id: str, request: AIItemKnowledgeGenerateRequest,
                               current_user: Dict[str, Any] = Depends(get_current_user)):
    item = _get_ai_knowledge_item(cookie_id, item_id, current_user)
    overview = str(request.overview or '').strip()
    if not overview:
        profile = request.profile if isinstance(request.profile, dict) else {}
        seed_overview = profile.get('overview') if isinstance(profile.get('overview'), dict) else {}
        overview = str(seed_overview.get('text') or '').strip()
    if not overview:
        raise HTTPException(status_code=400, detail='请先填写商品概览，再生成结构化草稿')
    seed = {'overview': {
        'text': overview,
        'source': 'user',
        'status': 'confirmed',
    }}
    source_hash = _item_knowledge_source_hash(item)
    try:
        generated = ai_reply_engine.generate_item_knowledge_draft({
            'title': item.get('item_title') or '',
            'price': item.get('item_price') or '',
            'desc': item.get('item_detail') or item.get('item_description') or '',
        }, cookie_id, seller_overview=overview)
        draft = ai_reply_engine.merge_generated_knowledge_with_seed(seed, generated)
        db_manager.save_ai_item_knowledge_draft(cookie_id, item_id, draft, source_hash)
        return {
            'message': '旧草稿已替换，新的AI结构化草稿已生成',
            'draft': draft,
            'source_detail_hash': source_hash,
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(
            f"生成商品知识草稿失败 "
            f"{_ai_log_reference(cookie_id, 'account')}/"
            f"{_ai_log_reference(item_id, 'item')}: {type(e).__name__}"
        )
        raise HTTPException(status_code=500, detail='AI草稿生成失败，请检查AI配置')


@ai_router.post("/ai-item-knowledge/{cookie_id}/{item_id}/copy")
def copy_ai_item_knowledge(cookie_id: str, item_id: str, request: AIItemKnowledgeCopyRequest,
                           current_user: Dict[str, Any] = Depends(get_current_user)):
    _get_ai_knowledge_item(cookie_id, item_id, current_user)
    if not request.target_item_ids:
        raise HTTPException(status_code=400, detail='请选择至少一个目标商品')
    try:
        result = db_manager.copy_ai_item_knowledge_draft(
            cookie_id, item_id, request.target_item_ids
        )
        return {
            **result,
            'message': f"已覆盖 {len(result['copied_item_ids'])} 个商品草稿",
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@ai_router.put("/ai-item-knowledge/{cookie_id}/{item_id}/draft")
def save_ai_item_knowledge_draft(cookie_id: str, item_id: str, request: AIItemKnowledgeDraftRequest,
                                 current_user: Dict[str, Any] = Depends(get_current_user)):
    item = _get_ai_knowledge_item(cookie_id, item_id, current_user)
    try:
        db_manager.save_ai_item_knowledge_draft(
            cookie_id,
            item_id,
            request.profile,
            _item_knowledge_source_hash(item),
        )
        return {
            "message": "商品知识草稿已保存",
            **_item_knowledge_payload(cookie_id, item_id, item),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@ai_router.post("/ai-item-knowledge/{cookie_id}/{item_id}/publish")
def publish_ai_item_knowledge(cookie_id: str, item_id: str,
                              current_user: Dict[str, Any] = Depends(get_current_user)):
    item = _get_ai_knowledge_item(cookie_id, item_id, current_user)
    try:
        profile = db_manager.publish_ai_item_knowledge(cookie_id, item_id)
        return {
            "message": f"商品知识第 {profile['version']} 版已发布",
            **_item_knowledge_payload(cookie_id, item_id, item),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@ai_router.get("/ai-item-knowledge/{cookie_id}/{item_id}/versions")
def get_ai_item_knowledge_versions(cookie_id: str, item_id: str,
                                   current_user: Dict[str, Any] = Depends(get_current_user)):
    _get_ai_knowledge_item(cookie_id, item_id, current_user)
    return {'versions': db_manager.get_ai_item_knowledge_versions(cookie_id, item_id)}


@ai_router.post("/ai-item-knowledge/{cookie_id}/{item_id}/rollback/{version}")
def rollback_ai_item_knowledge(cookie_id: str, item_id: str, version: int,
                               current_user: Dict[str, Any] = Depends(get_current_user)):
    item = _get_ai_knowledge_item(cookie_id, item_id, current_user)
    try:
        profile = db_manager.rollback_ai_item_knowledge(cookie_id, item_id, version)
        return {
            "message": f"已回滚并发布为第 {profile['version']} 版",
            **_item_knowledge_payload(cookie_id, item_id, item),
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


def _require_owned_cookie(cookie_id: str, user_id: int) -> None:
    if cookie_id not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限访问该闲鱼账号")


def _account_verification_identities(cookie_id: str) -> list[str]:
    details = db_manager.get_cookie_details(cookie_id) or {}
    values = [
        str(cookie_id or '').strip(),
        str(details.get('xianyu_unb') or '').strip(),
    ]
    return list(dict.fromkeys(value for value in values if value))


def _account_verification_images(cookie_id: str) -> list[Path]:
    return list_private_verification_images(
        _account_verification_identities(cookie_id)
    )


def _latest_account_verification_image(cookie_id: str) -> Optional[Path]:
    return latest_private_verification_image(
        _account_verification_identities(cookie_id)
    )


def _remove_account_verification_images(cookie_id: str) -> None:
    for image_path in _account_verification_images(cookie_id):
        remove_verification_image(str(image_path))


def _current_session_refresh_status(cookie_id: str) -> Dict[str, Any]:
    refresh_status = db_manager.get_account_session_refresh(cookie_id)
    state = refresh_status.get('state')
    registry_active = active_refresh_registry.is_active(cookie_id)
    if state in {'refreshing', 'verification_required'} and not registry_active:
        _remove_account_verification_images(cookie_id)
        db_manager.update_account_session_refresh(
            cookie_id,
            state='action_required',
            trigger=refresh_status.get('trigger') or 'automatic',
            message='需要手动开始一次验证，当前没有正在运行的官方浏览器',
            error_code='browser_session_missing',
        )
        refresh_status = db_manager.get_account_session_refresh(cookie_id)
        state = refresh_status.get('state')
    if (
        state in {'refreshing', 'verification_required'}
        and registry_active
        and refresh_status.get('expires_at')
        and time.time() > float(refresh_status['expires_at'])
    ):
        active_refresh_registry.cancel(cookie_id)
        _remove_account_verification_images(cookie_id)
        db_manager.update_account_session_refresh(
            cookie_id,
            state='timeout',
            trigger=refresh_status.get('trigger') or 'automatic',
            message='身份验证已超时，请重新发起刷新',
            error_code='verification_timeout',
        )
        refresh_status = db_manager.get_account_session_refresh(cookie_id)
    refresh_status['browser_active'] = active_refresh_registry.browser_active(cookie_id)
    refresh_status['verification_image_url'] = (
        f"/face-verification/screenshot/{cookie_id}"
        if refresh_status.get('state') == 'verification_required'
        and _latest_account_verification_image(cookie_id) is not None
        else ''
    )
    return refresh_status


@accounts_router.get("/api/accounts/{cookie_id}/session-status")
def get_account_session_status(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    _require_owned_cookie(cookie_id, current_user['user_id'])
    status_data = _current_session_refresh_status(cookie_id)
    session_id = f"cookie-refresh:{cookie_id}"
    registry = get_session_registry()
    if registry.get(session_id):
        registry.update(
            session_id,
            status=status_data.get('state') or 'idle',
            error_code=status_data.get('error_code') or '',
            error_message=status_data.get('message') or '',
        )
    return {'success': True, 'data': status_data}


@accounts_router.post("/api/accounts/{cookie_id}/session-refresh")
async def refresh_account_session(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    _require_owned_cookie(cookie_id, current_user['user_id'])
    try:
        task = db_manager.create_client_renewal_task(
            user_id=current_user['user_id'],
            cookie_id=cookie_id,
            trigger='manual_client_device',
        )
    except ClientBrowserError as exc:
        if exc.error_code == 'renewal_task_exists':
            return {
                'success': True,
                'status': 'client_device_pending',
                'message': '当前设备续期任务已经在进行中',
                'data': _current_session_refresh_status(cookie_id),
            }
        if exc.error_code == 'client_device_binding_required':
            db_manager.update_account_session_refresh(
                cookie_id, state='manual_reauth_required', trigger='manual',
                message='绑定当前设备后可恢复自动续期',
                error_code='client_device_binding_required',
            )
            return {
                'success': False,
                'status': 'client_device_binding_required',
                'message': '请先绑定当前 Chrome 或 Edge 作为此账号的续期设备',
                'reauth_action': 'bind_client_device',
                'data': db_manager.get_account_session_refresh(cookie_id),
            }
        _raise_client_browser_error(exc)
    db_manager.update_account_session_refresh(
        cookie_id, state='refreshing', trigger='manual_client_device',
        message='等待绑定的当前设备领取续期任务',
        error_code='client_device_renewal_pending',
        expires_at=task['expires_at'],
    )
    return {
        'success': True,
        'status': 'client_device_pending',
        'message': '已发送到绑定的当前设备',
        'data': _current_session_refresh_status(cookie_id),
    }


@accounts_router.post("/api/accounts/{cookie_id}/session-refresh/cancel")
def cancel_account_session_refresh(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    _require_owned_cookie(cookie_id, current_user['user_id'])
    status_info = _current_session_refresh_status(cookie_id)
    cancelled = db_manager.cancel_active_client_renewal_task(
        user_id=current_user['user_id'], cookie_id=cookie_id
    ) or active_refresh_registry.cancel(cookie_id)
    if not cancelled:
        raise HTTPException(status_code=409, detail="当前没有正在运行的刷新任务")
    _remove_account_verification_images(cookie_id)
    db_manager.update_account_session_refresh(
        cookie_id, state='cancelled', trigger=status_info.get('trigger') or 'manual',
        message='Cookie 刷新已取消', error_code='cancelled',
    )
    get_session_registry().update(
        f"cookie-refresh:{cookie_id}",
        status='cancelled',
        error_code='cancelled',
        error_message='Cookie 刷新已取消',
    )
    return {'success': True, 'message': '刷新已取消'}


@accounts_router.post("/api/accounts/{cookie_id}/session-refresh/show-browser")
def show_account_session_refresh_browser(
    cookie_id: str,
    request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    _require_server_browser_access(request, current_user)
    _require_owned_cookie(cookie_id, current_user['user_id'])
    if not active_refresh_registry.show_browser(cookie_id):
        raise HTTPException(status_code=404, detail="没有正在等待人工操作的官方浏览器会话")
    return {'success': True, 'message': '已请求在本机显示同一官方会话'}


@accounts_router.get("/api/diagnostics/auto-reply/{cookie_id}")
def diagnose_auto_reply(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """诊断指定账号的自动回复链路"""
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        _require_owned_cookie(cookie_id, user_id)

        cookie_info = db_manager.get_cookie_details(cookie_id) or {}
        cookie_value = user_cookies.get(cookie_id, '')
        ai_settings = db_manager.get_ai_reply_settings(cookie_id)
        default_reply = db_manager.get_default_reply(cookie_id) or {}
        status_enabled = db_manager.get_cookie_status(cookie_id)
        refresh_status = _current_session_refresh_status(cookie_id)

        manager_ready = cookie_manager.manager is not None
        manager_has_cookie = False
        task_running = False
        task_done = False
        task_error = ''
        task_status = {}
        recent_runtime_error = ''
        if manager_ready:
            manager_has_cookie = cookie_id in getattr(cookie_manager.manager, 'cookies', {})
            task_status = getattr(cookie_manager.manager, 'task_status', {}).get(cookie_id, {}) or {}
            task = getattr(cookie_manager.manager, 'tasks', {}).get(cookie_id)
            if task:
                task_done = task.done()
                task_running = not task_done
                if task_done:
                    try:
                        exc = task.exception()
                        task_error = str(exc) if exc else ''
                    except Exception as exc_check_error:
                        task_error = str(exc_check_error)
            if not task_running:
                recent_runtime_error = task_status.get('last_error') or task_error or ''

        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM keywords WHERE cookie_id = ?", (cookie_id,))
            keyword_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM default_replies WHERE cookie_id = ?", (cookie_id,))
            default_reply_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM ai_conversations WHERE cookie_id = ?", (cookie_id,))
            conversation_count = cursor.fetchone()[0]
            cursor.execute('''
                SELECT role, content, created_at
                FROM ai_conversations
                WHERE cookie_id = ?
                ORDER BY created_at DESC
                LIMIT 5
            ''', (cookie_id,))
            recent_conversations = [
                {'role': row[0], 'content': (row[1] or '')[:120], 'created_at': row[2]}
                for row in cursor.fetchall()
            ]
            cursor.execute('''
                SELECT event_type, event_description, processing_result, processing_status, error_message, created_at, updated_at
                FROM risk_control_logs
                WHERE cookie_id = ?
                ORDER BY id DESC
                LIMIT 1
            ''', (cookie_id,))
            risk_row = cursor.fetchone()
            latest_risk_control = {
                'event_type': risk_row[0],
                'event_description': risk_row[1],
                'processing_result': risk_row[2],
                'processing_status': risk_row[3],
                'error_message': risk_row[4],
                'created_at': risk_row[5],
                'updated_at': risk_row[6],
            } if risk_row and refresh_status.get('state') in {'refreshing', 'verification_required'} else None

        issues = []
        if not status_enabled:
            issues.append("账号已暂停，自动回复不会运行")
        if len(cookie_value) <= 50:
            issues.append("Cookie 内容过短，可能无效")
        if not manager_ready:
            issues.append("CookieManager 未就绪")
        elif not manager_has_cookie:
            issues.append("运行中的账号管理器没有加载该账号，需要重启服务")
        if manager_ready and manager_has_cookie and not task_running:
            issues.append(recent_runtime_error or task_error or "实时监听任务未运行")
        elif recent_runtime_error:
            issues.append(f"实时监听最近失败: {recent_runtime_error[:120]}")
        if not ai_settings.get('ai_enabled'):
            issues.append("账号 AI 回复未启用")
        if not ai_settings.get('api_key'):
            issues.append("未配置 AI API Key")
        if not ai_settings.get('model_name'):
            issues.append("未配置 AI 模型")
        if keyword_count == 0 and not ai_settings.get('ai_enabled') and not default_reply.get('enabled'):
            issues.append("关键词、AI、默认回复都未配置，无法自动回复")

        refresh_state = refresh_status.get('state')
        if refresh_state == 'action_required':
            issues.append(refresh_status.get('message') or "需要手动开始一次验证")
        elif refresh_state == 'verification_required':
            issues.append("Cookie 刷新正在等待身份验证，请在账号卡片中完成验证")
        elif refresh_state == 'refreshing':
            issues.append("Cookie 正在自动刷新，请稍候")
        elif refresh_state in {'failed', 'timeout'}:
            updated_at = refresh_status.get('updated_at')
            if is_runtime_event_active(
                updated_at,
                refresh_status.get('last_success_at'),
                max_age_seconds=600,
            ):
                issues.append(refresh_status.get('message') or "最近一次 Cookie 刷新失败")
        issues = list(dict.fromkeys(issues))

        return {
            "success": True,
            "data": {
                "cookie_id": cookie_id,
                "ready": len(issues) == 0,
                "issues": issues,
                "diagnosed_at": time.time(),
                "account": {
                    "enabled": bool(status_enabled),
                    "cookie_length": len(cookie_value),
                    "has_login_username": bool(cookie_info.get('username')),
                    "has_login_password": bool(cookie_info.get('password')),
                    "login_credentials_valid": bool(
                        cookie_info.get('password')
                        and is_valid_account_login_username(cookie_info.get('username'))
                    ),
                    "show_browser": bool(cookie_info.get('show_browser')),
                },
                "runtime": {
                    "manager_ready": manager_ready,
                    "manager_has_cookie": manager_has_cookie,
                    "task_running": task_running,
                    "task_done": task_done,
                    "task_error": task_error,
                    "task_status": task_status,
                    "recent_runtime_error": recent_runtime_error,
                    "latest_risk_control": latest_risk_control,
                },
                "session": refresh_status,
                "reply": {
                    "keyword_count": keyword_count,
                    "default_reply_count": default_reply_count,
                    "default_reply_enabled": bool(default_reply.get('enabled')),
                    "ai_enabled": bool(ai_settings.get('ai_enabled')),
                    "ai_model": ai_settings.get('model_name'),
                    "ai_base_url": ai_settings.get('base_url'),
                    "has_ai_key": bool(ai_settings.get('api_key')),
                    "conversation_count": conversation_count,
                    "recent_conversations": recent_conversations,
                }
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"自动回复诊断异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")



# ==================== 日志管理API ====================

@admin_router.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None,
                   _: Dict[str, Any] = Depends(require_admin)):
    """获取实时系统日志"""
    try:
        # 获取文件日志收集器
        collector = get_file_log_collector()

        # 获取日志
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}


@admin_router.get("/risk-control-logs")
async def get_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"获取风控日志失败: {str(e)}", admin_user)
        return {
            "success": False,
            "message": f"获取风控日志失败: {str(e)}",
            "data": [],
            "total": 0
        }


@admin_router.delete("/risk-control-logs/{log_id}")
async def delete_risk_control_log(
    log_id: int,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """删除风控日志记录（管理员专用）"""
    try:
        log_with_user('info', f"删除风控日志记录: {log_id}", admin_user)

        success = db_manager.delete_risk_control_log(log_id)

        if success:
            log_with_user('info', f"风控日志删除成功: {log_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"风控日志删除失败: {log_id}", admin_user)
            return {"success": False, "message": "删除失败，记录可能不存在"}

    except Exception as e:
        log_with_user('error', f"删除风控日志失败: {log_id} - {str(e)}", admin_user)
        return {"success": False, "message": f"删除失败: {str(e)}"}


@admin_router.get("/logs/stats")
async def get_log_stats(_: Dict[str, Any] = Depends(require_admin)):
    """获取日志统计信息"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}


@admin_router.post("/logs/clear")
async def clear_logs(_: Dict[str, Any] = Depends(require_admin)):
    """清空日志"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "日志已清空"}

    except Exception as e:
        return {"success": False, "message": f"清空日志失败: {str(e)}"}


# ==================== 商品管理API ====================

@content_router.post("/items/get-all-from-account")
async def get_all_items_from_account(
    request: dict,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """从指定账号获取所有商品信息"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 校验账号归属：禁止用他人账号的 Cookie 向闲鱼发起请求
        user_cookies = db_manager.get_all_cookies(current_user['user_id'])
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权操作此账号")

        # 获取指定账号的cookie信息
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(
            cookies_str,
            cookie_id,
            register_instance=False,
        )

        # 调用获取所有商品信息的方法（自动分页）
        logger.info(
            "开始同步账号 {} 的在售商品",
            _masked_identifier(cookie_id),
        )
        result = await xianyu_instance.get_all_items()

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            saved_count = result.get('total_saved', 0)
            active_count = result.get('active_count', total_count)
            hidden_count = result.get('hidden_count', 0)
            images_updated = result.get('images_updated', 0)
            failed_count = result.get('failed_count', 0)
            logger.info(
                "账号商品同步完成: account={}, pages={}, active={}, hidden={}, images_updated={}, failed={}",
                _masked_identifier(cookie_id),
                total_pages,
                active_count,
                hidden_count,
                images_updated,
                failed_count,
            )
            return {
                "success": True,
                "message": f"同步完成：在售 {active_count} 件，隐藏历史 {hidden_count} 件，更新图片 {images_updated} 件",
                "total_count": total_count,
                "total_pages": total_pages,
                "saved_count": saved_count,
                "active_count": active_count,
                "hidden_count": hidden_count,
                "images_updated": images_updated,
                "failed_count": failed_count,
            }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


@content_router.post("/items/get-by-page")
async def get_items_by_page(
    request: dict,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """从指定账号按页获取商品信息"""
    xianyu_instance = None
    try:
        # 验证参数
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 验证分页参数
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "页码和每页数量必须是数字"}

        if page_number < 1:
            return {"success": False, "message": "页码必须大于0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "每页数量必须在1-100之间"}

        # 先按当前租户读取 Cookie，禁止使用其他租户的平台会话。
        user_cookies = db_manager.get_all_cookies(current_user['user_id'])
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权操作此账号")

        cookies_str = user_cookies[cookie_id]
        if not cookies_str:
            return {"success": False, "message": "账号cookies为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(
            cookies_str,
            cookie_id,
            register_instance=False,
        )

        # 调用获取指定页商品信息的方法
        logger.info(
            "开始按页获取商品: account={}, page={}, page_size={}",
            _masked_identifier(cookie_id),
            page_number,
            page_size,
        )
        result = await xianyu_instance.get_item_list_info(page_number, page_size)

        if result.get('error'):
            logger.warning(
                "按页获取商品失败: account={}, page={}",
                _masked_identifier(cookie_id),
                page_number,
            )
            return {"success": False, "message": "获取商品信息失败"}

        current_count = result.get('current_count', 0)
        logger.info(
            "按页获取商品完成: account={}, page={}, count={}",
            _masked_identifier(cookie_id),
            page_number,
            current_count,
        )
        return {
            "success": True,
            "message": f"成功获取第{page_number}页 {current_count} 个商品",
            "page_number": page_number,
            "page_size": page_size,
            "current_count": current_count,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(
            "按页获取商品异常: account={}, error_type={}",
            _masked_identifier(request.get('cookie_id')),
            type(exc).__name__,
        )
        raise HTTPException(status_code=500, detail="获取商品信息失败")
    finally:
        if xianyu_instance is not None:
            try:
                await xianyu_instance.close_session()
            except Exception as exc:
                logger.warning(
                    "关闭商品同步会话失败: account={}, error_type={}",
                    _masked_identifier(request.get('cookie_id')),
                    type(exc).__name__,
                )


# ------------------------- 用户设置接口 -------------------------

@settings_router.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@settings_router.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新用户设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"更新用户设置: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"用户设置更新成功: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"用户设置更新失败: {key}", current_user)
            raise HTTPException(status_code=400, detail='更新失败')
    except Exception as e:
        log_with_user('error', f"更新用户设置异常: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@settings_router.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取用户特定设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='设置不存在')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 管理员专用接口 -------------------------

@admin_router.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有用户信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询所有用户信息", admin_user)
        users = db_manager.get_all_users()

        # 为每个用户添加统计信息
        for user in users:
            user_id = user['id']
            # 统计用户的Cookie数量
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # 统计用户的卡券数量
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # 隐藏密码字段
            if 'password_hash' in user:
                del user['password_hash']
            if 'password_hash_v2' in user:
                del user['password_hash_v2']

        log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.delete('/admin/users/{user_id}')
async def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除用户（管理员专用）"""
    from db_manager import db_manager
    try:
        # 不能删除管理员自己
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 获取要删除的用户信息
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="用户不存在")

        log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # 删除用户及其相关数据
        success = db_manager.delete_user_and_data(user_id)

        if success:
            import cookie_manager

            manager = cookie_manager.manager
            if manager is not None:
                reconcile = await manager.reconcile_from_db()
                if not reconcile.get("success"):
                    logger.error(
                        "用户删除后运行态对账未完成: "
                        f"user_id={user_id}, failed={reconcile.get('failed', 0)}"
                    )
                    raise HTTPException(
                        status_code=503,
                        detail="用户数据已删除，但运行态对账未完成，请人工检查",
                    )
            log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"用户 {user_to_delete['username']} 删除成功"}
        else:
            log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="删除失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.get('/admin/risk-control-logs')
async def get_admin_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"查询风控日志失败: {str(e)}", admin_user)
        return {"success": False, "message": f"查询失败: {str(e)}", "data": [], "total": 0}


@admin_router.get('/admin/cookies')
def get_admin_cookies(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有Cookie信息（管理员专用）"""
    try:
        log_with_user('info', "查询所有Cookie信息", admin_user)

        if cookie_manager.manager is None:
            return {
                "success": True,
                "cookies": [],
                "message": "CookieManager 未就绪"
            }

        # 获取所有用户的cookies
        from db_manager import db_manager
        all_users = db_manager.get_all_users()
        all_cookies = []

        for user in all_users:
            user_id = user['id']
            user_cookies = db_manager.get_all_cookies(user_id)
            for cookie_id, cookie_value in user_cookies.items():
                # 获取cookie详细信息
                cookie_details = db_manager.get_cookie_details(cookie_id)
                cookie_info = {
                    'cookie_id': cookie_id,
                    'user_id': user_id,
                    'username': user['username'],
                    'nickname': cookie_details.get('remark', '') if cookie_details else '',
                    'enabled': cookie_manager.manager.get_cookie_status(cookie_id)
                }
                all_cookies.append(cookie_info)

        log_with_user('info', f"获取到 {len(all_cookies)} 个Cookie", admin_user)
        return {
            "success": True,
            "cookies": all_cookies,
            "total": len(all_cookies)
        }

    except Exception as e:
        log_with_user('error', f"获取Cookie信息失败: {str(e)}", admin_user)
        return {
            "success": False,
            "cookies": [],
            "message": f"获取失败: {str(e)}"
        }


@admin_router.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_admin),
                   lines: int = 100,
                   level: str = None):
    """获取系统日志（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

        # 查找日志文件
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"找到日志文件: {log_files}")

        if not log_files:
            logger.warning("未找到日志文件")
            return {"logs": [], "message": "未找到日志文件", "success": False}

        # 获取最新的日志文件
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"使用最新日志文件: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"读取到 {len(all_lines)} 行日志")

                # 如果指定了日志级别，进行过滤
                if level:
                    filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                    logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                else:
                    filtered_lines = all_lines

                # 获取最后N行
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"取最后 {len(recent_lines)} 行日志")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"读取日志文件失败: {str(e)}")
            log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

        log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
        logger.info(f"成功返回 {len(logs)} 条日志记录")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"获取系统日志失败: {str(e)}")
        log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
        return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

@admin_router.get('/admin/log-files')
def list_log_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出所有可用的系统日志文件"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询日志文件列表", admin_user)

        log_dir = "logs"
        if not os.path.exists(log_dir):
            logger.warning("日志目录不存在")
            return {"success": True, "files": []}

        log_pattern = os.path.join(log_dir, "xianyu_*.log")
        log_files = glob.glob(log_pattern)

        files_info = []
        for file_path in log_files:
            try:
                stat_info = os.stat(file_path)
                files_info.append({
                    "name": os.path.basename(file_path),
                    "size": stat_info.st_size,
                    "modified_at": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                    "modified_ts": stat_info.st_mtime
                })
            except OSError as e:
                logger.warning(f"读取日志文件信息失败 {file_path}: {e}")

        # 按修改时间倒序排序
        files_info.sort(key=lambda item: item.get("modified_ts", 0), reverse=True)

        logger.info(f"返回日志文件列表，共 {len(files_info)} 个文件")
        return {"success": True, "files": files_info}

    except Exception as e:
        logger.error(f"获取日志文件列表失败: {str(e)}")
        log_with_user('error', f"获取日志文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.get('/admin/logs/export')
def export_log_file(file: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """导出指定的日志文件"""
    import os
    from fastapi.responses import StreamingResponse

    try:
        if not file:
            raise HTTPException(status_code=400, detail="缺少文件参数")

        safe_name = os.path.basename(file)
        log_dir = os.path.abspath("logs")
        target_path = os.path.abspath(os.path.join(log_dir, safe_name))

        # 防止目录遍历
        if not target_path.startswith(log_dir):
            log_with_user('warning', f"尝试访问非法日志文件: {file}", admin_user)
            raise HTTPException(status_code=400, detail="非法的日志文件路径")

        if not os.path.exists(target_path):
            log_with_user('warning', f"日志文件不存在: {file}", admin_user)
            raise HTTPException(status_code=404, detail="日志文件不存在")

        log_with_user('info', f"导出日志文件: {safe_name}", admin_user)
        def iter_file(path: str):
            file_handle = open(path, 'rb')
            try:
                while True:
                    chunk = file_handle.read(8192)
                    if not chunk:
                        break
                    yield chunk
            finally:
                file_handle.close()

        headers = {
            "Content-Disposition": f'attachment; filename="{safe_name}"'
        }
        return StreamingResponse(
            iter_file(target_path),
            media_type='text/plain; charset=utf-8',
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出日志文件失败: {str(e)}")
        log_with_user('error', f"导出日志文件失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取系统统计信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询系统统计信息", admin_user)

        # 用户统计
        all_users = db_manager.get_all_users()
        total_users = len(all_users)

        # Cookie统计
        all_cookies = db_manager.get_all_cookies()
        total_cookies = len(all_cookies)

        # 活跃账号统计（启用状态的账号）
        active_cookies = 0
        for cookie_id in all_cookies.keys():
            status = db_manager.get_cookie_status(cookie_id)
            if status:
                active_cookies += 1

        # 卡券统计
        all_cards = db_manager.get_all_cards()
        total_cards = len(all_cards) if all_cards else 0

        # 关键词统计
        all_keywords = db_manager.get_all_keywords()
        total_keywords = sum(len(kw_list) for kw_list in all_keywords.values())

        # 订单统计
        total_orders = 0
        try:
            orders = db_manager.get_all_orders()
            total_orders = len(orders) if orders else 0
        except:
            pass

        stats = {
            "total_users": total_users,
            "total_cookies": total_cookies,
            "active_cookies": active_cookies,
            "total_cards": total_cards,
            "total_keywords": total_keywords,
            "total_orders": total_orders
        }

        log_with_user('info', f"系统统计信息查询完成: {stats}", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- BI报表分析接口 -------------------------


def _dashboard_period(
    range_key: str,
    start_date: Optional[str],
    end_date: Optional[str],
) -> Dict[str, str]:
    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else datetime.now().date()
        if range_key == "custom":
            if not start_date or not end_date:
                raise ValueError("自定义时间范围需要开始和结束日期")
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        elif range_key == "yesterday":
            end = end - timedelta(days=1)
            start = end
        else:
            days = {"today": 1, "3days": 3, "7days": 7, "30days": 30}[range_key]
            start = end - timedelta(days=days - 1)
        if start > end:
            raise ValueError("开始日期不能晚于结束日期")
    except (KeyError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc) or "时间范围无效") from exc
    period_days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=period_days - 1)
    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "previous_start_date": previous_start.isoformat(),
        "previous_end_date": previous_end.isoformat(),
    }


@orders_router.get('/api/dashboard/summary')
def get_dashboard_summary(
    range_key: Literal['today', 'yesterday', '3days', '7days', '30days', 'custom'] = Query('7days', alias='range'),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    period = _dashboard_period(range_key, start_date, end_date)
    # 仪表盘一律只统计当前登录用户自己的数据，admin 也不例外（不再提供全局视图）
    scoped_user_id = current_user['user_id']
    valid_statuses = list(DASHBOARD_ANALYTICS_STATUSES)
    current = db_manager.get_order_analytics(
        start_date=period['start_date'],
        end_date=period['end_date'],
        user_id=scoped_user_id,
        include_statuses=valid_statuses,
    )
    previous = db_manager.get_order_analytics(
        start_date=period['previous_start_date'],
        end_date=period['previous_end_date'],
        user_id=scoped_user_id,
        include_statuses=valid_statuses,
    )
    if 'error' in current or 'error' in previous:
        raise HTTPException(
            status_code=500,
            detail=current.get('error') or previous.get('error') or '仪表盘统计失败',
        )
    single_day = period['start_date'] == period['end_date']
    if single_day:
        hourly = db_manager.get_traffic_analytics(
            start_date=period['start_date'],
            end_date=period['end_date'],
            user_id=scoped_user_id,
            include_statuses=valid_statuses,
        )
        if 'error' in hourly:
            raise HTTPException(status_code=500, detail=hourly.get('error') or '仪表盘时段统计失败')
        current['hourly_stats'] = hourly.get('hourly', [])
    return {
        "success": True,
        "scope": "user",
        "range": period,
        "trend_granularity": "hour" if single_day else "day",
        "stats": db_manager.get_dashboard_stats(scoped_user_id),
        "current": current,
        "previous": previous,
        "item_names": db_manager.get_dashboard_item_names(
            scoped_user_id,
            [item.get("item_id") for item in current["item_stats"]],
        ),
    }

@orders_router.get('/analytics/orders')
def get_order_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    获取订单分析数据（BI报表）

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询订单分析数据: {start_date} - {end_date}", current_user)

        # 获取当前用户的ID
        user_id = current_user['user_id']

        # 定义有效订单状态（只统计这几种状态）
        valid_statuses = list(DASHBOARD_ANALYTICS_STATUSES)

        # 调用数据库分析函数，传入包含状态
        analytics_data = db_manager.get_order_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )

        if 'error' in analytics_data:
            log_with_user('error', f"获取订单分析数据失败: {analytics_data['error']}", current_user)
            raise HTTPException(status_code=500, detail=analytics_data['error'])

        log_with_user('info', "订单分析数据查询成功", current_user)
        return analytics_data

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"获取订单分析数据失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@orders_router.get('/analytics/orders/valid')
def get_valid_orders(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    获取有效订单详情列表（用于统计中的订单明细）

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询有效订单列表: {start_date} - {end_date}", current_user)

        # 获取当前用户的ID
        user_id = current_user['user_id']

        # 定义有效订单状态
        valid_statuses = list(DASHBOARD_ANALYTICS_STATUSES)

        # 调用数据库函数获取有效订单
        orders = db_manager.get_orders_for_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )

        log_with_user('info', f"查询到 {len(orders)} 个有效订单", current_user)
        return {"orders": orders}

    except Exception as e:
        log_with_user('error', f"获取有效订单列表失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@orders_router.get('/analytics/traffic')
def get_traffic_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    订单时段分析（经营驾驶舱）

    按平台订单时间快照分桶到东八区小时/星期，回报覆盖率。
    只统计当前登录用户自己名下有效订单（租户隔离）。

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    try:
        log_with_user('info', f"查询订单时段分析: {start_date} - {end_date}", current_user)
        user_id = current_user['user_id']
        valid_statuses = list(DASHBOARD_ANALYTICS_STATUSES)
        data = db_manager.get_traffic_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )
        if 'error' in data:
            log_with_user('error', f"获取订单时段分析失败: {data['error']}", current_user)
            raise HTTPException(status_code=500, detail=data['error'])
        log_with_user('info', "订单时段分析查询成功", current_user)
        return data
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"获取订单时段分析失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@orders_router.get('/analytics/buyers')
def get_buyer_behavior_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    买家行为分析（经营驾驶舱）

    复购、下单频次分布、买家贡献榜；只做订单可直接得出的行为量，
    绝不刻画客户类型/年龄/职业/画像标签。只统计当前登录用户自己名下
    有效订单（租户隔离）。

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    try:
        log_with_user('info', f"查询买家行为分析: {start_date} - {end_date}", current_user)
        user_id = current_user['user_id']
        valid_statuses = list(VALID_ORDER_STATUSES)
        data = db_manager.get_buyer_behavior_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )
        if 'error' in data:
            log_with_user('error', f"获取买家行为分析失败: {data['error']}", current_user)
            raise HTTPException(status_code=500, detail=data['error'])
        log_with_user('info', "买家行为分析查询成功", current_user)
        return data
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"获取买家行为分析失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@orders_router.get('/analytics/items/performance')
def get_item_performance_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """成交商品表现；来源是订单，不把成交趋势命名为流量。"""
    data = db_manager.get_item_performance_analytics(
        user_id=current_user['user_id'],
        start_date=start_date,
        end_date=end_date,
    )
    if 'error' in data:
        raise HTTPException(status_code=500, detail=data['error'])
    return data


@orders_router.get('/analytics/items/traffic')
def get_item_traffic_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    cookie_id: Optional[str] = None,
    item_id: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Verified seller-backend exposure/view/want deltas with sufficiency gates."""
    try:
        return db_manager.get_item_traffic_analytics(
            user_id=current_user['user_id'],
            start_date=start_date,
            end_date=end_date,
            cookie_id=cookie_id,
            item_id=item_id,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@orders_router.get('/analytics/items/metrics/status')
def get_item_metric_status(
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Return account-scoped canary state without exposing account credentials."""
    from item_metric_service import item_metric_collection_status

    user_id = int(current_user['user_id'])
    accounts = [
        item_metric_collection_status(
            db_manager,
            user_id=user_id,
            cookie_id=cookie_id,
        )
        for cookie_id in db_manager.get_all_cookies(user_id)
    ]
    return {
        "adapter_available": any(
            bool(account.get("adapter_available")) for account in accounts
        ),
        "enabled_accounts": sum(
            int(bool(account.get("collection_enabled"))) for account in accounts
        ),
        "accounts": accounts,
    }


@orders_router.post('/analytics/items/metrics/sync')
async def sync_item_metrics(
    request: ItemMetricSyncRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Run one visible-account canary through the registered verified adapter."""
    from item_metric_scheduler import item_metric_scheduler
    from item_metric_service import collect_item_metrics_once

    user_cookies = db_manager.get_all_cookies(current_user['user_id'])
    if request.cookie_id not in user_cookies:
        raise HTTPException(status_code=404, detail="账号不存在或无权访问")
    result = await collect_item_metrics_once(
        db_manager,
        user_id=current_user['user_id'],
        cookie_id=request.cookie_id,
        cookie_string=user_cookies[request.cookie_id],
        canary=True,
    )
    if result.get("collection_enabled"):
        await item_metric_scheduler.start()
    status_code = 200 if result.get("success") else 409
    return JSONResponse(status_code=status_code, content=result)

# ------------------------- 指定商品回复接口 -------------------------

@content_router.get("/itemReplays")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品回复信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复信息失败: {str(e)}")

@content_router.get("/itemReplays/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")

@content_router.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品回复失败: {str(e)}")

@content_router.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除商品回复失败: {str(e)}")

class ItemToDelete(BaseModel):
    cookie_id: str
    item_id: str

class BatchDeleteRequest(BaseModel):
    items: List[ItemToDelete]

@content_router.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@content_router.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item_reply = db_manager.get_item_reply(cookie_id, item_id)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取商品回复失败: error_type={type(e).__name__}")
        raise HTTPException(status_code=500, detail="获取商品回复失败")


# ------------------------- 数据库备份和恢复接口 -------------------------

@admin_router.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_admin)):
    """下载数据库备份文件（管理员专用）"""
    import os
    from fastapi.responses import FileResponse
    from datetime import datetime

    try:
        log_with_user('info', "请求下载数据库备份", admin_user)

        # 使用db_manager的实际数据库路径
        from db_manager import db_manager
        db_file_path = db_manager.db_path

        # 检查数据库文件是否存在
        if not os.path.exists(db_file_path):
            log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="数据库文件不存在")

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.post('/admin/backup/upload')
async def upload_database_backup(
    admin_user: Dict[str, Any] = Depends(require_admin),
):
    """在线替换生产 SQLite 已永久关闭；停服回滚走运维 runbook。"""
    raise HTTPException(
        status_code=409,
        detail="在线数据库恢复已关闭，请按停服回滚流程执行",
    )

@admin_router.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出服务器上的备份文件（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询备份文件列表", admin_user)

        # 查找备份文件（在data目录中）
        backup_files = glob.glob("data/xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

        # 按修改时间倒序排列
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统管理接口 -------------------------

@admin_router.post('/admin/reload-cache')
async def reload_system_cache(admin_user: Dict[str, Any] = Depends(require_admin)):
    """刷新系统缓存（管理员专用）"""
    try:
        log_with_user('info', "刷新系统缓存", admin_user)

        # 这里可以添加实际的缓存刷新逻辑
        # 例如：重新加载配置、清理内存缓存等

        log_with_user('info', "系统缓存刷新成功", admin_user)
        return {"success": True, "message": "系统缓存已刷新"}

    except Exception as e:
        log_with_user('error', f"刷新系统缓存失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 数据管理接口 -------------------------

@admin_router.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'ai_training_rules',
            'ai_item_knowledge_profiles', 'ai_item_knowledge_versions', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay",
            'risk_control_logs'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许访问该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除指定表的指定记录（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'ai_training_rules',
            'ai_item_knowledge_profiles', 'ai_item_knowledge_versions', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许操作该表")

        # 特殊保护：不能删除管理员用户
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 删除记录
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@admin_router.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """清空指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"清空表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'ai_training_rules',
            'ai_item_knowledge_profiles', 'ai_item_knowledge_versions', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        # 不允许清空用户表
        if table_name == 'users':
            log_with_user('warning', "尝试清空用户表", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空用户表")

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空该表")

        # 清空表数据
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
            return {"success": True, "message": "清空成功"}
        else:
            log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="清空失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 商品多规格管理API
@content_router.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(
    cookie_id: str,
    item_id: str,
    spec_data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品的多规格状态"""
    try:
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(current_user['user_id'])
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"商品多规格状态已{'开启' if is_multi_spec else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 商品多数量发货管理API
@content_router.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(
    cookie_id: str,
    item_id: str,
    delivery_data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品的多数量发货状态"""
    try:
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(current_user['user_id'])
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"商品多数量发货状态已{'开启' if multi_quantity_delivery else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ItemDeliveryModeUpdate(BaseModel):
    mode: Literal['off', 'resource', 'invite']
    card_id: Optional[int] = None


class ItemDeliveryModeBatch(BaseModel):
    cookie_id: str
    item_ids: List[str] = Field(default_factory=list, max_length=500)
    mode: Literal['off', 'resource', 'invite']
    card_id: Optional[int] = None


def _delivery_mode_error(result: Dict[str, Any]) -> HTTPException:
    code = str(result.get('error') or 'update_failed')
    if code in {'item_not_found', 'resource_not_found'}:
        return HTTPException(status_code=404, detail=code)
    if code in {'resource_disabled', 'api_resource_not_validated'}:
        return HTTPException(status_code=409, detail=code)
    return HTTPException(status_code=400, detail=code)


@content_router.put("/items/{cookie_id}/{item_id}/delivery-mode")
def update_item_delivery_mode(
    cookie_id: str,
    item_id: str,
    payload: ItemDeliveryModeUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = current_user['user_id']
    if cookie_id not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限操作该账号商品")
    result = db_manager.set_item_delivery_mode(
        cookie_id,
        item_id,
        payload.mode,
        user_id,
        card_id=payload.card_id,
    )
    if result.get('outcome') != 'updated':
        raise _delivery_mode_error(result)
    return {
        "message": "发货方式已保存",
        "item_id": item_id,
        "mode": result.get('mode'),
        "card_id": result.get('card_id'),
    }


@content_router.post("/items/delivery-modes/batch")
def update_item_delivery_modes_batch(
    payload: ItemDeliveryModeBatch,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    user_id = current_user['user_id']
    if payload.cookie_id not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限操作该账号商品")
    item_ids = [str(value).strip() for value in payload.item_ids if str(value).strip()]
    if not item_ids:
        raise HTTPException(status_code=400, detail="请至少选择一个商品")
    return db_manager.set_item_delivery_modes_batch(
        payload.cookie_id,
        item_ids,
        payload.mode,
        user_id,
        card_id=payload.card_id,
    )


class ItemDeliveryBindingUpdate(BaseModel):
    """Compatibility shape; None now means explicit off, never keyword fallback."""
    card_id: Optional[int] = None


class ItemDeliveryBindingBatch(BaseModel):
    cookie_id: str
    item_ids: List[str] = Field(default_factory=list, max_length=500)
    card_id: Optional[int] = None


@content_router.put("/items/{cookie_id}/{item_id}/delivery-binding")
def update_item_delivery_binding(
    cookie_id: str,
    item_id: str,
    update_data: ItemDeliveryBindingUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Bind one card to one item, or clear the binding."""
    user_id = current_user['user_id']
    if cookie_id not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限操作该账号商品")

    if not db_manager.set_item_delivery_card(cookie_id, item_id, update_data.card_id, user_id):
        raise HTTPException(status_code=404, detail="商品不存在或卡密不可用")
    return {
        "message": "自动发货卡密已绑定" if update_data.card_id else "自动发货卡密已解除绑定",
        "card_id": update_data.card_id,
    }


@content_router.post("/items/delivery-bindings/batch")
def update_item_delivery_bindings_batch(
    payload: ItemDeliveryBindingBatch,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Bind the same card to several items in one action."""
    user_id = current_user['user_id']
    if payload.cookie_id not in db_manager.get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限操作该账号商品")
    item_ids = [str(item_id).strip() for item_id in payload.item_ids if str(item_id).strip()]
    if not item_ids:
        raise HTTPException(status_code=400, detail="请至少选择一个商品")

    updated, failed = 0, []
    for item_id in item_ids:
        if db_manager.set_item_delivery_card(payload.cookie_id, item_id, payload.card_id, user_id):
            updated += 1
        else:
            failed.append(item_id)
    if not updated:
        raise HTTPException(status_code=404, detail="商品不存在或卡密不可用")
    return {
        "message": f"已更新 {updated} 个商品的发货设置",
        "updated": updated,
        "failed": failed,
    }


class ItemInviteAutoFulfillmentUpdate(BaseModel):
    invite_auto_fulfillment: bool


@content_router.put("/items/{cookie_id}/{item_id}/invite-auto-fulfillment")
def update_item_invite_auto_fulfillment(
    cookie_id: str,
    item_id: str,
    update_data: ItemInviteAutoFulfillmentUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user),
):
    """Toggle the invite bridge for one item owned by the current user."""
    user_cookies = db_manager.get_all_cookies(current_user['user_id'])
    if cookie_id not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限操作该账号商品")

    enabled = update_data.invite_auto_fulfillment
    if not db_manager.update_item_invite_auto_fulfillment_status(
        cookie_id,
        item_id,
        enabled,
    ):
        raise HTTPException(status_code=404, detail="商品不存在")
    return {
        "message": f"邀请自动发货已{'开启' if enabled else '关闭'}",
        "invite_auto_fulfillment": enabled,
    }





# ==================== 订单管理接口 ====================

# 平台事实字段：只允许订单同步与发货/履约路径写入，人工编辑接口一律不接受。
ORDER_PLATFORM_STATE_FIELDS = frozenset({'order_status', 'system_shipped'})

# 人工可改的本地字段（导入订单的补录与收货信息更正）。
ORDER_MANUAL_EDITABLE_FIELDS = frozenset({
    'item_id', 'buyer_id', 'spec_name', 'spec_value',
    'quantity', 'amount', 'created_at',
    'receiver_name', 'receiver_phone', 'receiver_address',
})


def _forged_order_state_fields(order: Dict[str, Any],
                               update_data: Optional[Dict[str, Any]]) -> List[str]:
    """找出试图把平台态改成与当前存量不同值的字段；原样回显不算改写。"""
    forged: List[str] = []
    for field in sorted(ORDER_PLATFORM_STATE_FIELDS):
        if field not in (update_data or {}):
            continue
        submitted = (update_data or {})[field]
        stored = (order or {}).get(field)
        if field == 'system_shipped':
            changed = bool(submitted) != bool(stored)
        else:
            changed = str(submitted or '').strip() != str(stored or '').strip()
        if changed:
            forged.append(field)
    return forged


def get_orders_db() -> Any:
    """订单路由统一的可替换数据库依赖。"""
    return db_manager


async def _persist_order_sync_cookie(
    cookie_id: str,
    cookie_string: str,
    orders_db: Any,
) -> bool:
    """Persist one verified account-scoped Cookie update and refresh its listener."""
    if not orders_db.update_cookie_account_info(cookie_id, cookie_value=cookie_string):
        return False
    if cookie_manager.manager:
        runtime_update = cookie_manager.manager.update_cookie(
            cookie_id,
            cookie_string,
            save_to_db=False,
        )
        if hasattr(runtime_update, '__await__'):
            await runtime_update
    return True


async def _sync_recent_orders(
    current_user: Dict[str, Any],
    orders_db: Any,
    cookie_id: Optional[str] = None,
    days: int = 90,
) -> JSONResponse:
    user_cookies = orders_db.get_all_cookies(current_user["user_id"])
    if cookie_id:
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=404, detail="账号不存在或无权访问")
        user_cookies = {cookie_id: user_cookies[cookie_id]}
    if not user_cookies:
        raise HTTPException(status_code=400, detail="当前没有可同步的闲鱼账号")

    async def persist_cookie(account_id: str, cookie_string: str) -> bool:
        return await _persist_order_sync_cookie(
            account_id,
            cookie_string,
            orders_db,
        )

    client = XianyuOrderListClient()
    coordinator = OrderSyncCoordinator(
        orders_db,
        discoverer=client.discover,
        cookie_updater=persist_cookie,
    )
    account_results = []
    total_summary = new_order_sync_summary()
    scalar_summary_keys = (
        "total_seen",
        "discovered",
        "status_updated",
        "details_updated",
        "unchanged",
        "failed",
        "status_unconfirmed",
    )
    for account_id, cookie_string in user_cookies.items():
        account_details = orders_db.get_cookie_details(account_id) or {}
        result = await coordinator.sync_account(
            cookie_id=account_id,
            cookie_string=cookie_string,
            days=days,
            user_agent=str(account_details.get('browser_user_agent') or ''),
        )
        account_results.append({"cookie_id": account_id, **result})
        for key in scalar_summary_keys:
            total_summary[key] += int((result.get("summary") or {}).get(key) or 0)
        account_coverage = (result.get("summary") or {}).get("field_coverage") or {}
        for field in SYNC_COVERAGE_FIELDS:
            target = total_summary["field_coverage"][field]
            source = account_coverage.get(field) or {}
            target["covered"] += int(source.get("covered") or 0)
            target["total"] += int(source.get("total") or 0)

    for field in SYNC_COVERAGE_FIELDS:
        coverage = total_summary["field_coverage"][field]
        coverage["rate"] = round(
            coverage["covered"] / coverage["total"], 4
        ) if coverage["total"] else 0.0

    skipped_reauth = [
        row["cookie_id"]
        for row in account_results
        if row.get("skipped") and row.get("skip_reason") == "skipped_reauth"
    ]
    active_results = [row for row in account_results if not row.get("skipped")]
    requires_login = [
        row["cookie_id"]
        for row in active_results
        if row.get("requires_login")
    ]
    successful = sum(1 for row in active_results if row.get("success"))
    active_count = len(active_results)
    all_active_successful = successful == active_count
    partial = bool(skipped_reauth) or any(
        row.get("partial") or not row.get("success") for row in active_results
    )
    permission_denied = any(
        row.get("error_code") == "platform_permission_denied"
        for row in account_results
    )
    payload = {
        "success": all_active_successful,
        "partial": partial,
        "message": (
            "登录状态待恢复的账号已跳过订单同步"
            if not active_count and skipped_reauth
            else "订单同步完成，登录状态待恢复的账号已跳过"
            if all_active_successful and skipped_reauth
            else "订单同步完成"
            if all_active_successful
            else "登录状态已过期，请先在账号管理更新登录状态"
            if requires_login and successful == 0
            else "平台拒绝订单访问，请检查卖家订单权限"
            if permission_denied and successful == 0
            else "订单同步部分完成"
        ),
        "days": days,
        "summary": total_summary,
        "requires_login": requires_login,
        "skipped_reauth": skipped_reauth,
        "accounts": account_results,
    }
    status_code = 409 if requires_login and active_count and successful == 0 else 200
    return JSONResponse(status_code=status_code, content=payload)


@orders_router.post('/api/orders/sync')
async def sync_recent_orders(
    request: OrderSyncRequest,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """Discover and reconcile recent seller orders with truthful partial failures."""
    return await _sync_recent_orders(
        current_user=current_user,
        orders_db=orders_db,
        cookie_id=request.cookie_id,
        days=request.days,
    )


def _compose_order_display(order: Dict[str, Any], catalog_item: Dict[str, str],
                           profile: Optional[Dict[str, Any]] = None) -> None:
    """列表/详情共用的展示补全：快照优先、目录兜底，并标注身份状态。

    item_identity: snapshot / catalog_fallback / missing
    buyer_identity: snapshot / profile / history_unsaved / missing
    """
    snapshot_title = str(order.get('item_title') or '')
    snapshot_image = str(order.get('item_image') or '')
    order['item_title'] = snapshot_title or catalog_item.get('item_title', '')
    order['item_image'] = snapshot_image or catalog_item.get('item_image', '')
    order['item_title_source'] = str(order.get('item_title_source') or (
        'catalog' if not snapshot_title and order['item_title'] else ''
    ))
    order['item_image_source'] = str(order.get('item_image_source') or (
        'catalog' if not snapshot_image and order['item_image'] else ''
    ))
    order['item_price'] = catalog_item.get('item_price', '')
    if snapshot_title or snapshot_image:
        order['item_identity'] = 'snapshot'
    elif order['item_title'] or order['item_image']:
        order['item_identity'] = 'catalog_fallback'
    else:
        order['item_identity'] = 'missing'
    snapshot_nickname = str(order.get('buyer_nickname') or '')
    snapshot_avatar = str(order.get('buyer_avatar_url') or '')
    profile_name = str((profile or {}).get('display_name') or '')
    profile_avatar = str((profile or {}).get('avatar_url') or '')
    order['buyer_display_name'] = snapshot_nickname or profile_name
    order['buyer_avatar_url'] = snapshot_avatar or profile_avatar
    buyer_snapshot_source = str(order.get('buyer_snapshot_source') or '')
    buyer_nickname_source = str(
        order.get('buyer_nickname_source') or buyer_snapshot_source
    )
    buyer_avatar_source = str(
        order.get('buyer_avatar_source') or buyer_snapshot_source
    )
    order['buyer_display_name_source'] = (
        buyer_nickname_source if snapshot_nickname
        else str((profile or {}).get('display_name_source') or '') if profile_name
        else ''
    )
    order['buyer_avatar_source'] = (
        buyer_avatar_source if snapshot_avatar
        else str((profile or {}).get('avatar_source') or '') if profile_avatar
        else ''
    )
    if snapshot_nickname or snapshot_avatar:
        order['buyer_identity'] = 'snapshot'
    elif profile_name or profile_avatar:
        order['buyer_identity'] = 'profile'
    elif buyer_snapshot_source == 'history_unsaved':
        order['buyer_identity'] = 'history_unsaved'
    else:
        order['buyer_identity'] = 'missing'


def _parse_order_query_date(
    value: Any,
    field_name: str,
) -> Optional[datetime]:
    """校验订单筛选日期；保持外部 YYYY-MM-DD 字符串契约。"""
    if value is None or not isinstance(value, str):
        return None
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "invalid_order_date",
                "field": field_name,
                "message": "日期必须是有效的 YYYY-MM-DD 日历日期",
            },
        ) from exc
    if parsed.strftime("%Y-%m-%d") != value:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "invalid_order_date",
                "field": field_name,
                "message": "日期必须是有效的 YYYY-MM-DD 日历日期",
            },
        )
    return parsed


@orders_router.get('/api/orders')
def get_user_orders(
    current_user: Dict[str, Any] = Depends(get_current_user),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    cookie_id: Optional[str] = Query(None, description="筛选Cookie ID"),
    status: Optional[str] = Query(None, description="筛选状态"),
    search: Optional[str] = Query(None, max_length=100, description="搜索订单号/商品/买家"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    orders_db: Any = Depends(get_orders_db),
):
    """获取当前用户的订单信息（服务端筛选与分页；列表不含收货隐私字段）"""
    try:
        user_id = current_user['user_id']
        log_with_user('info', f"查询用户订单信息 (page={page}, page_size={page_size})", current_user)
        parsed_start = _parse_order_query_date(start_date, "start_date")
        parsed_end = _parse_order_query_date(end_date, "end_date")
        if parsed_start and parsed_end and parsed_start > parsed_end:
            raise HTTPException(
                status_code=422,
                detail={
                    "code": "invalid_order_date_range",
                    "message": "开始日期不得晚于结束日期",
                },
            )
        normalized_start = (
            parsed_start.strftime("%Y-%m-%d") if parsed_start else None
        )
        normalized_end = parsed_end.strftime("%Y-%m-%d") if parsed_end else None

        user_cookies = orders_db.get_all_cookies(user_id)
        if cookie_id:
            # 未授权账号与 /api/orders/sync 同语义：404，不再静默回退全账号
            if cookie_id not in user_cookies:
                raise HTTPException(status_code=404, detail="账号不存在或无权访问")
            scope_ids = [cookie_id]
        else:
            scope_ids = list(user_cookies.keys())

        result = orders_db.query_orders(
            scope_ids, status=status, search=search or '',
            start_date=normalized_start, end_date=normalized_end,
            page=page, page_size=page_size,
        )
        for order in result['items']:
            order['id'] = order['order_id']
            catalog_item = {
                'item_title': order.pop('catalog_title', '') or '',
                'item_image': order.pop('catalog_image', '') or '',
                'item_price': order.pop('catalog_price', '') or '',
            }
            profile = {
                'display_name': order.pop('profile_display_name', '') or '',
                'avatar_url': order.pop('profile_avatar_url', '') or '',
                'profile_source': order.pop('profile_source', '') or '',
                'display_name_source': order.pop(
                    'profile_display_name_source', ''
                ) or '',
                'avatar_source': order.pop('profile_avatar_source', '') or '',
            }
            _compose_order_display(order, catalog_item, profile)

        total = result['total']
        total_pages = max(1, (total + page_size - 1) // page_size) if total else 0
        log_with_user('info', f"用户订单查询成功，共 {total} 条记录，第 {page}/{total_pages or 1} 页", current_user)
        return {
            "success": True,
            "data": result['items'],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user(
            'error',
            f"查询用户订单失败 ({type(e).__name__})",
            current_user,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "code": "order_query_failed",
                "message": "订单列表查询失败，请稍后重试",
            },
        )


@orders_router.get('/api/orders/{order_id}')
def get_order_detail(
    order_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """获取订单详情（含收货信息与成交快照全字段）"""
    try:
        user_id = current_user['user_id']
        log_with_user('info', f"查询订单详情: {order_id}", current_user)

        user_cookies = orders_db.get_all_cookies(user_id)
        order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
        if not order or order.get('cookie_id') not in user_cookies:
            log_with_user('warning', f"订单不存在或无权访问: {order_id}", current_user)
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        cookie_id = order['cookie_id']
        catalog_item = orders_db.get_item_catalog_lookup([cookie_id]).get(
            (str(cookie_id), str(order.get('item_id') or '')), {})
        profile = orders_db.get_customer_profile(
            str(cookie_id),
            str(order.get('buyer_id') or ''),
        )
        _compose_order_display(order, catalog_item, profile)
        log_with_user('info', f"订单详情查询成功: {order_id}", current_user)
        return {"success": True, "data": order}

    except HTTPException:
        raise
    except Exception as e:
        log_with_user(
            'error',
            f"查询订单详情失败 ({type(e).__name__})",
            current_user,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "code": "order_detail_query_failed",
                "message": "订单详情查询失败，请稍后重试",
            },
        )


# 订单商品图媒体缓存目录（不在 /static 挂载下，必须经鉴权端点访问）
ORDER_ITEM_IMAGE_CACHE_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'data', 'media_cache', 'order_items'
)
ORDER_ITEM_IMAGE_MAX_BYTES = 8 * 1024 * 1024
ORDER_ITEM_IMAGE_TIMEOUT_SECONDS = 10
ORDER_ITEM_IMAGE_MAX_REDIRECTS = 3
ORDER_ITEM_IMAGE_FAILURE_TTL_SECONDS = 300
ORDER_ITEM_IMAGE_FAILURE_CACHE_MAX_ENTRIES = 512
ORDER_ITEM_IMAGE_MAX_CONCURRENT_DOWNLOADS = 4
ORDER_ITEM_IMAGE_TRUSTED_HOST_SUFFIXES = (
    'alicdn.com',
    'goofish.com',
    'tbcdn.cn',
)
_ORDER_ITEM_IMAGE_REDIRECT_STATUSES = {301, 302, 303, 307, 308}
_ORDER_ITEM_IMAGE_FAILURE_CACHE: OrderedDict[str, Tuple[float, int, str]] = (
    OrderedDict()
)
_ORDER_ITEM_IMAGE_INFLIGHT: Dict[Tuple[int, str], asyncio.Task] = {}
_ORDER_ITEM_IMAGE_SEMAPHORES: Dict[int, asyncio.Semaphore] = {}
_ORDER_ITEM_IMAGE_STATE_LOCK = threading.Lock()


def _order_item_image_cache_path(image_url: str) -> Tuple[str, str]:
    """由源地址导出确定性缓存键与落盘路径（重编码后统一为 JPEG）。"""
    cache_key = hashlib.sha256(image_url.encode('utf-8')).hexdigest()[:32] + '.jpg'
    return cache_key, os.path.join(ORDER_ITEM_IMAGE_CACHE_DIR, cache_key)


def _get_order_item_image_failure(cache_key: str) -> Optional[Tuple[int, str]]:
    now = time.monotonic()
    with _ORDER_ITEM_IMAGE_STATE_LOCK:
        cached = _ORDER_ITEM_IMAGE_FAILURE_CACHE.get(cache_key)
        if cached is None:
            return None
        expires_at, status_code, reason = cached
        if expires_at <= now:
            _ORDER_ITEM_IMAGE_FAILURE_CACHE.pop(cache_key, None)
            return None
        _ORDER_ITEM_IMAGE_FAILURE_CACHE.move_to_end(cache_key)
        return status_code, reason


def _remember_order_item_image_failure(
    cache_key: str,
    status_code: int,
    reason: str,
) -> None:
    with _ORDER_ITEM_IMAGE_STATE_LOCK:
        _ORDER_ITEM_IMAGE_FAILURE_CACHE[cache_key] = (
            time.monotonic() + ORDER_ITEM_IMAGE_FAILURE_TTL_SECONDS,
            status_code,
            reason,
        )
        _ORDER_ITEM_IMAGE_FAILURE_CACHE.move_to_end(cache_key)
        while (
            len(_ORDER_ITEM_IMAGE_FAILURE_CACHE)
            > ORDER_ITEM_IMAGE_FAILURE_CACHE_MAX_ENTRIES
        ):
            _ORDER_ITEM_IMAGE_FAILURE_CACHE.popitem(last=False)


def _forget_order_item_image_failure(cache_key: str) -> None:
    with _ORDER_ITEM_IMAGE_STATE_LOCK:
        _ORDER_ITEM_IMAGE_FAILURE_CACHE.pop(cache_key, None)


def _trusted_order_image_url(image_url: str) -> Tuple[str, int]:
    parsed = urlsplit(str(image_url or '').strip())
    host = str(parsed.hostname or '').rstrip('.').lower()
    if (
        parsed.scheme.lower() != 'https'
        or not host
        or parsed.username is not None
        or parsed.password is not None
        or parsed.fragment
    ):
        raise ValueError("untrusted image URL")
    try:
        port = parsed.port or 443
    except ValueError as exc:
        raise ValueError("invalid image URL port") from exc
    if port != 443:
        raise ValueError("untrusted image URL port")
    if not any(
        host == suffix or host.endswith(f'.{suffix}')
        for suffix in ORDER_ITEM_IMAGE_TRUSTED_HOST_SUFFIXES
    ):
        raise ValueError("untrusted image host")
    return host, port


async def _resolve_order_image_host(image_url: str) -> Tuple[str, ...]:
    """解析并拒绝任何非公网地址；每次重定向都会重新调用。"""
    host, port = _trusted_order_image_url(image_url)
    try:
        literal = ipaddress.ip_address(host)
        addresses = (str(literal),)
    except ValueError:
        loop = asyncio.get_running_loop()
        records = await loop.getaddrinfo(
            host,
            port,
            type=socket.SOCK_STREAM,
        )
        addresses = tuple(dict.fromkeys(str(record[4][0]) for record in records))
    if not addresses:
        raise ValueError("image host did not resolve")
    for address in addresses:
        ip = ipaddress.ip_address(address)
        if (
            ip.is_loopback
            or ip.is_private
            or ip.is_link_local
            or ip.is_reserved
            or ip.is_multicast
            or ip.is_unspecified
            or not ip.is_global
        ):
            raise ValueError("image host resolved to a non-public address")
    return addresses


class _PinnedOrderImageResolver:
    """aiohttp resolver：只返回已验证的公网地址，阻断校验后的 DNS 重绑定。"""

    def __init__(self, host: str, addresses: Tuple[str, ...]):
        self._records: Dict[str, Tuple[str, ...]] = {host: addresses}

    def pin(self, host: str, addresses: Tuple[str, ...]) -> None:
        self._records[host] = addresses

    async def resolve(self, host: str, port: int = 0, family: int = socket.AF_UNSPEC):
        addresses = self._records.get(str(host).rstrip('.').lower())
        if not addresses:
            raise OSError("host was not validated")
        return [
            {
                "hostname": host,
                "host": address,
                "port": port,
                "family": socket.AF_INET6 if ':' in address else socket.AF_INET,
                "proto": 0,
                "flags": 0,
            }
            for address in addresses
        ]

    async def close(self) -> None:
        self._records.clear()


async def _build_pinned_order_image_connector(image_url: str) -> Any:
    import aiohttp

    host, _port = _trusted_order_image_url(image_url)
    addresses = await _resolve_order_image_host(image_url)
    resolver = _PinnedOrderImageResolver(host, addresses)
    connector = aiohttp.TCPConnector(resolver=resolver, use_dns_cache=False)
    setattr(connector, '_order_image_resolver', resolver)
    return connector


async def _download_order_item_image(
    session: Any,
    image_url: str,
    resolver: _PinnedOrderImageResolver,
) -> bytes:
    current_url = image_url
    for redirect_count in range(ORDER_ITEM_IMAGE_MAX_REDIRECTS + 1):
        if redirect_count:
            addresses = await _resolve_order_image_host(current_url)
            host, _port = _trusted_order_image_url(current_url)
            resolver.pin(host, addresses)
        async with session.get(current_url, allow_redirects=False) as source_response:
            if source_response.status in _ORDER_ITEM_IMAGE_REDIRECT_STATUSES:
                if redirect_count >= ORDER_ITEM_IMAGE_MAX_REDIRECTS:
                    raise HTTPException(
                        status_code=404,
                        detail={"reason": "source_expired"},
                    )
                location = str(source_response.headers.get('Location') or '').strip()
                if not location:
                    raise HTTPException(
                        status_code=404,
                        detail={"reason": "source_expired"},
                    )
                current_url = urljoin(current_url, location)
                _trusted_order_image_url(current_url)
                continue
            if source_response.status != 200:
                raise HTTPException(
                    status_code=404,
                    detail={"reason": "source_expired"},
                )
            content_type = str(
                source_response.headers.get('Content-Type') or ''
            ).split(';', 1)[0].strip().lower()
            if not content_type.startswith('image/'):
                raise HTTPException(
                    status_code=422,
                    detail={"reason": "unsupported_format"},
                )
            content_length = str(
                source_response.headers.get('Content-Length') or ''
            ).strip()
            if content_length.isdigit() and int(content_length) > ORDER_ITEM_IMAGE_MAX_BYTES:
                raise HTTPException(
                    status_code=422,
                    detail={"reason": "unsupported_format"},
                )
            raw = await source_response.content.read(ORDER_ITEM_IMAGE_MAX_BYTES + 1)
            if len(raw) > ORDER_ITEM_IMAGE_MAX_BYTES:
                raise HTTPException(
                    status_code=422,
                    detail={"reason": "unsupported_format"},
                )
            return raw
    raise HTTPException(status_code=404, detail={"reason": "source_expired"})


def _publish_order_item_image_atomically(image: Any, cache_path: str) -> None:
    os.makedirs(ORDER_ITEM_IMAGE_CACHE_DIR, mode=0o700, exist_ok=True)
    descriptor, temporary_path = tempfile.mkstemp(
        prefix='.order-image-',
        suffix='.tmp',
        dir=ORDER_ITEM_IMAGE_CACHE_DIR,
    )
    os.close(descriptor)
    try:
        image.save(temporary_path, format='JPEG', quality=85)
        os.chmod(temporary_path, 0o600)
        os.replace(temporary_path, cache_path)
    finally:
        Path(temporary_path).unlink(missing_ok=True)


def _transcode_order_item_image(raw: bytes, cache_path: str) -> None:
    from PIL import Image as PILImage, UnidentifiedImageError

    try:
        image = PILImage.open(io.BytesIO(raw))
        image = image.convert('RGB')
        _publish_order_item_image_atomically(image, cache_path)
    except UnidentifiedImageError as exc:
        raise HTTPException(
            status_code=422,
            detail={"reason": "unsupported_format"},
        ) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={"reason": "unsupported_format"},
        ) from exc


async def _materialize_order_item_image(
    cache_key: str,
    image_url: str,
    cache_path: str,
) -> None:
    import aiohttp

    loop_id = id(asyncio.get_running_loop())
    semaphore = _ORDER_ITEM_IMAGE_SEMAPHORES.setdefault(
        loop_id,
        asyncio.Semaphore(ORDER_ITEM_IMAGE_MAX_CONCURRENT_DOWNLOADS),
    )
    try:
        async with semaphore:
            if os.path.isfile(cache_path):
                return
            timeout = aiohttp.ClientTimeout(
                total=ORDER_ITEM_IMAGE_TIMEOUT_SECONDS,
            )
            connector = await _build_pinned_order_image_connector(image_url)
            resolver = getattr(connector, '_order_image_resolver')
            try:
                async with aiohttp.ClientSession(
                    timeout=timeout,
                    connector=connector,
                    connector_owner=False,
                ) as session:
                    raw = await _download_order_item_image(
                        session,
                        image_url,
                        resolver,
                    )
            finally:
                await connector.close()
            await asyncio.to_thread(
                _transcode_order_item_image,
                raw,
                cache_path,
            )
    except HTTPException as exc:
        detail = exc.detail if isinstance(exc.detail, dict) else {}
        reason = str(detail.get('reason') or 'source_expired')
        status_code = 422 if exc.status_code == 422 else 404
        _remember_order_item_image_failure(
            cache_key,
            status_code,
            reason,
        )
        raise
    except asyncio.CancelledError:
        raise
    except Exception as exc:
        _remember_order_item_image_failure(
            cache_key,
            404,
            'source_expired',
        )
        raise HTTPException(
            status_code=404,
            detail={"reason": "source_expired"},
        ) from exc
    else:
        _forget_order_item_image_failure(cache_key)


async def _ensure_order_item_image(
    cache_key: str,
    image_url: str,
    cache_path: str,
) -> None:
    if os.path.isfile(cache_path):
        return
    loop = asyncio.get_running_loop()
    inflight_key = (id(loop), cache_key)
    task = _ORDER_ITEM_IMAGE_INFLIGHT.get(inflight_key)
    if task is None or task.done():
        task = loop.create_task(
            _materialize_order_item_image(
                cache_key,
                image_url,
                cache_path,
            )
        )
        _ORDER_ITEM_IMAGE_INFLIGHT[inflight_key] = task

        def forget(completed: asyncio.Task) -> None:
            if _ORDER_ITEM_IMAGE_INFLIGHT.get(inflight_key) is completed:
                _ORDER_ITEM_IMAGE_INFLIGHT.pop(inflight_key, None)

        task.add_done_callback(forget)
    await asyncio.shield(task)


@orders_router.get('/api/orders/{order_id}/item-image')
async def get_order_item_image(order_id: str,
                               retry: bool = Query(False),
                               current_user: Dict[str, Any] = Depends(get_current_user),
                               orders_db: Any = Depends(get_orders_db)):
    """返回订单商品图的应用内缓存字节流；失败返回机读原因。

    只服务订单绑定的图片地址（快照优先、目录兜底），不做任意 URL 代理。
    失败区分：not_saved（无可用图源）/ source_expired（源站拉取失败）/
    unsupported_format（源不是可解码图片，含 HEIC）。
    """
    user_id = current_user['user_id']
    user_cookies = orders_db.get_all_cookies(user_id)
    order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
    if not order or order.get('cookie_id') not in user_cookies:
        raise HTTPException(status_code=404, detail={"reason": "not_found"})

    cookie_id = str(order['cookie_id'])
    snapshot_image = str(order.get('item_image') or '')
    image_url = snapshot_image
    if not image_url:
        catalog_item = orders_db.get_item_catalog_lookup([cookie_id]).get(
            (cookie_id, str(order.get('item_id') or '')), {})
        image_url = str(catalog_item.get('item_image') or '')
    if not image_url:
        raise HTTPException(status_code=404, detail={"reason": "not_saved"})
    try:
        _trusted_order_image_url(image_url)
    except ValueError:
        raise HTTPException(status_code=404, detail={"reason": "source_expired"})

    cache_key, cache_path = _order_item_image_cache_path(image_url)
    response_headers = {"Cache-Control": "private, max-age=86400"}
    if os.path.isfile(cache_path):
        return FileResponse(cache_path, media_type='image/jpeg', headers=response_headers)
    if retry:
        _forget_order_item_image_failure(cache_key)
    else:
        cached_failure = _get_order_item_image_failure(cache_key)
        if cached_failure is not None:
            status_code, reason = cached_failure
            raise HTTPException(
                status_code=status_code,
                detail={"reason": reason},
            )

    await _ensure_order_item_image(cache_key, image_url, cache_path)

    # 缓存键只在图源仍是订单快照时回写（目录兜底图不建立订单级联）
    if image_url == snapshot_image:
        orders_db.set_order_item_image_cache_key(order_id, cache_key, snapshot_image)
    return FileResponse(cache_path, media_type='image/jpeg', headers=response_headers)


@orders_router.delete('/api/orders/{order_id}')
def delete_order(
    order_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """删除订单"""
    try:
        user_id = current_user['user_id']
        log_with_user('info', f"删除订单: {order_id}", current_user)

        # 获取用户的所有Cookie
        user_cookies = orders_db.get_all_cookies(user_id)

        # 验证订单属于当前用户
        order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        if order.get('cookie_id') not in user_cookies:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        # 删除时把归属条件带进 WHERE，先查后删之间的归属变化不会导致越权删除
        success = orders_db.delete_order(order_id, cookie_ids=list(user_cookies))
        if success:
            log_with_user('info', f"订单删除成功: {order_id}", current_user)
            return {"success": True, "message": "删除成功"}
        else:
            raise HTTPException(status_code=500, detail="删除失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"删除订单失败: {str(e)}")


@orders_router.post('/api/orders/{order_id}/refresh')
async def refresh_single_order(
    order_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """用卖家订单列表刷新单条订单；详情直连适配器验真前不启动浏览器。"""
    try:
        user_id = current_user['user_id']
        log_with_user('info', f"刷新单条订单: {order_id}", current_user)

        # 获取用户的所有Cookie
        user_cookies = orders_db.get_all_cookies(user_id)

        # 验证订单存在且属于当前用户
        order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        cookie_id = order.get('cookie_id')
        if not cookie_id or cookie_id not in user_cookies:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        cookies_str = user_cookies[cookie_id]
        if not cookies_str:
            raise HTTPException(status_code=400, detail="Cookie无效")

        client = XianyuOrderListClient()
        target_observed = False

        async def discover_target(**kwargs):
            nonlocal target_observed
            discovery = await client.discover(
                target_order_id=order_id,
                **kwargs,
            )
            if not discovery.get("success"):
                return discovery
            matching = [
                row for row in discovery.get("orders") or []
                if str(row.get("order_id") or "") == order_id
            ]
            target_observed = bool(matching)
            return {**discovery, "orders": matching}

        async def persist_cookie(account_id: str, cookie_string: str) -> bool:
            return await _persist_order_sync_cookie(
                account_id,
                cookie_string,
                orders_db,
            )

        account_details = orders_db.get_cookie_details(cookie_id) or {}
        result = await OrderSyncCoordinator(
            orders_db,
            discoverer=discover_target,
            cookie_updater=persist_cookie,
        ).sync_account(
            cookie_id=cookie_id,
            cookie_string=cookies_str,
            days=365,
            user_agent=str(account_details.get('browser_user_agent') or ''),
        )

        requires_login = bool(result.get("requires_login"))
        if not target_observed and result.get("success"):
            result = {
                **result,
                "success": False,
                "partial": False,
                "error_code": "order_not_observed",
                "message": "本轮订单列表未返回该订单，已保留原状态",
            }

        refreshed = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies)) or order
        error_code = str(result.get("error_code") or "")
        if error_code == "status_unconfirmed":
            message = "订单已获取部分字段，但平台状态仍待确认"
        else:
            message = result.get("message") or (
                "订单状态已刷新" if result.get("success") else "订单刷新未完成"
            )
        payload = {
            "success": bool(result.get("success")),
            "partial": bool(result.get("partial")),
            "error_code": error_code,
            "requires_login": requires_login,
            "message": message,
            "summary": result.get("summary") or new_order_sync_summary(),
            "fields_obtained": result.get("fields_obtained") or [],
            "data": {
                "order_id": order_id,
                "order_status": refreshed.get("order_status") or refreshed.get("status") or "unknown",
                "status_changed": bool((result.get("summary") or {}).get("status_updated")),
                "details_changed": bool((result.get("summary") or {}).get("details_updated")),
            },
        }
        log_with_user(
            'info',
            f"订单刷新结束: success={payload['success']}, partial={payload['partial']}, code={error_code or 'none'}",
            current_user,
        )
        return JSONResponse(
            status_code=409 if requires_login else 200,
            content=payload,
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"刷新订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"刷新订单失败: {str(e)}")


@orders_router.put('/api/orders/{order_id}')
async def update_order(
    order_id: str,
    update_data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """Update explicitly supplied local fields without hidden platform requests."""
    try:
        user_id = current_user['user_id']
        log_with_user(
            'info',
            f"更新订单: {order_id}, 字段数: {len(update_data or {})}",
            current_user,
        )

        user_cookies = orders_db.get_all_cookies(user_id)
        order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")
        if order.get('cookie_id') not in user_cookies:
            raise HTTPException(status_code=404, detail="订单不存在或无权访问")

        # 平台态（order_status）与发货态（system_shipped）只能由订单同步与履约路径写入。
        # 原样回显不报错，只忽略；一旦试图改成别的值就显式拒绝，不做静默丢弃。
        forged_fields = _forged_order_state_fields(order, update_data)
        if forged_fields:
            log_with_user(
                'warning',
                f"拒绝本地改写订单平台状态: {order_id}, 字段={forged_fields}",
                current_user,
            )
            raise HTTPException(
                status_code=403,
                detail={
                    "code": "platform_state_readonly",
                    "message": "发货与支付状态只能由订单同步或发货流程写入",
                    "fields": forged_fields,
                },
            )

        filtered_data = {
            key: value
            for key, value in (update_data or {}).items()
            if key in ORDER_MANUAL_EDITABLE_FIELDS
        }
        ignored_fields = sorted(
            key for key in (update_data or {})
            if key not in ORDER_MANUAL_EDITABLE_FIELDS
        )

        if not filtered_data:
            return {
                "success": True,
                "message": "没有可更新字段",
                "data": order,
                "refreshed": False,
                "ignored_fields": ignored_fields,
            }

        success = orders_db.insert_or_update_order(
            order_id=order_id,
            **filtered_data
        )

        if success:
            log_with_user('info', f"订单更新成功: {order_id}", current_user)
            updated_order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
            return {
                "success": True,
                "message": "更新成功",
                "data": updated_order,
                "refreshed": False,
                "ignored_fields": ignored_fields,
            }
        raise HTTPException(status_code=500, detail="更新失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"更新订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"更新订单失败: {str(e)}")


@orders_router.post('/api/orders/refresh', deprecated=True)
async def refresh_orders_status(
    cookie_id: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """Compatibility endpoint backed by the bounded direct order-list sync."""
    # 兼容旧前端；状态筛选不再缩小核对范围，避免漏掉签收或退款变化。
    return await _sync_recent_orders(
        current_user=current_user,
        orders_db=orders_db,
        cookie_id=cookie_id,
        days=90,
    )


@orders_router.post('/api/orders/manual-ship')
async def manual_ship_orders(
    order_ids: List[str] = Body(..., description="订单ID列表"),
    ship_mode: str = Body(..., description="发货模式: status_only（仅修改发货状态）或 full_delivery（完整发货流程）"),
    custom_content: Optional[str] = Body(None, description="自定义发货内容（保留兼容）"),
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """
    手动发货

    发货模式：
    - status_only: 仅在闲鱼标记为已发货（不发送卡券给买家）
    - full_delivery: 完整发货流程（匹配卡券、发送卡券给买家、标记发货状态）
    """
    try:
        from XianyuAutoAsync import (
            AUTO_DELIVERY_SOURCE_PAID_NOTICE,
            XianyuLive,
        )

        user_id = current_user['user_id']
        log_with_user('info', f"开始手动发货: 订单数量={len(order_ids)}, 模式={ship_mode}", current_user)

        # 验证发货模式
        if ship_mode not in ['status_only', 'full_delivery']:
            raise HTTPException(status_code=400, detail="发货模式必须是 status_only 或 full_delivery")

        # 获取用户的所有Cookie
        user_cookies = orders_db.get_all_cookies(user_id)

        success_count = 0
        failed_count = 0
        results = []

        # 遍历每个订单
        for order_id in order_ids:
            try:
                # 获取订单信息（归属条件直接进 WHERE）
                order = orders_db.get_order_by_id(order_id, cookie_ids=list(user_cookies))
                if not order:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '订单不存在或无权访问'
                    })
                    failed_count += 1
                    continue

                # 验证订单属于当前用户
                cookie_id = order.get('cookie_id')
                if cookie_id not in user_cookies:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '订单不存在或无权访问'
                    })
                    failed_count += 1
                    continue

                item_id = order.get('item_id')
                buyer_id = order.get('buyer_id')

                if ship_mode == 'status_only':
                    # ====== 仅修改闲鱼发货状态 ======
                    if not item_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少商品ID'
                        })
                        failed_count += 1
                        continue

                    # 获取cookies_str用于创建独立session
                    cookies_str = user_cookies.get(cookie_id)
                    if not cookies_str:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '无法获取账号Cookie信息'
                        })
                        failed_count += 1
                        continue

                    # 创建独立的aiohttp session（避免跨异步上下文问题）
                    import aiohttp
                    from secure_confirm_decrypted import SecureConfirm

                    try:
                        async with aiohttp.ClientSession(
                            headers={'cookie': cookies_str},
                            timeout=aiohttp.ClientTimeout(total=30)
                        ) as session:
                            confirm = SecureConfirm(session, cookies_str, cookie_id, None)
                            confirm_result = await confirm.auto_confirm(order_id, item_id)

                        if confirm_result and confirm_result.get('success'):
                            # 更新本地数据库状态
                            orders_db.insert_or_update_order(
                                order_id=order_id,
                                cookie_id=cookie_id,
                                order_status='shipped',
                                system_shipped=True
                            )
                            results.append({
                                'order_id': order_id,
                                'success': True,
                                'message': '已成功修改闲鱼发货状态'
                            })
                            success_count += 1
                        else:
                            error_msg = confirm_result.get('error', '未知错误') if confirm_result else '确认发货返回空结果'
                            results.append({
                                'order_id': order_id,
                                'success': False,
                                'message': f'修改发货状态失败: {error_msg}'
                            })
                            failed_count += 1
                    except Exception as e:
                        log_with_user('error', f"确认发货异常: {str(e)}", current_user)
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': f'确认发货异常: {str(e)}'
                        })
                        failed_count += 1

                elif ship_mode == 'full_delivery':
                    # ====== 完整发货流程：匹配卡券 + 发送卡券 + 修改状态 ======
                    if not item_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少商品ID，无法匹配发货规则'
                        })
                        failed_count += 1
                        continue

                    # 邀请自动发货商品由邀请服务独占卡密库存并发货，本地"完整发货"会与其
                    # 重复发码；这里直接拦截（仅 full_delivery，status_only 不受影响，运营
                    # 仍可用 status_only 在闲鱼标记发货状态）。
                    if orders_db.is_invite_auto_fulfillment_enabled(cookie_id, item_id):
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '该商品是邀请自动发货商品，由邀请服务发货，不能走本地完整发货以免重复发码；如需仅标记发货状态请改用 status_only。'
                        })
                        failed_count += 1
                        continue

                    if not buyer_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少买家ID，无法发送卡券'
                        })
                        failed_count += 1
                        continue

                    # 必须有运行中的实例（需要WebSocket发送消息）
                    live_instance = XianyuLive.get_instance(cookie_id)
                    if not live_instance:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '该账号未在线运行，无法执行完整发货。请先启动账号。'
                        })
                        failed_count += 1
                        continue

                    if not live_instance.ws or live_instance.ws.closed:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '该账号WebSocket连接已断开，无法发送消息。请等待重连后重试。'
                        })
                        failed_count += 1
                        continue

                    # 查找与买家的chat_id（优先从订单记录获取，回退到AI对话记录）
                    chat_id = order.get('chat_id') or ''
                    if not chat_id:
                        chat_id = orders_db.find_chat_id_by_buyer(cookie_id, buyer_id)
                    if not chat_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '未找到与该买家的聊天记录，无法发送卡券消息。请等待买家发送消息后重试。'
                        })
                        failed_count += 1
                        continue

                    payment_check = await live_instance._verify_paid_order_for_delivery(
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                    )
                    if not payment_check.get("allowed"):
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'partial': False,
                            'error_code': payment_check.get('error_code') or 'payment_unconfirmed',
                            'requires_login': bool(payment_check.get('requires_login')),
                            'manual_review': False,
                            'message': payment_check.get('reason') or '未确认订单处于待发货状态',
                        })
                        failed_count += 1
                        continue

                    # 检查多数量发货；开启后只信任本轮平台查询返回的数量。
                    quantity_to_send = 1
                    multi_quantity_delivery = orders_db.get_item_multi_quantity_delivery_status(cookie_id, item_id)
                    if multi_quantity_delivery:
                        trusted_quantity = payment_check.get('quantity')
                        if trusted_quantity is None:
                            results.append({
                                'order_id': order_id,
                                'success': False,
                                'partial': False,
                                'error_code': 'quantity_unconfirmed',
                                'requires_login': False,
                                'manual_review': False,
                                'message': '平台订单列表未返回可信购买数量'
                            })
                            failed_count += 1
                            continue
                        quantity_to_send = trusted_quantity

                    fulfillment = await live_instance._execute_fulfillment_attempt(
                        websocket=live_instance.ws,
                        order_id=order_id,
                        item_id=item_id,
                        buyer_id=buyer_id,
                        chat_id=chat_id,
                        expected_quantity=quantity_to_send,
                        delivery_source=AUTO_DELIVERY_SOURCE_PAID_NOTICE,
                        confirm_platform=True,
                        item_title=str(order.get('item_title') or ''),
                        database=orders_db,
                    )
                    if fulfillment.get('success'):
                        already_completed = bool(fulfillment.get('already_completed'))
                        results.append({
                            'order_id': order_id,
                            'success': True,
                            'partial': False,
                            'error_code': None,
                            'requires_login': False,
                            'manual_review': False,
                            'attempt_id': fulfillment.get('attempt_id'),
                            'sent_count': fulfillment.get('sent_count', quantity_to_send),
                            'expected_count': quantity_to_send,
                            'message': (
                                '订单已有已提交的完整发货记录，未重复发送'
                                if already_completed
                                else f'完整发货成功，已发送{quantity_to_send}条卡券信息给买家'
                            )
                        })
                        success_count += 1
                    else:
                        manual_review = bool(fulfillment.get('manual_review'))
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'partial': bool(fulfillment.get('partial')),
                            'error_code': fulfillment.get('error_code') or 'fulfillment_failed',
                            'requires_login': False,
                            'manual_review': manual_review,
                            'attempt_id': fulfillment.get('attempt_id'),
                            'sent_count': int(fulfillment.get('sent_count') or 0),
                            'expected_count': quantity_to_send,
                            'message': (
                                '完整发货未完成，状态不确定，已转人工复核'
                                if manual_review
                                else '完整发货未开始或发货内容不完整'
                            )
                        })
                        failed_count += 1

            except Exception as e:
                results.append({
                    'order_id': order_id,
                    'success': False,
                    'message': str(e)
                })
                failed_count += 1
                log_with_user('error', f"发货订单 {order_id} 时发生异常: {str(e)}", current_user)

        log_with_user('info', f"手动发货完成: 成功{success_count}个, 失败{failed_count}个", current_user)

        return {
            "success": True,
            "message": f"发货完成: 成功{success_count}个, 失败{failed_count}个",
            "total": len(order_ids),
            "success_count": success_count,
            "failed_count": failed_count,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"手动发货失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"手动发货失败: {str(e)}")


# 订单导入的合法字段映射：前端字段名 -> insert_or_update_order 参数名。
# status_text/pay_time/item_title/item_price 无对应订单列（商品信息以目录 join 为准），
# 不在映射内的字段一律丢弃并在结果 message 中提示，避免传入非法关键字参数。
_IMPORT_ORDER_PARAM_MAPPING = {
    'item_id': 'item_id',
    'buyer_id': 'buyer_id',
    'receiver_name': 'receiver_name',
    'receiver_phone': 'receiver_phone',
    'receiver_address': 'receiver_address',
    'receiver_city': 'receiver_city',
    'status': 'order_status',  # 注意：前端用 status，后端用 order_status
    'order_time': 'created_at',
    'quantity': 'quantity',
    'amount': 'amount',
    'item_image': 'item_image',
}
_ORDER_IMPORT_MAX_BYTES = 5 * 1024 * 1024
_ORDER_IMPORT_MAX_ROWS = 10_000
_ORDER_IMPORT_MAX_COLUMNS = 50
_ORDER_IMPORT_HEADER_ALIASES = {
    '订单号': 'order_id',
    '账号ID': 'cookie_id',
    '商品ID': 'item_id',
    '买家ID': 'buyer_id',
    '状态': 'status',
    '金额': 'amount',
    '数量': 'quantity',
    '订单时间': 'order_time',
    '收货人': 'receiver_name',
    '手机号': 'receiver_phone',
    '收货地址': 'receiver_address',
    '城市': 'receiver_city',
    '商品图片': 'item_image',
}


def _map_import_order_params(order_data: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
    """把导入的订单数据映射为 insert_or_update_order 的合法关键字参数。

    返回 (合法参数字典, 被忽略的字段名列表)。order_id/cookie_id 由调用方单独处理。
    """
    params: Dict[str, Any] = {}
    ignored: List[str] = []
    for field, value in order_data.items():
        if value is None or field in ('order_id', 'cookie_id'):
            continue
        param_name = _IMPORT_ORDER_PARAM_MAPPING.get(field)
        if param_name:
            params[param_name] = value
        else:
            ignored.append(field)
    return params, ignored


def _cell_import_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, 'isoformat'):
        return value.isoformat(sep=' ')
    if isinstance(value, str):
        return value.strip()
    return value


def _orders_from_xlsx(raw: bytes, filename: str) -> List[Dict[str, Any]]:
    if not str(filename or '').lower().endswith('.xlsx'):
        raise HTTPException(status_code=415, detail="仅支持 .xlsx 文件")
    if not raw:
        raise HTTPException(status_code=400, detail="Excel 文件为空")
    if len(raw) > _ORDER_IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Excel 文件超过 5MB")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        if sheet.max_row > _ORDER_IMPORT_MAX_ROWS + 1 or sheet.max_column > _ORDER_IMPORT_MAX_COLUMNS:
            raise HTTPException(status_code=413, detail="Excel 行列数超过限制")
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Excel 缺少表头")
        headers = [
            _ORDER_IMPORT_HEADER_ALIASES.get(str(value or '').strip(), str(value or '').strip())
            for value in header_row
        ]
        if not headers or any(not header for header in headers) or len(set(headers)) != len(headers):
            raise HTTPException(status_code=400, detail="Excel 表头为空或重复")
        missing = [field for field in ('order_id', 'cookie_id') if field not in headers]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Excel 缺少必需表头: {', '.join(missing)}",
            )
        orders = []
        for row in rows:
            values = [_cell_import_value(value) for value in row[:len(headers)]]
            if not any(value not in (None, '') for value in values):
                continue
            orders.append(dict(zip(headers, values)))
        return orders
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件损坏或格式不受支持")
    finally:
        if 'workbook' in locals():
            workbook.close()


async def _parse_order_import_request(request: Request) -> List[Dict[str, Any]]:
    """解析两个并列导入契约：程序化 JSON 对象数组，或 multipart `.xlsx` 文件。

    `.xlsx` 是唯一支持的电子表格格式；JSON 数组是既有程序化 API 能力，
    不属于电子表格格式限制。
    """
    content_type = str(request.headers.get('content-type') or '').lower()
    if content_type.startswith('application/json'):
        content_length = request.headers.get('content-length')
        if content_length and content_length.isdigit() and int(content_length) > _ORDER_IMPORT_MAX_BYTES:
            raise HTTPException(status_code=413, detail="导入内容超过 5MB")
        try:
            payload = await request.json()
        except Exception:
            raise HTTPException(status_code=400, detail="JSON 格式错误")
        if not isinstance(payload, list) or any(not isinstance(row, dict) for row in payload):
            raise HTTPException(status_code=422, detail="JSON 必须是订单对象数组")
        if len(payload) > _ORDER_IMPORT_MAX_ROWS:
            raise HTTPException(status_code=413, detail="订单行数超过限制")
        return payload
    if content_type.startswith('multipart/form-data'):
        form = await request.form()
        upload = form.get('file')
        if upload is None or not hasattr(upload, 'read'):
            raise HTTPException(status_code=400, detail="缺少 Excel 文件")
        raw = await upload.read(_ORDER_IMPORT_MAX_BYTES + 1)
        if len(raw) > _ORDER_IMPORT_MAX_BYTES:
            raise HTTPException(status_code=413, detail="Excel 文件超过 5MB")
        # 最多 5MB 的工作簿解析是纯 CPU 阻塞，放到线程里跑
        return await asyncio.to_thread(
            _orders_from_xlsx, raw, str(getattr(upload, 'filename', '') or '')
        )
    raise HTTPException(
        status_code=415,
        detail="仅支持程序化 JSON 数组或 multipart .xlsx 文件",
    )


@orders_router.post('/api/orders/import')
async def import_orders(
    request: Request,
    current_user: Dict[str, Any] = Depends(get_current_user),
    orders_db: Any = Depends(get_orders_db),
):
    """批量导入订单：支持程序化 JSON 数组与 multipart `.xlsx` 两个并列契约。"""
    try:
        orders = await _parse_order_import_request(request)
        user_id = current_user['user_id']
        log_with_user('info', f"开始导入订单: 订单数量={len(orders)}", current_user)

        # 获取用户的所有Cookie
        user_cookies = orders_db.get_all_cookies(user_id)

        success_count = 0
        failed_count = 0
        results = []

        # 必需字段验证
        required_fields = ['order_id', 'cookie_id']

        for order_data in orders:
            try:
                # 验证必需字段
                missing_fields = [f for f in required_fields if not order_data.get(f)]
                if missing_fields:
                    results.append({
                        'order_id': order_data.get('order_id', 'unknown'),
                        'success': False,
                        'message': f'缺少必需字段: {", ".join(missing_fields)}'
                    })
                    failed_count += 1
                    continue

                order_id = str(order_data['order_id'])
                cookie_id = str(order_data['cookie_id'])

                # 验证Cookie属于当前用户
                if cookie_id not in user_cookies:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '无权操作此账号的订单'
                    })
                    failed_count += 1
                    continue

                # 检查订单是否已存在（这里必须不带归属条件：越权订单要能被看见并拒绝）
                existing_order = await asyncio.to_thread(
                    orders_db.get_order_by_id, order_id
                )

                # 已存在订单必须归属当前用户名下的账号：
                # 归属其他账号（含 cookie_id 为 NULL 的历史孤儿订单）一律拒绝，
                # 防止通过导入接口接管/认领他人订单
                if existing_order and existing_order.get('cookie_id') not in user_cookies:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '无权操作此订单'
                    })
                    failed_count += 1
                    continue

                # 映射为 insert_or_update_order 的合法参数，映射外字段丢弃并提示
                mapped_params, ignored_fields = _map_import_order_params(order_data)
                insert_params = {
                    'order_id': order_id,
                    'cookie_id': cookie_id,
                    **mapped_params,
                }

                # 使用 insert_or_update_order 统一处理；整批导入逐行写入会长时间
                # 独占库锁，放到线程里执行让事件循环在行之间保持可用
                written = await asyncio.to_thread(
                    orders_db.insert_or_update_order, **insert_params
                )
                if not written:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '写入订单失败'
                    })
                    failed_count += 1
                    continue

                message = '订单已更新' if existing_order else '订单已导入'
                if ignored_fields:
                    message += f"（已忽略字段: {', '.join(ignored_fields)}）"
                results.append({
                    'order_id': order_id,
                    'success': True,
                    'message': message
                })

                success_count += 1

            except Exception as e:
                results.append({
                    'order_id': order_data.get('order_id', 'unknown'),
                    'success': False,
                    'message': str(e)
                })
                failed_count += 1
                log_with_user('error', f"导入订单时发生异常: {str(e)}", current_user)

        log_with_user('info', f"导入订单完成: 成功{success_count}个, 失败{failed_count}个", current_user)

        return {
            "success": True,
            "message": f"导入完成: 成功{success_count}个, 失败{failed_count}个",
            "total": len(orders),
            "success_count": success_count,
            "failed_count": failed_count,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"导入订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"导入订单失败: {str(e)}")


# ==================== 前端 SPA Catch-All 路由 ====================
# 必须放在所有 API 路由之后，用于处理前端 SPA 的直接访问
# 这样用户直接访问 /dashboard、/accounts 等前端路由时，会返回 index.html
# 然后由 React Router 在客户端处理路由

# 定义不需要返回前端页面的路径前缀（API 路径）
API_PREFIXES = ['/api/', '/internal/', '/static/', '/health', '/login', '/logout', '/register', '/verify', '/check-default-password', '/change-password', '/change-admin-password']

@frontend_router.get('/{path:path}', response_class=HTMLResponse)
async def catch_all_route(path: str):
    """
    Catch-all 路由：处理所有未匹配的 GET 请求
    如果是 API 请求，返回 404；否则返回前端 index.html
    """
    # 检查是否是 API 请求
    full_path = f'/{path}'
    for prefix in API_PREFIXES:
        if full_path.startswith(prefix):
            raise HTTPException(status_code=404, detail="Not Found")

    # 返回前端页面
    return await serve_frontend()


include_domain_routers(app)


# 移除自动启动，由Start.py或手动启动
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)
