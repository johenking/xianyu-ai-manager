"""订单中心 API 契约：隐私/授权、媒体安全、JSON 与 `.xlsx` 并列导入。"""

import io
import asyncio
import os
import socket
from pathlib import Path
import tempfile
import unittest
from contextlib import contextmanager
from unittest.mock import AsyncMock, patch

import aiohttp
from fastapi.testclient import TestClient
from openpyxl import Workbook
from PIL import Image as PILImage

from db_manager import DBManager
import reply_server


def _png_bytes() -> bytes:
    buffer = io.BytesIO()
    PILImage.new("RGB", (2, 2), color=(200, 40, 40)).save(buffer, format="PNG")
    return buffer.getvalue()


class _FakeContent:
    def __init__(self, data: bytes):
        self._data = data

    async def read(self, _limit: int) -> bytes:
        return self._data


def _fake_aiohttp_session(
    status: int = 200,
    body: bytes = b"",
    error: Exception = None,
    headers: dict = None,
    calls: list = None,
    session_kwargs: list = None,
):
    """替身 aiohttp.ClientSession：让媒体端点在测试里不发真实网络请求。"""

    class FakeResponse:
        def __init__(self):
            self.status = status
            self.content = _FakeContent(body)
            self.headers = headers or {"Content-Type": "image/png"}

        async def __aenter__(self):
            if error is not None:
                raise error
            return self

        async def __aexit__(self, *_exc):
            return False

    class FakeSession:
        def __init__(self, *args, **kwargs):
            if session_kwargs is not None:
                session_kwargs.append(dict(kwargs))

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_exc):
            return False

        def get(self, _url, **_kwargs):
            if calls is not None:
                calls.append(_url)
            return FakeResponse()

    return patch("aiohttp.ClientSession", FakeSession)


def _fake_aiohttp_sequence(responses: list, calls: list):
    class FakeResponse:
        def __init__(self, spec):
            self.status = spec["status"]
            self.content = _FakeContent(spec.get("body", b""))
            self.headers = spec.get("headers", {})

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_exc):
            return False

    class FakeSession:
        def __init__(self, *args, **kwargs):
            self._responses = iter(responses)

        async def __aenter__(self):
            return self

        async def __aexit__(self, *_exc):
            return False

        def get(self, url, **_kwargs):
            calls.append(url)
            return FakeResponse(next(self._responses))

    return patch("aiohttp.ClientSession", FakeSession)


@contextmanager
def _public_image_network(**kwargs):
    """隔离 DNS 与 HTTP；任何媒体测试都不访问真实网络。"""
    with patch(
        "reply_server._resolve_order_image_host",
        new=AsyncMock(return_value=("203.0.113.10",)),
        create=True,
    ), _fake_aiohttp_session(**kwargs):
        yield


class OrderApiContractTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.previous_key_file = os.environ.get("SYSTEM_SECRET_KEY_FILE")
        os.environ["SYSTEM_SECRET_KEY_FILE"] = str(self.root / ".system-key")
        self.db = DBManager(str(self.root / "orders-api.db"))
        self.assertTrue(self.db.create_user("seller-one", "one@example.test", "Strong-pass-2026!"))
        self.assertTrue(self.db.create_user("seller-two", "two@example.test", "Strong-pass-2026!"))
        self.user_one = self.db.get_user_by_username("seller-one")
        self.user_two = self.db.get_user_by_username("seller-two")
        with self.db.lock:
            self.db.conn.executemany(
                "INSERT INTO cookies (id, value, user_id) VALUES (?, ?, ?)",
                (
                    ("acct-one", "unb=1; cookie2=x", self.user_one["id"]),
                    ("acct-two", "unb=2; cookie2=x", self.user_two["id"]),
                ),
            )
            self.db.conn.commit()
        self.db.insert_or_update_order(
            order_id="order-1", item_id="item-1", buyer_id="buyer-1",
            amount="¥12.50", order_status="pending_ship", cookie_id="acct-one",
            created_at="2026-07-20 10:00:00", receiver_name="张三",
            receiver_phone="13800000000", receiver_address="幸福路1号", receiver_city="福州",
        )
        self.db.apply_order_sync_update(
            order_id="order-1", cookie_id="acct-one", incoming_status="pending_ship",
            status_source="order_list",
            item_snapshot={"item_title": "成交快照标题", "source": "order_list"},
            buyer_snapshot={"buyer_nickname": "买家甲", "source": "order_list"},
        )
        self.db.insert_or_update_order(
            order_id="order-2", item_id="item-2", buyer_id="buyer-2",
            amount="¥99.00", order_status="completed", cookie_id="acct-two",
        )
        self.original_db = reply_server.db_manager
        reply_server.db_manager = self.db
        reply_server.SESSION_TOKENS.clear()
        self.media_dir = str(self.root / "media")
        self._media_patch = patch.object(
            reply_server, "ORDER_ITEM_IMAGE_CACHE_DIR", self.media_dir
        )
        self._media_patch.start()
        self.client = TestClient(reply_server.app, raise_server_exceptions=False)

    def tearDown(self):
        self.client.close()
        self._media_patch.stop()
        reply_server.SESSION_TOKENS.clear()
        reply_server.db_manager = self.original_db
        self.db.close()
        if self.previous_key_file is None:
            os.environ.pop("SYSTEM_SECRET_KEY_FILE", None)
        else:
            os.environ["SYSTEM_SECRET_KEY_FILE"] = self.previous_key_file
        self.tempdir.cleanup()

    def headers_for(self, user):
        token, _ = reply_server.create_login_session(user)
        return {"Authorization": f"Bearer {token}"}

    def test_every_order_route_uses_the_replaceable_database_dependency(self):
        missing_dependency = []
        for route in reply_server.orders_router.routes:
            if not getattr(route, "path", "").startswith("/api/orders"):
                continue
            dependencies = {
                dependency.call
                for dependency in getattr(route, "dependant", ()).dependencies
            }
            if reply_server.get_orders_db not in dependencies:
                missing_dependency.append(route.path)
        self.assertEqual(missing_dependency, [])

    def test_list_hides_receiver_privacy_and_serves_snapshot_display(self):
        response = self.client.get("/api/orders", headers=self.headers_for(self.user_one))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total"], 1)
        row = payload["data"][0]
        for private_field in ("receiver_name", "receiver_phone", "receiver_address", "receiver_city"):
            self.assertNotIn(private_field, row)
        self.assertEqual(row["item_title"], "成交快照标题")
        self.assertEqual(row["item_identity"], "snapshot")
        self.assertEqual(row["buyer_display_name"], "买家甲")
        self.assertEqual(row["buyer_identity"], "snapshot")
        self.assertEqual(row["buyer_display_name_source"], "order_list")

    def test_list_search_and_unowned_cookie_semantics(self):
        headers = self.headers_for(self.user_one)
        hit = self.client.get("/api/orders", params={"search": "快照标题"}, headers=headers)
        self.assertEqual(hit.json()["total"], 1)
        miss = self.client.get("/api/orders", params={"search": "不存在的词"}, headers=headers)
        self.assertEqual(miss.json()["total"], 0)
        # 未授权账号：404，而不是静默回退到全账号
        foreign = self.client.get("/api/orders", params={"cookie_id": "acct-two"}, headers=headers)
        self.assertEqual(foreign.status_code, 404)

    def test_list_exposes_independent_profile_field_sources(self):
        self.db.upsert_customer_observation(
            "acct-two", "buyer-2", "档案昵称", "", "order_detail", 1000.0,
        )
        self.db.upsert_customer_observation(
            "acct-two", "buyer-2", "", "https://a/catalog.jpg", "catalog", 2000.0,
        )
        response = self.client.get(
            "/api/orders", headers=self.headers_for(self.user_two),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["data"][0]
        self.assertEqual(row["buyer_identity"], "profile")
        self.assertEqual(row["buyer_display_name"], "档案昵称")
        self.assertEqual(row["buyer_display_name_source"], "order_detail")
        self.assertEqual(row["buyer_avatar_url"], "https://a/catalog.jpg")
        self.assertEqual(row["buyer_avatar_source"], "catalog")

    def test_list_exposes_independent_order_buyer_field_sources(self):
        self.db.apply_order_sync_update(
            order_id="order-1",
            cookie_id="acct-one",
            incoming_status="pending_ship",
            buyer_snapshot={
                "buyer_nickname": "详情昵称",
                "source": "order_detail",
            },
        )
        self.db.apply_order_sync_update(
            order_id="order-1",
            cookie_id="acct-one",
            incoming_status="pending_ship",
            buyer_snapshot={
                "buyer_avatar_url": "https://a/realtime.jpg",
                "source": "realtime_message",
            },
        )

        response = self.client.get(
            "/api/orders", headers=self.headers_for(self.user_one),
        )
        self.assertEqual(response.status_code, 200)
        row = response.json()["data"][0]
        self.assertEqual(row["buyer_display_name"], "详情昵称")
        self.assertEqual(row["buyer_display_name_source"], "order_detail")
        self.assertEqual(row["buyer_avatar_url"], "https://a/realtime.jpg")
        self.assertEqual(row["buyer_avatar_source"], "realtime_message")

        self.db.apply_order_sync_update(
            order_id="order-1",
            cookie_id="acct-one",
            incoming_status="pending_ship",
            buyer_snapshot={
                "buyer_avatar_url": "https://a/detail.jpg",
                "source": "order_detail",
            },
        )
        upgraded = self.client.get(
            "/api/orders", headers=self.headers_for(self.user_one),
        ).json()["data"][0]
        self.assertEqual(upgraded["buyer_avatar_url"], "https://a/detail.jpg")
        self.assertEqual(upgraded["buyer_avatar_source"], "order_detail")

    def test_import_cannot_take_over_existing_order_from_another_cookie(self):
        before = self.db.get_order_by_id("order-1")
        response = self.client.post(
            "/api/orders/import",
            headers=self.headers_for(self.user_two),
            json=[{
                "order_id": "order-1",
                "cookie_id": "acct-two",
                "item_id": "item-takeover",
                "buyer_id": "buyer-takeover",
                "amount": "999.00",
                "order_status": "completed",
                "receiver_name": "越权收件人",
            }],
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["success_count"], 0)
        self.assertEqual(payload["failed_count"], 1)
        self.assertFalse(payload["results"][0]["success"])
        # 端点层显式归属校验：给出明确拒绝语义，而不是依赖 DB 层守卫的通用失败
        self.assertEqual(payload["results"][0]["message"], "无权操作此订单")
        self.assertEqual(self.db.get_order_by_id("order-1"), before)

    def test_import_cannot_claim_orphan_order(self):
        # 历史孤儿订单（cookie_id=NULL）不允许任何用户通过导入认领
        with self.db.lock:
            self.db.conn.execute(
                "INSERT INTO orders (order_id, item_id, order_status) VALUES ('orphan-1', 'item-x', 'unknown')"
            )
            self.db.conn.commit()
        response = self.client.post(
            "/api/orders/import",
            headers=self.headers_for(self.user_one),
            json=[{
                "order_id": "orphan-1",
                "cookie_id": "acct-one",
                "amount": "1.00",
            }],
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["failed_count"], 1)
        self.assertEqual(payload["results"][0]["message"], "无权操作此订单")
        orphan = self.db.get_order_by_id("orphan-1")
        self.assertIsNone(orphan.get("cookie_id"))
        self.assertNotEqual(orphan.get("amount"), "1.00")

    def test_detail_returns_receiver_fields_and_enforces_ownership(self):
        headers = self.headers_for(self.user_one)
        detail = self.client.get("/api/orders/order-1", headers=headers)
        self.assertEqual(detail.status_code, 200)
        data = detail.json()["data"]
        self.assertEqual(data["receiver_name"], "张三")
        self.assertEqual(data["receiver_phone"], "13800000000")
        self.assertEqual(data["buyer_display_name"], "买家甲")
        foreign = self.client.get("/api/orders/order-2", headers=headers)
        self.assertEqual(foreign.status_code, 404)

    def test_item_image_not_saved_when_no_source(self):
        # order-2 无快照且目录无此商品
        response = self.client.get(
            "/api/orders/order-2/item-image", headers=self.headers_for(self.user_two)
        )
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"]["reason"], "not_saved")

    def _attach_snapshot_image(self, image_url="https://img.alicdn.com/item-1.png"):
        self.db.apply_order_sync_update(
            order_id="order-1", cookie_id="acct-one", incoming_status="pending_ship",
            status_source="order_detail",
            item_snapshot={"item_image": image_url,
                           "source": "order_detail"},
        )

    def test_item_image_source_expired_and_unsupported_format(self):
        self._attach_snapshot_image()
        headers = self.headers_for(self.user_one)
        with _public_image_network(status=404):
            expired = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(expired.status_code, 404)
        self.assertEqual(expired.json()["detail"]["reason"], "source_expired")

        with _public_image_network(error=OSError("network down")):
            network = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(network.status_code, 404)
        self.assertEqual(network.json()["detail"]["reason"], "source_expired")

        with _public_image_network(status=200, body=b"not an image / heic-like"):
            bad_format = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(bad_format.status_code, 422)
        self.assertEqual(bad_format.json()["detail"]["reason"], "unsupported_format")

    def test_item_image_success_caches_and_writes_cache_key(self):
        self._attach_snapshot_image()
        headers = self.headers_for(self.user_one)
        with _public_image_network(status=200, body=_png_bytes()):
            first = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.headers["content-type"], "image/jpeg")
        cached_files = os.listdir(self.media_dir)
        self.assertEqual(len(cached_files), 1)
        detail = self.db.get_order_by_id("order-1")
        self.assertEqual(detail["item_image_cache_key"], cached_files[0])
        # 二次访问直接走磁盘缓存：让替身网络层抛错也必须成功
        with _public_image_network(error=AssertionError("网络层不应被调用")):
            second = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(second.status_code, 200)

    def test_item_image_rejects_non_https_and_untrusted_hosts_before_network(self):
        headers = self.headers_for(self.user_one)
        for image_url in (
            "http://img.alicdn.com/plaintext.png",
            "https://127.0.0.1/internal.png",
            "https://assets.example.test/untrusted.png",
        ):
            with self.subTest(image_url=image_url):
                # 每个子用例重置成交快照，允许同来源测试值覆盖。
                with self.db.lock:
                    self.db.conn.execute(
                        "UPDATE orders SET item_image = ?, item_image_cache_key = ''"
                        " WHERE order_id = 'order-1'",
                        (image_url,),
                    )
                    self.db.conn.commit()
                calls = []
                with _fake_aiohttp_session(
                    status=200, body=_png_bytes(), calls=calls,
                ):
                    response = self.client.get(
                        "/api/orders/order-1/item-image", headers=headers,
                    )
                self.assertEqual(response.status_code, 404)
                self.assertEqual(response.json()["detail"]["reason"], "source_expired")
                self.assertEqual(calls, [], "不可信 URL 必须在发起请求前被拒绝")

    def test_item_image_rejects_private_dns_result(self):
        self._attach_snapshot_image()
        with patch(
            "reply_server._resolve_order_image_host",
            new=AsyncMock(side_effect=ValueError("private address")),
            create=True,
        ), _fake_aiohttp_session(status=200, body=_png_bytes(), calls=[]):
            response = self.client.get(
                "/api/orders/order-1/item-image", headers=self.headers_for(self.user_one),
            )
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"]["reason"], "source_expired")

    def test_item_image_rejects_non_image_content_type(self):
        self._attach_snapshot_image()
        with _public_image_network(
            status=200,
            body=_png_bytes(),
            headers={"Content-Type": "text/html; charset=utf-8"},
        ):
            response = self.client.get(
                "/api/orders/order-1/item-image", headers=self.headers_for(self.user_one),
            )
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.json()["detail"]["reason"], "unsupported_format")

    def test_item_image_validates_and_follows_trusted_redirect(self):
        self._attach_snapshot_image()
        calls = []
        responses = [
            {
                "status": 302,
                "headers": {"Location": "https://gw.alicdn.com/item-1-final.png"},
            },
            {
                "status": 200,
                "headers": {"Content-Type": "image/png"},
                "body": _png_bytes(),
            },
        ]
        with patch(
            "reply_server._resolve_order_image_host",
            new=AsyncMock(return_value=("203.0.113.10",)),
            create=True,
        ) as resolve_mock, _fake_aiohttp_sequence(responses, calls):
            response = self.client.get(
                "/api/orders/order-1/item-image", headers=self.headers_for(self.user_one),
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            calls,
            [
                "https://img.alicdn.com/item-1.png",
                "https://gw.alicdn.com/item-1-final.png",
            ],
        )
        self.assertEqual(resolve_mock.await_count, 2)

    def test_item_image_rejects_private_dns_after_redirect(self):
        self._attach_snapshot_image()
        calls = []
        dns_hosts = []
        responses = [{
            "status": 302,
            "headers": {"Location": "https://gw.alicdn.com/private.png"},
        }]

        async def fake_getaddrinfo(host, port, **_kwargs):
            dns_hosts.append(host)
            address = (
                "93.184.216.34"
                if host == "img.alicdn.com"
                else "10.0.0.8"
            )
            return [
                (
                    socket.AF_INET,
                    socket.SOCK_STREAM,
                    socket.IPPROTO_TCP,
                    "",
                    (address, port),
                )
            ]

        with patch.object(
            asyncio.BaseEventLoop,
            "getaddrinfo",
            new=AsyncMock(side_effect=fake_getaddrinfo),
        ), _fake_aiohttp_sequence(responses, calls):
            response = self.client.get(
                "/api/orders/order-1/item-image",
                headers=self.headers_for(self.user_one),
            )
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"]["reason"], "source_expired")
        self.assertEqual(calls, ["https://img.alicdn.com/item-1.png"])
        self.assertEqual(dns_hosts, ["img.alicdn.com", "gw.alicdn.com"])

    def test_item_image_enforces_ten_second_total_timeout_and_maps_expiry(self):
        self._attach_snapshot_image()
        session_kwargs = []
        with _public_image_network(
            error=asyncio.TimeoutError(),
            session_kwargs=session_kwargs,
        ):
            response = self.client.get(
                "/api/orders/order-1/item-image",
                headers=self.headers_for(self.user_one),
            )
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"]["reason"], "source_expired")
        self.assertEqual(len(session_kwargs), 1)
        timeout = session_kwargs[0]["timeout"]
        self.assertIsInstance(timeout, aiohttp.ClientTimeout)
        self.assertEqual(timeout.total, 10)

    def test_item_image_rejects_declared_and_streamed_oversize_payloads(self):
        self._attach_snapshot_image()
        headers = self.headers_for(self.user_one)
        with _public_image_network(
            status=200,
            body=_png_bytes(),
            headers={
                "Content-Type": "image/png",
                "Content-Length": str(reply_server.ORDER_ITEM_IMAGE_MAX_BYTES + 1),
            },
        ):
            declared = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(declared.status_code, 422)
        self.assertEqual(declared.json()["detail"]["reason"], "unsupported_format")

        with _public_image_network(
            status=200,
            body=b"x" * (reply_server.ORDER_ITEM_IMAGE_MAX_BYTES + 1),
            headers={"Content-Type": "image/png"},
        ):
            streamed = self.client.get("/api/orders/order-1/item-image", headers=headers)
        self.assertEqual(streamed.status_code, 422)
        self.assertEqual(streamed.json()["detail"]["reason"], "unsupported_format")

    def test_image_connector_pins_the_validated_public_address(self):
        async def exercise():
            with patch(
                "reply_server._resolve_order_image_host",
                new=AsyncMock(return_value=("93.184.216.34",)),
            ), patch("aiohttp.TCPConnector") as connector:
                await reply_server._build_pinned_order_image_connector(
                    "https://img.alicdn.com/item.jpg"
                )
                resolver = connector.call_args.kwargs["resolver"]
                records = await resolver.resolve("img.alicdn.com", 443)
                self.assertEqual(records[0]["host"], "93.184.216.34")
                self.assertEqual(records[0]["hostname"], "img.alicdn.com")

        asyncio.run(exercise())

    def test_item_image_cache_publish_is_atomic(self):
        self._attach_snapshot_image()
        with patch("reply_server.os.replace", wraps=os.replace) as replace_mock:
            with _public_image_network(status=200, body=_png_bytes()):
                response = self.client.get(
                    "/api/orders/order-1/item-image", headers=self.headers_for(self.user_one),
                )
        self.assertEqual(response.status_code, 200)
        replace_mock.assert_called_once()

    def test_item_image_requires_ownership(self):
        response = self.client.get(
            "/api/orders/order-1/item-image", headers=self.headers_for(self.user_two)
        )
        self.assertEqual(response.status_code, 404)

    def test_programmatic_import_accepts_json_order_array_contract(self):
        headers = self.headers_for(self.user_one)
        response = self.client.post(
            "/api/orders/import",
            json=[
                {"order_id": "imported-1", "cookie_id": "acct-one",
                 "item_id": "item-9", "status": "pending_ship", "amount": "5.00",
                 "unknown_field": "会被忽略"},
                {"order_id": "imported-2", "cookie_id": "acct-two"},
            ],
            headers=headers,
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["success_count"], 1)
        self.assertEqual(payload["failed_count"], 1)
        self.assertIn("已忽略字段", payload["results"][0]["message"])
        self.assertIn("无权操作", payload["results"][1]["message"])

    def test_spreadsheet_import_accepts_real_xlsx_upload_contract(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["order_id", "cookie_id", "item_id", "status", "amount"])
        sheet.append(["xlsx-1", "acct-one", "item-x", "pending_ship", "8.50"])
        sheet.append(["", "acct-one", "empty-order", "pending_ship", "1.00"])
        payload = io.BytesIO()
        workbook.save(payload)

        response = self.client.post(
            "/api/orders/import",
            files={
                "file": (
                    "orders.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=self.headers_for(self.user_one),
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["success_count"], 1)
        self.assertEqual(body["failed_count"], 1)
        self.assertEqual(self.db.get_order_by_id("xlsx-1")["amount"], "8.50")

    def test_import_rejects_unsafe_excel_shapes(self):
        headers = self.headers_for(self.user_one)
        empty = self.client.post(
            "/api/orders/import",
            files={"file": ("orders.xlsx", b"", "application/octet-stream")},
            headers=headers,
        )
        wrong_extension = self.client.post(
            "/api/orders/import",
            files={"file": ("orders.csv", b"order_id,cookie_id", "text/csv")},
            headers=headers,
        )
        oversized = self.client.post(
            "/api/orders/import",
            files={"file": (
                "orders.xlsx",
                b"x" * (reply_server._ORDER_IMPORT_MAX_BYTES + 1),
                "application/octet-stream",
            )},
            headers=headers,
        )
        workbook = Workbook()
        workbook.active.append(["item_id", "amount"])
        workbook.active.append(["item-x", "1.00"])
        missing_headers_payload = io.BytesIO()
        workbook.save(missing_headers_payload)
        missing_headers = self.client.post(
            "/api/orders/import",
            files={"file": (
                "orders.xlsx",
                missing_headers_payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=headers,
        )
        self.assertEqual(
            [empty.status_code, wrong_extension.status_code, oversized.status_code,
             missing_headers.status_code],
            [400, 415, 413, 400],
        )
        self.assertEqual(wrong_extension.json()["detail"], "仅支持 .xlsx 文件")
        self.assertEqual(oversized.json()["detail"], "Excel 文件超过 5MB")

    def test_import_rejects_legacy_xls_with_clear_contract_error(self):
        response = self.client.post(
            "/api/orders/import",
            files={"file": (
                "orders.xls",
                b"legacy-binary-excel",
                "application/vnd.ms-excel",
            )},
            headers=self.headers_for(self.user_one),
        )
        self.assertEqual(response.status_code, 415)
        self.assertEqual(response.json()["detail"], "仅支持 .xlsx 文件")

    def test_import_rejects_xlsx_over_row_and_column_limits(self):
        headers = self.headers_for(self.user_one)

        too_many_rows = Workbook(write_only=True)
        row_sheet = too_many_rows.create_sheet()
        row_sheet.append(["order_id", "cookie_id"])
        for index in range(reply_server._ORDER_IMPORT_MAX_ROWS + 1):
            row_sheet.append([f"row-{index}", "acct-one"])
        row_payload = io.BytesIO()
        too_many_rows.save(row_payload)
        row_response = self.client.post(
            "/api/orders/import",
            files={"file": (
                "too-many-rows.xlsx",
                row_payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=headers,
        )

        too_many_columns = Workbook(write_only=True)
        column_sheet = too_many_columns.create_sheet()
        column_sheet.append(
            ["order_id", "cookie_id"]
            + [f"extra_{index}" for index in range(reply_server._ORDER_IMPORT_MAX_COLUMNS - 1)]
        )
        column_sheet.append(
            ["column-order", "acct-one"]
            + ["x"] * (reply_server._ORDER_IMPORT_MAX_COLUMNS - 1)
        )
        column_payload = io.BytesIO()
        too_many_columns.save(column_payload)
        column_response = self.client.post(
            "/api/orders/import",
            files={"file": (
                "too-many-columns.xlsx",
                column_payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=headers,
        )

        self.assertEqual(row_response.status_code, 413)
        self.assertEqual(column_response.status_code, 413)
        self.assertEqual(row_response.json()["detail"], "Excel 行列数超过限制")
        self.assertEqual(column_response.json()["detail"], "Excel 行列数超过限制")

    def test_unowned_order_mutations_do_not_reveal_existence(self):
        headers = self.headers_for(self.user_two)
        foreign_requests = (
            self.client.get("/api/orders/order-1", headers=headers),
            self.client.get("/api/orders/order-1/item-image", headers=headers),
            self.client.delete("/api/orders/order-1", headers=headers),
            self.client.post("/api/orders/order-1/refresh", headers=headers),
            self.client.put("/api/orders/order-1", json={"amount": "1"}, headers=headers),
        )
        missing_requests = (
            self.client.get("/api/orders/missing-order", headers=headers),
            self.client.get("/api/orders/missing-order/item-image", headers=headers),
            self.client.delete("/api/orders/missing-order", headers=headers),
            self.client.post("/api/orders/missing-order/refresh", headers=headers),
            self.client.put("/api/orders/missing-order", json={"amount": "1"}, headers=headers),
        )
        self.assertEqual([response.status_code for response in foreign_requests], [404] * 5)
        self.assertEqual(
            [response.json()["detail"] for response in foreign_requests],
            [response.json()["detail"] for response in missing_requests],
        )

        manual = self.client.post(
            "/api/orders/manual-ship",
            json={"order_ids": ["order-1"], "ship_mode": "status_only"},
            headers=headers,
        )
        self.assertEqual(manual.status_code, 200)
        self.assertEqual(manual.json()["results"][0]["message"], "订单不存在或无权访问")


if __name__ == "__main__":
    unittest.main()
